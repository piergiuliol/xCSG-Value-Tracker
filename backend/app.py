"""
app.py — FastAPI routes for xCSG Value Tracker
Phase 1 realignment (April 2026).

IMPORTANT: app.mount("/", StaticFiles(...)) MUST be the LAST line.
"""
import csv
import io
import json
import os
import tempfile
from pathlib import Path
from typing import List, Optional

from fastapi import Depends, FastAPI, HTTPException, Query, Response, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from backend import auth
from backend import database as db
from backend import metrics as mtx
from backend import pioneers as pioneers_mod
from backend.pioneers import PioneerInUseError
from backend.models import (
    ActivityLogEntry,
    AppSettings,
    AppSettingsUpdate,
    CategoryCreate,
    CategoryUpdate,
    ExpertAssessmentMetrics,
    ExpertContextResponse,
    ExpertResponseCreate,
    LoginRequest,
    LoginResponse,
    MetricsSummary,
    ProjectPioneerEntry,
    ProjectPioneerUpdate,
    PioneerCreate,
    PioneerUpdate,
    PioneerSummary,
    PracticeCreate,
    PracticeUpdate,
    PracticeRoleEntry,
    PracticeRolesUpdate,
    ProjectCreate,
    ProjectUpdate,
    RegisterRequest,
    ScalingGates,
    TrendData,
    UserInfo,
)
from backend.schema import (
    EXPERT_FIELDS,
    MAX_PIONEERS_PER_PROJECT,
    MAX_ROUNDS_PER_PIONEER,
    build_schema_response,
    missing_required_fields,
)


# ── Expert field options (derived from schema.py) ────────────────────────────

EXPERT_FIELD_OPTIONS = {
    key: {
        "key": key,
        "label": defn["label"],
        "section": defn["section"],
        "options": defn.get("options", []),
        "type": defn.get("type", "categorical"),
        "has_legacy": False,
    }
    for key, defn in EXPERT_FIELDS.items()
}


def _normalize_project_payload(data: dict) -> dict:
    normalized = dict(data)
    if normalized.get("working_days") is not None and normalized.get("xcsg_calendar_days") is None:
        normalized["xcsg_calendar_days"] = str(normalized["working_days"])
    if normalized.get("revision_depth") and normalized.get("xcsg_revision_rounds") is None:
        normalized["xcsg_revision_rounds"] = normalized["revision_depth"]
    # Auto-compute working_days from dates if not provided.
    # Prefer actual delivery; fall back to expected delivery when actual is blank.
    end_date = normalized.get("date_delivered") or normalized.get("date_expected_delivered")
    if normalized.get("working_days") is None and normalized.get("date_started") and end_date:
        try:
            from datetime import date as _date
            s = _date.fromisoformat(normalized["date_started"])
            e = _date.fromisoformat(end_date)
            # Business days approximation: total days * 5/7
            total = (e - s).days
            if total >= 0:
                normalized["working_days"] = max(1, round(total * 5 / 7))
        except (ValueError, TypeError):
            pass
    return normalized

# ── App init ──────────────────────────────────────────────────────────────────

app = FastAPI(title="xCSG Value Tracker", version="2.1.0")

# CORS
_raw_origins = os.environ.get(
    "CORS_ORIGINS",
    '["http://localhost:3000", "http://localhost:8765", "http://127.0.0.1:8765"]',
)
try:
    CORS_ORIGINS = json.loads(_raw_origins)
except Exception:
    CORS_ORIGINS = ["http://localhost:8765"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_event():
    # init_db() owns all migrations — don't duplicate here.
    db.init_db()


# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "2.1.0"}


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/api/auth/login", response_model=LoginResponse)
async def login(body: LoginRequest):
    user_row = db.get_user_by_username(body.username)
    if not user_row or not auth.verify_password(body.password, user_row["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password",
        )
    token = auth.create_token(user_row["id"], user_row["username"], user_row["role"])
    db.log_activity(user_row["id"], "login", details=f"User {user_row['username']} logged in")

    return LoginResponse(
        access_token=token,
        expires_in=auth.JWT_EXPIRY_HOURS * 3600,
        user=UserInfo(
            id=user_row["id"],
            username=user_row["username"],
            name=user_row["username"].title(),
            role=user_row["role"],
        ),
    )


@app.post("/api/auth/register", status_code=201)
async def register(body: RegisterRequest, current_user: dict = Depends(auth.get_current_user_admin)):
    existing = db.get_user_by_username(body.username)
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    hashed = auth.hash_password(body.password)
    user_id = db.create_user(body.username, body.email, hashed, body.role)
    return {"id": user_id, "username": body.username, "role": body.role}


@app.get("/api/users")
async def list_users(current_user: dict = Depends(auth.get_current_user_admin)):
    return db.list_users()


@app.put("/api/users/{user_id}")
async def update_user_endpoint(
    user_id: int,
    body: dict,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    allowed = {}
    if "role" in body and body["role"] in ("admin", "analyst", "viewer"):
        allowed["role"] = body["role"]
    if "email" in body:
        allowed["email"] = body["email"]
    if "password" in body and body["password"]:
        allowed["password_hash"] = auth.hash_password(body["password"])
    db.update_user(user_id, allowed)
    return {"ok": True}


@app.delete("/api/users/{user_id}", status_code=204)
async def delete_user_endpoint(
    user_id: int,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    if int(current_user["sub"]) == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete_user(user_id)


@app.put("/api/auth/password")
async def change_own_password(
    body: dict,
    current_user: dict = Depends(auth.get_current_user),
):
    user = db.get_user_by_id(int(current_user["sub"]))
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not auth.verify_password(body.get("current_password", ""), user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    new_pw = body.get("new_password", "")
    if len(new_pw) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    db.update_user(int(current_user["sub"]), {"password_hash": auth.hash_password(new_pw)})
    return {"ok": True}


# ── Project Categories ───────────────────────────────────────────────────────

@app.get("/api/categories")
async def list_categories(current_user: dict = Depends(auth.get_current_user)):
    cats = db.list_categories()
    conn = db.get_connection()
    try:
        counts = {}
        for row in conn.execute("SELECT category_id, COUNT(*) as cnt FROM projects GROUP BY category_id").fetchall():
            counts[row["category_id"]] = row["cnt"]
    finally:
        conn.close()
    for c in cats:
        c["project_count"] = counts.get(c["id"], 0)
    return cats


@app.post("/api/categories", status_code=201)
async def create_category(
    body: CategoryCreate,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    try:
        cat_id = db.create_category(body.name, body.description)
    except Exception:
        raise HTTPException(status_code=400, detail="Category name already exists")
    cat = db.get_category(cat_id)
    db.log_activity(
        current_user["sub"],
        "category_created",
        details=f"Created category '{body.name}'",
    )
    return dict(cat)


@app.put("/api/categories/{category_id}")
async def update_category(
    category_id: int,
    body: CategoryUpdate,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    cat = db.get_category(category_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    try:
        db.update_category(category_id, body.name, body.description)
    except Exception:
        raise HTTPException(status_code=400, detail="Category name already exists")
    db.log_activity(
        current_user["sub"],
        "category_updated",
        details=f"Updated category '{body.name}'",
    )
    return dict(db.get_category(category_id))


@app.put("/api/categories/{category_id}/practices")
async def set_category_practices_endpoint(
    category_id: int,
    body: dict,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    """Replace the practice attributions for a category. Body: {"practice_ids": [1, 2, ...]}."""
    if not db.get_category(category_id):
        raise HTTPException(status_code=404, detail="Category not found")
    practice_ids = body.get("practice_ids", [])
    if not isinstance(practice_ids, list):
        raise HTTPException(status_code=400, detail="practice_ids must be a list")
    for pid in practice_ids:
        if not db.get_practice(int(pid)):
            raise HTTPException(status_code=400, detail=f"Invalid practice id: {pid}")
    db.set_practices_for_category(category_id, practice_ids)
    db.log_activity(
        current_user["sub"],
        "category_practices_updated",
        details=f"Updated practices for category #{category_id}: {practice_ids}",
    )
    # Return the fresh category (with practices list).
    cats = db.list_categories()
    return next((c for c in cats if c["id"] == category_id), {})


@app.delete("/api/categories/{category_id}", status_code=204)
async def delete_category(
    category_id: int,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    cat = db.get_category(category_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    if db.category_has_projects(category_id):
        raise HTTPException(status_code=400, detail="Cannot delete category with existing projects")
    db.delete_category(category_id)
    db.log_activity(
        current_user["sub"],
        "category_deleted",
        details=f"Deleted category '{cat['name']}'",
    )


# ── Practices ────────────────────────────────────────────────────────────────

@app.get("/api/practices")
async def list_practices_endpoint(current_user: dict = Depends(auth.get_current_user)):
    return db.list_practices()


@app.post("/api/practices", status_code=201)
async def create_practice_endpoint(
    body: PracticeCreate,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    try:
        practice_id = db.create_practice(body.code, body.name, body.description)
    except Exception:
        raise HTTPException(status_code=400, detail="Practice code already exists")
    practice = db.get_practice(practice_id)
    db.log_activity(
        current_user["sub"],
        "practice_created",
        details=f"Created practice '{body.code}'",
    )
    return dict(practice)


@app.put("/api/practices/{practice_id}")
async def update_practice_endpoint(
    practice_id: int,
    body: PracticeUpdate,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    practice = db.get_practice(practice_id)
    if not practice:
        raise HTTPException(status_code=404, detail="Practice not found")
    db.update_practice(practice_id, body.name, body.description)
    db.log_activity(
        current_user["sub"],
        "practice_updated",
        details=f"Updated practice '{body.name}'",
    )
    return dict(db.get_practice(practice_id))


@app.get("/api/practices/{practice_id}/roles")
def get_practice_roles(practice_id: int, user=Depends(auth.get_current_user)):
    practice = db.get_practice(practice_id)
    if not practice:
        raise HTTPException(status_code=404, detail="Practice not found")
    return db.list_practice_roles(practice_id)


@app.put("/api/practices/{practice_id}/roles")
def update_practice_roles(
    practice_id: int,
    payload: PracticeRolesUpdate,
    user=Depends(auth.get_current_user_admin),
):
    practice = db.get_practice(practice_id)
    if not practice:
        raise HTTPException(status_code=404, detail="Practice not found")
    db.replace_practice_roles(
        practice_id,
        [r.model_dump() for r in payload.roles],
    )
    return db.list_practice_roles(practice_id)


@app.delete("/api/practices/{practice_id}", status_code=204)
async def delete_practice_endpoint(
    practice_id: int,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    practice = db.get_practice(practice_id)
    if not practice:
        raise HTTPException(status_code=404, detail="Practice not found")
    ok = db.delete_practice(practice_id)
    if not ok:
        raise HTTPException(status_code=400, detail="Cannot delete practice with existing projects")
    db.log_activity(
        current_user["sub"],
        "practice_deleted",
        details=f"Deleted practice '{practice['code']}'",
    )


# ── Projects ─────────────────────────────────────────────────────────────────

@app.get("/api/projects")
async def list_projects(
    status_filter: Optional[str] = Query(None, alias="status"),
    category_id: Optional[int] = Query(None),
    practice_id: Optional[int] = Query(None),
    pioneer: Optional[str] = Query(None),
    client: Optional[str] = Query(None),
    current_user: dict = Depends(auth.get_current_user),
):
    projects = db.list_projects(
        status_filter=status_filter,
        category_id=category_id,
        practice_id=practice_id,
        pioneer=pioneer,
        client=client,
    )

    enriched_projects = []
    for project in projects:
        result = dict(project)
        result["pioneers"] = db.list_pioneers(project["id"])
        responses = db.get_all_project_responses(project["id"])
        if responses:
            project_dict = dict(project)
            project_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(project["id"])
            project_dict["legacy_team"] = db.list_legacy_team(project["id"])
            result["metrics"] = mtx.compute_averaged_project_metrics(project_dict, responses)
            result["response_count"] = len(responses)
        else:
            result["metrics"] = None
            result["response_count"] = 0
        enriched_projects.append(result)

    return enriched_projects


@app.post("/api/projects", status_code=201)
async def create_project(
    body: ProjectCreate,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    cat = db.get_category(body.category_id)
    if not cat:
        raise HTTPException(status_code=400, detail="Invalid category_id")

    if body.practice_id is not None:
        if not db.get_practice(body.practice_id):
            raise HTTPException(status_code=400, detail="Invalid practice_id")
        if not db.is_practice_allowed_for_category(body.category_id, body.practice_id):
            allowed = db.get_practices_for_category(body.category_id)
            codes = [a["code"] for a in allowed]
            raise HTTPException(
                status_code=400,
                detail=f"Practice is not allowed for this category. Allowed: {codes}",
            )

    if len(body.pioneers) > MAX_PIONEERS_PER_PROJECT:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_PIONEERS_PER_PROJECT} pioneers per project")
    if body.default_rounds > MAX_ROUNDS_PER_PIONEER:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_ROUNDS_PER_PIONEER} rounds per pioneer")
    for p in body.pioneers:
        if p.total_rounds is not None and p.total_rounds > MAX_ROUNDS_PER_PIONEER:
            raise HTTPException(status_code=400, detail=f"Maximum {MAX_ROUNDS_PER_PIONEER} rounds per pioneer")

    data = _normalize_project_payload(body.model_dump())
    data["created_by"] = current_user["sub"]
    data["pioneers"] = [
        {
            "pioneer_id": p.pioneer_id,
            "first_name": p.first_name,
            "last_name": p.last_name,
            "email": p.email,
            "total_rounds": p.total_rounds,
            "day_rate": p.day_rate,
            "role_name": p.role_name,
        }
        for p in body.pioneers
    ]

    norm = db.get_norm_by_category(body.category_id)
    if norm:
        data["legacy_overridden"] = (
            data.get("legacy_calendar_days") != norm["typical_calendar_days"]
            or data.get("legacy_revision_rounds") != norm["typical_revision_rounds"]
        )
    else:
        data["legacy_overridden"] = False

    project_id = db.create_project(data)
    row = db.get_project(project_id)
    result = dict(row)
    result["pioneers"] = db.list_pioneers(project_id)
    db.log_activity(
        current_user["sub"],
        "project_created",
        project_id=project_id,
        details=f"Created project '{data['project_name']}' ({cat['name']})",
    )
    return result


@app.get("/api/projects/{project_id}")
async def get_project(
    project_id: int,
    current_user: dict = Depends(auth.get_current_user),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    result = dict(row)
    result["pioneers"] = db.list_pioneers(project_id)
    result["legacy_team"] = db.list_legacy_team(project_id)
    responses = db.get_all_project_responses(project_id)
    if responses:
        project_dict = dict(row)
        project_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(project_id)
        project_dict["legacy_team"] = result["legacy_team"]
        result["metrics"] = mtx.compute_averaged_project_metrics(project_dict, responses)
        result["response_count"] = len(responses)
    else:
        result["metrics"] = None
        result["response_count"] = 0
    return result


@app.put("/api/projects/{project_id}")
async def update_project(
    project_id: int,
    body: ProjectUpdate,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    data = {k: v for k, v in _normalize_project_payload(body.model_dump()).items() if v is not None}
    db.update_project(project_id, data)
    db.log_activity(
        current_user["sub"],
        "project_updated",
        project_id=project_id,
        details=f"Updated project #{project_id}",
    )
    updated = db.get_project(project_id)
    result = dict(updated)
    result["pioneers"] = db.list_pioneers(project_id)
    result["legacy_team"] = db.list_legacy_team(project_id)
    return result


@app.delete("/api/projects/{project_id}", status_code=204)
async def delete_project(
    project_id: int,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    db.log_activity(
        current_user["sub"],
        "project_deleted",
        details=f"Deleted project #{project_id} ({row['project_name']})",
    )
    db.delete_project(project_id)


@app.patch("/api/deliverables/{project_id}")
async def patch_deliverable(
    project_id: int,
    body: ProjectUpdate,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    if body.client_pulse is None:
        raise HTTPException(status_code=400, detail="client_pulse is required")
    db.update_project_client_pulse(project_id, body.client_pulse)
    return dict(db.get_project(project_id))


# ── Pioneer Management ──────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/pioneers")
async def list_project_pioneers(
    project_id: int,
    current_user: dict = Depends(auth.get_current_user),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    return db.list_pioneers(project_id)


@app.post("/api/projects/{project_id}/pioneers", status_code=201)
async def add_project_pioneer(
    project_id: int,
    body: ProjectPioneerEntry,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    current_count = len(db.list_pioneers(project_id))
    if current_count >= MAX_PIONEERS_PER_PROJECT:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_PIONEERS_PER_PROJECT} pioneers per project")
    pioneer_id = db.add_pioneer(
        project_id=project_id,
        pioneer_id=body.pioneer_id,
        first_name=body.first_name,
        last_name=body.last_name,
        email=body.email,
        total_rounds=body.total_rounds,
        issued_by=current_user.get("sub"),
        day_rate=body.day_rate,
        role_name=body.role_name,
    )
    pioneers = db.list_pioneers(project_id)
    new_pioneer = next((p for p in pioneers if p["id"] == pioneer_id), None)
    label = (
        ((body.first_name or "").strip() + " " + (body.last_name or "").strip()).strip()
        or (new_pioneer.get("display_name") if new_pioneer else "")
    )
    db.log_activity(
        current_user["sub"],
        "pioneer_added",
        project_id=project_id,
        details=f"Added pioneer '{label}' to project #{project_id}",
    )
    return new_pioneer


@app.put("/api/projects/{project_id}/pioneers/{pioneer_id}")
async def update_project_pioneer(
    project_id: int,
    pioneer_id: int,
    body: ProjectPioneerUpdate,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    # exclude_unset so omitted fields don't appear at all; role_name=null is a
    # legitimate intentional clear, so we keep it even when None.
    data = body.model_dump(exclude_unset=True)
    # Strip None from fields where None means "leave unchanged" (all except role_name).
    data = {k: v for k, v in data.items() if v is not None or k == "role_name"}
    if not data:
        raise HTTPException(status_code=400, detail="No fields to update")
    try:
        db.update_pioneer(pioneer_id, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    pioneers = db.list_pioneers(project_id)
    updated = next((p for p in pioneers if p["id"] == pioneer_id), None)
    if not updated:
        raise HTTPException(status_code=404, detail="Pioneer not found")
    return updated


@app.delete("/api/projects/{project_id}/pioneers/{pioneer_id}", status_code=204)
async def delete_project_pioneer(
    project_id: int,
    pioneer_id: int,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    success = db.remove_pioneer(pioneer_id)
    if not success:
        raise HTTPException(status_code=400, detail="Cannot remove pioneer with existing responses")
    db.log_activity(
        current_user["sub"],
        "pioneer_removed",
        project_id=project_id,
        details=f"Removed pioneer #{pioneer_id} from project #{project_id}",
    )


@app.post("/api/pioneers/{pioneer_id}/rounds/{round_number}/issue")
async def issue_pioneer_round(
    pioneer_id: int,
    round_number: int,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    """Issue (or re-issue) a round token for a pioneer. Returns the new token row."""
    try:
        row = db.issue_round_token(pioneer_id, round_number, issued_by=current_user.get("sub"))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    db.log_activity(
        current_user["sub"],
        "round_issued",
        details=f"Issued round {round_number} for pioneer #{pioneer_id}",
    )
    return row


@app.delete("/api/pioneers/{pioneer_id}/rounds/{round_number}", status_code=204)
async def cancel_pioneer_round(
    pioneer_id: int,
    round_number: int,
    current_user: dict = Depends(auth.get_current_user_analyst),
):
    """Cancel a pending round token. Fails if the round is already completed."""
    ok = db.cancel_round_token(pioneer_id, round_number)
    if not ok:
        raise HTTPException(status_code=400, detail="Cannot cancel: round is completed or not issued")
    db.log_activity(
        current_user["sub"],
        "round_cancelled",
        details=f"Cancelled round {round_number} for pioneer #{pioneer_id}",
    )


@app.get("/api/pioneers/{pioneer_id}/rounds/{round_number}")
async def get_pioneer_round(
    pioneer_id: int,
    round_number: int,
    current_user: dict = Depends(auth.get_current_user),
):
    """Return the completed response + per-round metrics for a pioneer round."""
    responses = db.get_pioneer_responses(pioneer_id)
    response = next((r for r in responses if r.get("round_number") == round_number), None)
    if not response:
        raise HTTPException(status_code=404, detail="Round has not been submitted")

    with db._db() as conn:
        pp = conn.execute(
            """SELECT pp.project_id,
                      pio.first_name AS pioneer_first_name,
                      pio.last_name  AS pioneer_last_name,
                      TRIM(pio.first_name || ' ' || pio.last_name) AS pioneer_name
               FROM project_pioneers pp
               JOIN pioneers pio ON pio.id = pp.pioneer_id
               WHERE pp.id = ?""",
            (pioneer_id,),
        ).fetchone()
    if not pp:
        raise HTTPException(status_code=404, detail="Pioneer not found")

    project_row = db.get_project(pp["project_id"])
    project_dict = dict(project_row)
    merged = dict(project_dict)
    merged.update(dict(response))
    merged["pioneer_day_rates"] = db.get_pioneer_day_rates(pp["project_id"])
    merged["legacy_team"] = db.list_legacy_team(pp["project_id"])
    metrics = mtx.compute_project_metrics(merged)
    # Include project-level economics fields so the frontend can render the
    # economics card (renderEconomicsCard needs project.engagement_revenue etc.)
    pioneers = db.list_pioneers(pp["project_id"])
    return {
        "round_number": round_number,
        "pioneer_id": pioneer_id,
        "pioneer_name": pp["pioneer_name"],
        "submitted_at": response.get("submitted_at"),
        "response": dict(response),
        "metrics": metrics,
        "project": {
            "id": project_dict.get("id"),
            "project_name": project_dict.get("project_name"),
            "engagement_revenue": project_dict.get("engagement_revenue"),
            "scope_expansion_revenue": project_dict.get("scope_expansion_revenue"),
            "xcsg_pricing_model": project_dict.get("xcsg_pricing_model"),
            "currency": project_dict.get("currency"),
            "pioneers": pioneers,
        },
    }


# ── Schema (NO auth — public) ────────────────────────────────────────────────

@app.get("/api/schema")
async def get_schema():
    """Return full schema: fields, scores, sections, metrics, norms columns."""
    return build_schema_response()


# ── Expert (NO auth — token-based) ───────────────────────────────────────────

@app.get("/api/expert/options")
async def get_expert_options():
    """Return all field definitions for building the expert assessment form."""
    return EXPERT_FIELD_OPTIONS


@app.get("/api/expert/{token}", response_model=ExpertContextResponse)
async def get_expert_context(token: str):
    tok = db.get_round_token(token)
    if not tok:
        raise HTTPException(status_code=404, detail="Expert link is invalid or has expired")

    pioneer_id = tok["pioneer_id"]
    project_id = tok["project_id"]
    current_round = tok["round_number"]
    total_rounds = tok.get("total_rounds") or tok.get("default_rounds") or 1
    show_previous = bool(tok.get("show_previous") if tok.get("show_previous") is not None else tok.get("show_previous_answers"))
    already_completed = tok.get("completed_at") is not None

    previous_responses = None
    if show_previous:
        previous_responses = db.get_pioneer_responses(pioneer_id)

    show_other_pioneers = bool(tok.get("show_other_pioneers_answers"))
    other_pioneers_responses = None
    if show_other_pioneers:
        other_pioneers_responses = db.get_other_pioneer_responses(project_id, pioneer_id)

    return ExpertContextResponse(
        project_id=project_id,
        project_name=tok["project_name"],
        category_name=tok["category_name"],
        practice_code=tok.get("practice_code"),
        practice_name=tok.get("practice_name"),
        description=tok.get("description"),
        client_name=tok.get("client_name"),
        pioneer_name=tok["pioneer_name"],
        date_started=tok.get("date_started"),
        date_delivered=tok.get("date_delivered"),
        xcsg_team_size=tok["xcsg_team_size"],
        xcsg_calendar_days=tok.get("xcsg_calendar_days"),
        engagement_stage=tok.get("engagement_stage"),
        already_completed=already_completed,
        pioneer_id=pioneer_id,
        current_round=current_round,
        total_rounds=total_rounds,
        show_previous=show_previous,
        previous_responses=previous_responses,
        show_other_pioneers=show_other_pioneers,
        other_pioneers_responses=other_pioneers_responses,
    )


@app.get("/api/expert/{token}/metrics", response_model=ExpertAssessmentMetrics)
async def get_expert_metrics(token: str):
    """Return computed flywheel leg scores for a submitted expert assessment."""
    tok = db.get_round_token(token)
    if not tok:
        raise HTTPException(status_code=404, detail="Expert link is invalid or has expired")
    project_id = tok["project_id"]

    project_row = db.get_project(project_id)
    if not project_row:
        raise HTTPException(status_code=404, detail="Project not found")

    responses = db.get_all_project_responses(project_id)
    if not responses:
        raise HTTPException(status_code=404, detail="Assessment not yet submitted")

    project_dict = dict(project_row)
    project_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(project_id)
    project_dict["legacy_team"] = db.list_legacy_team(project_id)
    metrics = mtx.compute_averaged_project_metrics(project_dict, responses)
    return ExpertAssessmentMetrics(
        machine_first_score=metrics.get("machine_first_score"),
        senior_led_score=metrics.get("senior_led_score"),
        proprietary_knowledge_score=metrics.get("proprietary_knowledge_score"),
        client_impact=metrics.get("client_impact"),
        data_independence=metrics.get("data_independence"),
        ai_survival_rate=metrics.get("ai_survival_rate"),
        reuse_intent_score=metrics.get("reuse_intent_score"),
    )


@app.post("/api/expert/{token}", status_code=201)
async def submit_expert_response(token: str, body: ExpertResponseCreate):
    tok = db.get_round_token(token)
    if not tok:
        raise HTTPException(status_code=404, detail="Expert link is invalid or has expired")

    if tok.get("completed_at") is not None:
        return {"already_completed": True, "message": "This round has already been submitted."}

    pioneer_id = tok["pioneer_id"]
    project_id = tok["project_id"]
    current_round = tok["round_number"]
    total_rounds = tok.get("total_rounds") or tok.get("default_rounds") or 1

    data = body.model_dump()
    missing = missing_required_fields(data)
    if missing:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Survey is incomplete. Please answer every required field before submitting.",
                "missing_fields": missing,
            },
        )

    response_id = db.create_expert_response(pioneer_id, project_id, current_round, data)
    db.complete_round_token(token, response_id)
    db.update_project_status(project_id)

    try:
        db.log_activity(
            1,
            "expert_submitted",
            project_id=project_id,
            details=f"Expert assessment submitted for '{tok['project_name']}' (pioneer: {tok['pioneer_name']}, round {current_round}/{total_rounds})",
        )
    except Exception as e:
        import logging
        logging.warning(f"log_activity failed (non-fatal): {e}")

    project_row = db.get_project(project_id)
    responses = db.get_all_project_responses(project_id)
    project_dict = dict(project_row)
    project_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(project_id)
    project_dict["legacy_team"] = db.list_legacy_team(project_id)
    metrics = mtx.compute_averaged_project_metrics(project_dict, responses)

    # Auto-issue the next round token if the pioneer has more rounds remaining
    # and the next round doesn't already have a token.
    next_round_token: Optional[str] = None
    next_round_number = current_round + 1
    if next_round_number <= total_rounds:
        with db._db() as conn:
            existing = conn.execute(
                "SELECT token, completed_at FROM pioneer_round_tokens WHERE pioneer_id = ? AND round_number = ?",
                (pioneer_id, next_round_number),
            ).fetchone()
        if existing is None:
            try:
                issued = db.issue_round_token(pioneer_id, next_round_number, issued_by=None)
                next_round_token = issued.get("token")
                try:
                    db.log_activity(
                        1,
                        "round_auto_issued",
                        project_id=project_id,
                        details=(
                            f"Auto-issued round {next_round_number} for pioneer "
                            f"#{pioneer_id} ({tok['pioneer_name']}) after round "
                            f"{current_round} submission"
                        ),
                    )
                except Exception as e:
                    import logging
                    logging.warning(f"log_activity failed (non-fatal): {e}")
            except ValueError:
                # Defensive: if validation rejects (e.g. race), fall through
                # with next_round_token = None.
                next_round_token = None
        elif existing["completed_at"] is None:
            # Already issued (e.g. admin pre-issued it) — surface the token
            # so the UI can pick it up.
            next_round_token = existing["token"]

    return {
        "success": True,
        "message": "Assessment submitted successfully",
        "metrics": metrics,
        "current_round": current_round,
        "total_rounds": total_rounds,
        "next_round_token": next_round_token,
    }


# ── Legacy Norms ──────────────────────────────────────────────────────────────

@app.get("/api/norms/aggregates")
async def list_norm_aggregates(current_user: dict = Depends(auth.get_current_user)):
    return db.list_norm_aggregates()


# ── Metrics ───────────────────────────────────────────────────────────────────

def _filter_project_ids_by_pioneer(pioneer_ids: Optional[List[int]]) -> Optional[set]:
    """Return the set of project ids that have any of the given pioneer_ids
    assigned. Returns None if pioneer_ids is None/empty (no filter).
    Returns an empty set when pioneer_ids has values but none match — the
    caller must treat this as 'filter all projects out'."""
    if not pioneer_ids:
        return None
    placeholders = ",".join("?" for _ in pioneer_ids)
    with db._db() as conn:
        rows = conn.execute(
            f"SELECT DISTINCT project_id FROM project_pioneers WHERE pioneer_id IN ({placeholders})",
            tuple(pioneer_ids),
        ).fetchall()
    return {r["project_id"] for r in rows}


@app.get("/api/dashboard/metrics")
async def dashboard_metrics(
    pioneer_id: Optional[List[int]] = Query(None),
    current_user: dict = Depends(auth.get_current_user),
):
    project_id_filter = _filter_project_ids_by_pioneer(pioneer_id)
    all_projects = db.list_projects()
    if project_id_filter is not None:
        all_projects = [p for p in all_projects if p["id"] in project_id_filter]
    complete = _build_averaged_complete_projects()
    if project_id_filter is not None:
        complete = [p for p in complete if p["id"] in project_id_filter]
    return mtx.compute_dashboard_metrics(complete, all_projects)


@app.get("/api/dashboard/takeaways")
async def dashboard_takeaways(
    pioneer_id: Optional[List[int]] = Query(None),
    current_user: dict = Depends(auth.get_current_user),
):
    """Return {chart_id: takeaway_string} for each dashboard chart."""
    from backend.takeaways import compute_takeaways
    from backend.schema import DASHBOARD_CONFIG

    project_id_filter = _filter_project_ids_by_pioneer(pioneer_id)
    complete = _build_averaged_complete_projects()
    all_p = db.list_projects()
    if project_id_filter is not None:
        complete = [p for p in complete if p["id"] in project_id_filter]
        all_p = [p for p in all_p if p["id"] in project_id_filter]
    aggregates = mtx.compute_summary(complete, all_p)
    scaling_gates = mtx.compute_scaling_gates(complete)
    return compute_takeaways(complete, aggregates, scaling_gates, DASHBOARD_CONFIG["charts"])


@app.get("/api/projects/{project_id}/metrics")
async def project_metrics(
    project_id: int,
    current_user: dict = Depends(auth.get_current_user),
):
    row = db.get_project(project_id)
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")

    responses = db.get_all_project_responses(project_id)
    project_dict = dict(row)
    project_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(project_id)
    project_dict["legacy_team"] = db.list_legacy_team(project_id)
    return mtx.compute_averaged_project_metrics(project_dict, responses)


def _build_averaged_complete_projects() -> list:
    """Build list of averaged metrics dicts for all projects with responses."""
    projects_with_responses = db.list_complete_projects()
    result = []
    for p in projects_with_responses:
        responses = db.get_all_project_responses(p["id"])
        if responses:
            project_dict = dict(p)
            project_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(p["id"])
            project_dict["legacy_team"] = db.list_legacy_team(p["id"])
            avg = mtx.compute_averaged_project_metrics(project_dict, responses)
            avg["id"] = p["id"]
            avg["project_name"] = p["project_name"]
            avg["category_name"] = p["category_name"]
            avg["practice_code"] = p.get("practice_code")
            avg["practice_name"] = p.get("practice_name")
            # Pioneer name(s) for charts/exports — fetch from list_pioneers and
            # join with comma. The projects table no longer carries a
            # pioneer_name column (dropped in v1.9), so we resolve via JOIN.
            project_pioneers = db.list_pioneers(p["id"])
            avg["pioneer_name"] = ", ".join(
                pp.get("display_name") or pp.get("pioneer_name", "")
                for pp in project_pioneers
            )
            avg["date_started"] = p.get("date_started")
            avg["date_delivered"] = p.get("date_delivered")
            result.append(avg)
    return result


@app.get("/api/metrics/summary", response_model=MetricsSummary)
async def metrics_summary(current_user: dict = Depends(auth.get_current_user)):
    complete = _build_averaged_complete_projects()
    all_p = db.list_projects()
    summary = mtx.compute_summary(complete, all_p)
    return MetricsSummary(**summary)


@app.get("/api/metrics/projects")
async def metrics_projects(current_user: dict = Depends(auth.get_current_user)):
    return _build_averaged_complete_projects()


@app.get("/api/metrics/trends", response_model=TrendData)
async def metrics_trends(current_user: dict = Depends(auth.get_current_user)):
    complete = _build_averaged_complete_projects()
    from backend.models import TrendPoint
    return TrendData(points=[TrendPoint(**p) for p in complete])


@app.get("/api/metrics/scaling-gates", response_model=ScalingGates)
async def metrics_scaling_gates(current_user: dict = Depends(auth.get_current_user)):
    complete = _build_averaged_complete_projects()
    gates = mtx.compute_scaling_gates(complete)
    from backend.models import ScalingGate
    passed = sum(1 for g in gates if g["status"] == "pass")
    return ScalingGates(
        gates=[ScalingGate(**g) for g in gates],
        passed_count=passed,
        total_count=len(gates),
    )


# ── Activity Log ──────────────────────────────────────────────────────────────

@app.get("/api/activity")
async def list_activity(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: dict = Depends(auth.get_current_user),
):
    rows = db.list_activity(limit=limit, offset=offset)
    total = db.get_activity_count()
    return {"items": rows, "total": total, "limit": limit, "offset": offset}


# ── Monitoring ───────────────────────────────────────────────────────────────

@app.get("/api/monitoring")
async def get_monitoring(
    status_filter: Optional[str] = Query(None, alias="status"),
    current_user: dict = Depends(auth.get_current_user),
):
    projects = db.list_projects(status_filter=status_filter)
    results = []
    total_pending = 0
    for p in projects:
        pioneers = db.list_pioneers(p["id"])
        default_rounds = p.get("default_rounds", 1) or 1
        total_expected = sum((pp.get("total_rounds") or default_rounds) for pp in pioneers)
        total_completed = sum(pp.get("response_count", 0) for pp in pioneers)
        total_pending += max(total_expected - total_completed, 0)
        results.append({
            "id": p["id"],
            "project_name": p["project_name"],
            "category_name": p.get("category_name", ""),
            "status": p["status"],
            "pioneer_count": len(pioneers),
            "responses_completed": total_completed,
            "responses_expected": total_expected,
        })
    total_projects = len(results)
    complete_count = sum(1 for r in results if r["status"] == "complete")
    return {
        "projects": results,
        "total_projects": total_projects,
        "total_pending_responses": total_pending,
        "completion_rate": round(complete_count / total_projects * 100, 1) if total_projects else 0,
    }


# ── Export ────────────────────────────────────────────────────────────────────

def _build_export_workbook(all_projects: list, complete_projects: list):
    """Build the Excel export workbook with Raw Data, Computed Metrics, and Notes sheets."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # ── Sheet 1: Raw Data ──
    ws1 = wb.active
    ws1.title = "Raw Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="121F6B")

    headers1 = [
        "ID", "Project Name", "Category", "Pioneer", "Client",
        "Date Started", "Date Delivered",
        "xCSG Calendar Days", "xCSG Team Size", "xCSG Revisions",
        "Legacy Calendar Days", "Legacy Team Size", "Legacy Revisions",
        "Status", "Created At",
        # B — Machine-First (xcsg / legacy)
        "B1 xcsg", "B1 legacy", "B2 xcsg", "B2 legacy",
        "B3 xcsg", "B3 legacy", "B4 xcsg", "B4 legacy",
        # C — Senior-Led (xcsg only)
        "C1", "C2", "C3", "C4 senior hrs", "C5 junior hrs",
        # D — Proprietary Knowledge (xcsg / legacy)
        "D1 xcsg", "D1 legacy", "D2 xcsg", "D2 legacy",
        "D3 xcsg", "D3 legacy",
        # F — Value Creation (xcsg / legacy)
        "F1 xcsg", "F1 legacy", "F2 xcsg", "F2 legacy",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Build expert response lookup: all responses per project for raw data export
    responses_by_project = {}
    for p_item in all_projects:
        responses_by_project[p_item["id"]] = db.get_all_project_responses(p_item["id"])

    for p in all_projects:
        responses = responses_by_project.get(p["id"], [])
        # Get pioneer names for this project
        project_pioneers = db.list_pioneers(p["id"])
        pioneer_lookup = {pp["id"]: pp.get("display_name") or pp.get("pioneer_name", "") for pp in project_pioneers}
        if not responses:
            # No responses yet — output one row with project info only
            pioneer_names_str = ", ".join(pp.get("display_name") or pp.get("pioneer_name", "") for pp in project_pioneers) if project_pioneers else ""
            ws1.append([
                p["id"], p["project_name"], p.get("category_name", ""),
                pioneer_names_str, p.get("client_name", ""),
                p.get("date_started", ""), p.get("date_delivered", ""),
                p["xcsg_calendar_days"], p["xcsg_team_size"], p["xcsg_revision_rounds"],
                p.get("legacy_calendar_days", ""), "",
                p.get("legacy_revision_rounds", ""),
                p["status"], p["created_at"],
                "", "", "", "", "", "", "", "",
                "", "", "", "", "",
                "", "", "", "", "", "",
                "", "", "", "",
            ])
        else:
            for er in responses:
                pioneer_name = pioneer_lookup.get(er.get("pioneer_id"), "")
                ws1.append([
                    p["id"], p["project_name"], p.get("category_name", ""),
                    pioneer_name, p.get("client_name", ""),
                    p.get("date_started", ""), p.get("date_delivered", ""),
                    p["xcsg_calendar_days"], p["xcsg_team_size"], p["xcsg_revision_rounds"],
                    p.get("legacy_calendar_days", ""), "",
                    p.get("legacy_revision_rounds", ""),
                    p["status"], p["created_at"],
                    # B
                    er.get("b1_starting_point_xcsg", er.get("b1_starting_point", "")),
                    er.get("b1_starting_point_legacy", ""),
                    er.get("b2_research_sources_xcsg", er.get("b2_research_sources", "")),
                    er.get("b2_research_sources_legacy", ""),
                    er.get("b3_assembly_ratio_xcsg", er.get("b3_assembly_ratio", "")),
                    er.get("b3_assembly_ratio_legacy", ""),
                    er.get("b4_hypothesis_first_xcsg", er.get("b4_hypothesis_first", "")),
                    er.get("b4_hypothesis_first_legacy", ""),
                    # C
                    er.get("c1_specialization", ""), er.get("c2_directness", ""), er.get("c3_judgment_pct", ""),
                    er.get("c4_senior_hours", ""), er.get("c5_junior_hours", ""),
                    # D
                    er.get("d1_proprietary_data_xcsg", er.get("d1_proprietary_data", "")),
                    er.get("d1_proprietary_data_legacy", ""),
                    er.get("d2_knowledge_reuse_xcsg", er.get("d2_knowledge_reuse", "")),
                    er.get("d2_knowledge_reuse_legacy", ""),
                    er.get("d3_moat_test_xcsg", er.get("d3_moat_test", "")),
                    er.get("d3_moat_test_legacy", ""),
                    # F
                    er.get("f1_feasibility_xcsg", er.get("f1_feasibility", "")),
                    er.get("f1_feasibility_legacy", ""),
                    er.get("f2_productization_xcsg", er.get("f2_productization", "")),
                    er.get("f2_productization_legacy", ""),
                ])

    # ── Sheet 2: Computed Metrics ──
    ws2 = wb.create_sheet("Computed Metrics")
    headers2 = [
        "ID", "Project Name", "Category", "Pioneer", "Client",
        "xCSG Person-Days", "Legacy Person-Days", "Effort Ratio",
        "xCSG Revisions", "Legacy Revisions", "Quality Ratio", "Value Multiplier",
        "Machine-First Score", "Senior-Led Score", "Proprietary Knowledge Score",
        "Created At",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Use averaged metrics for the computed metrics sheet
    for p in complete_projects:
        responses = db.get_all_project_responses(p["id"])
        if responses:
            export_dict = dict(p)
            export_dict["pioneer_day_rates"] = db.get_pioneer_day_rates(p["id"])
            export_dict["legacy_team"] = db.list_legacy_team(p["id"])
            m = mtx.compute_averaged_project_metrics(export_dict, responses)
            # Get pioneer names for this project
            comp_pioneers = db.list_pioneers(p["id"])
            comp_pioneer_names = ", ".join(pp.get("display_name") or pp.get("pioneer_name", "") for pp in comp_pioneers) if comp_pioneers else ""
            ws2.append([
                m.get("id", p["id"]), m.get("project_name", p["project_name"]),
                m.get("category_name", p.get("category_name", "")),
                comp_pioneer_names,
                m.get("client_name", p.get("client_name", "")),
                m.get("xcsg_person_days"), m.get("legacy_person_days"), m.get("effort_ratio"),
                m.get("xcsg_revisions", p.get("xcsg_revision_rounds", "")),
                m.get("legacy_revisions", p.get("legacy_revision_rounds", "")),
                m.get("quality_ratio"), m.get("value_multiplier"),
                m.get("machine_first_score", ""), m.get("senior_led_score", ""),
                m.get("proprietary_knowledge_score", ""), m.get("created_at", p.get("created_at", "")),
            ])

    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ── Sheet 3: Notes ──
    ws3 = wb.create_sheet("Notes")
    headers3 = ["Project", "Category", "Practice", "Pioneer", "Round", "Submitted", "Notes"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill

    for note in db.list_all_notes():
        ws3.append([
            note.get("project_name", ""),
            note.get("category_name", "") or "",
            note.get("practice_code", "") or "",
            note.get("pioneer_name", "") or "",
            note.get("round_number", "") or "",
            note.get("submitted_at", "") or "",
            note.get("notes", "") or "",
        ])

    # Column widths for Notes sheet (manual defaults — Notes column wide, rest compact).
    notes_widths = {"A": 30, "B": 30, "C": 12, "D": 24, "E": 8, "F": 22, "G": 80}
    for col_letter, width in notes_widths.items():
        ws3.column_dimensions[col_letter].width = width

    # Wrap text in the Notes column so multi-line notes stay readable.
    wrap = Alignment(wrapText=True, vertical="top")
    for row in ws3.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.alignment = wrap

    return wb


# ── Notes feed ────────────────────────────────────────────────────────────────

@app.get("/api/notes")
async def list_notes(
    practice_code: Optional[str] = None,
    category_id: Optional[int] = None,
    pioneer_name: Optional[str] = None,
    delivered_from: Optional[str] = None,
    delivered_to: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(auth.get_current_user),
):
    """Filterable, searchable feed of every expert-submitted note.

    All filters are optional. ``search`` does a case-insensitive substring
    match on the notes text. Returns rows ordered by submitted_at DESC.
    """
    return db.list_all_notes(
        practice_code=practice_code,
        category_id=category_id,
        pioneer_name=pioneer_name,
        delivered_from=delivered_from,
        delivered_to=delivered_to,
        search=search,
    )


@app.get("/api/export/excel")
async def export_excel(current_user: dict = Depends(auth.get_current_user)):
    try:
        import openpyxl  # noqa: F401 — presence check only
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    from backend.excel_export import build_dashboard_sheets
    from backend.schema import DASHBOARD_CONFIG, METRICS

    all_p = db.list_projects()
    complete_raw = db.list_complete_projects()
    # Averaged per-project metrics dicts (carry productivity_ratio, delivery_speed,
    # output_quality, practice_code, category_name, pioneer_name, etc).
    complete_metrics = _build_averaged_complete_projects()
    wb = _build_export_workbook(all_p, complete_raw)

    # Append the 18 dashboard-aggregate sheets that let a user rebuild every
    # chart offline. These operate on the FULL portfolio — dashboard filters
    # are intentionally ignored.
    aggregates = mtx.compute_summary(complete_metrics, all_p)
    scaling_gates = mtx.compute_scaling_gates(complete_metrics)
    build_dashboard_sheets(
        wb,
        complete=complete_metrics,
        all_projects=all_p,
        aggregates=aggregates,
        scaling_gates=scaling_gates,
        chart_configs=DASHBOARD_CONFIG["charts"],
        metrics_defs=METRICS,
    )

    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", prefix="xCSG_Export_", delete=False, dir="/tmp"
    )
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="xCSG_Value_Export.xlsx",
    )


@app.get("/api/export/file/{name}")
async def download_export_file(
    name: str,
    current_user: dict = Depends(auth.get_current_user),
):
    path = f"/tmp/{name}"
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Export file not found")
    return FileResponse(path, media_type="application/octet-stream", filename=name)


PIONEERS_CSV_FIELDS = [
    "id",
    "first_name",
    "last_name",
    "email",
    "notes",
    "project_count",
    "rounds_completed",
    "rounds_expected",
    "completion_rate",
    "last_activity_at",
    "status",
    "avg_quality_score",
    "avg_value_gain",
    "avg_machine_first",
    "avg_senior_led",
    "avg_knowledge",
    "practices",
    "roles",
]


@app.get("/api/export/pioneers.csv")
def export_pioneers_csv(
    practice: Optional[List[str]] = Query(None),
    role: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    search: Optional[str] = Query(None),
    user=Depends(auth.get_current_user),
):
    rows = pioneers_mod.filter_pioneers_for_export(
        practice=practice, role=role, status=status, search=search,
    )

    # Flatten lists into joined strings for CSV.
    flattened = []
    for r in rows:
        flat = dict(r)
        flat["practices"] = ", ".join(
            f"{p['code']}({p['count']})" for p in r.get("practices", [])
        )
        flat["roles"] = ", ".join(
            f"{x['role_name']}×{x['count']}" for x in r.get("roles", [])
        )
        flat.pop("portfolio", None)  # detail-only field, not in list export
        flattened.append(flat)

    # Always emit the header row (even when filter matches nothing) so consumers
    # can rely on the columns. Fieldnames come from a fixed list — not derived
    # from the first row — so an empty result still has a parseable header.
    output = io.StringIO()
    writer = csv.DictWriter(
        output, fieldnames=PIONEERS_CSV_FIELDS, extrasaction="ignore"
    )
    writer.writeheader()
    writer.writerows(flattened)

    # Prepend UTF-8 BOM so Excel on Windows opens accented characters correctly.
    body = "﻿" + output.getvalue()

    return Response(
        content=body,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=pioneers.csv"},
    )


@app.get("/api/export/pioneer/{pioneer_id}.xlsx")
def export_pioneer_xlsx(
    pioneer_id: int,
    user=Depends(auth.get_current_user),
):
    pioneer = pioneers_mod.get_pioneer_with_metrics(pioneer_id)
    if not pioneer:
        raise HTTPException(status_code=404, detail="Pioneer not found")

    from openpyxl import Workbook
    wb = Workbook()

    # summary sheet
    summary = wb.active
    summary.title = "summary"
    summary_cols = [
        "id", "first_name", "last_name", "email", "notes", "status",
        "project_count", "rounds_completed", "rounds_expected", "completion_rate",
        "last_activity_at",
        "avg_quality_score", "avg_value_gain",
        "avg_machine_first", "avg_senior_led", "avg_knowledge",
        "practices", "roles",
    ]
    summary.append(summary_cols)
    practices_str = ", ".join(
        f"{p['code']}({p['count']})" for p in pioneer.get("practices", [])
    )
    roles_str = ", ".join(
        f"{x['role_name']}×{x['count']}" for x in pioneer.get("roles", [])
    )
    summary.append([
        pioneer["id"], pioneer.get("first_name"), pioneer.get("last_name"),
        pioneer["email"], pioneer.get("notes"),
        pioneer["status"],
        pioneer["project_count"], pioneer["rounds_completed"], pioneer["rounds_expected"],
        pioneer.get("completion_rate"), pioneer.get("last_activity_at"),
        pioneer.get("avg_quality_score"), pioneer.get("avg_value_gain"),
        pioneer.get("avg_machine_first"), pioneer.get("avg_senior_led"),
        pioneer.get("avg_knowledge"),
        practices_str, roles_str,
    ])

    # portfolio sheet
    portfolio = wb.create_sheet("portfolio")
    portfolio_cols = [
        "project_id", "project_name", "practice_code", "role_name", "day_rate",
        "rounds_completed", "rounds_expected", "status", "last_activity_at",
    ]
    portfolio.append(portfolio_cols)
    for entry in pioneer.get("portfolio") or []:
        portfolio.append([entry.get(c) for c in portfolio_cols])

    # Stream the workbook as a response.
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pioneer-{pioneer_id}.xlsx"},
    )


# ── Settings ─────────────────────────────────────────────────────────────────

@app.get("/api/settings", response_model=AppSettings)
async def get_settings(current_user: dict = Depends(auth.get_current_user)):
    return db.get_app_settings()


@app.put("/api/settings", response_model=AppSettings)
async def update_settings(
    payload: AppSettingsUpdate,
    current_user: dict = Depends(auth.get_current_user_admin),
):
    db.update_app_settings(default_currency=payload.default_currency)
    return db.get_app_settings()


# ── Pioneer Registry (Phase 3a) ───────────────────────────────────────────────

@app.get("/api/pioneers")
def list_pioneers_endpoint(
    search: Optional[str] = Query(None),
    user=Depends(auth.get_current_user),
):
    """Optional ?search filter matches first_name, last_name, the combined
    'first last' display name, OR email (case-insensitive substring)."""
    if not search:
        return db.list_pioneers_with_metrics()
    return pioneers_mod.filter_pioneers_for_export(search=search)


@app.get("/api/pioneers/{pioneer_id}")
def get_pioneer_endpoint(pioneer_id: int, user=Depends(auth.get_current_user)):
    pioneer = db.get_pioneer_with_metrics(pioneer_id)
    if not pioneer:
        raise HTTPException(status_code=404, detail="Pioneer not found")
    return pioneer


@app.post("/api/pioneers", status_code=201)
def create_pioneer_endpoint(
    payload: PioneerCreate,
    response: Response,
    user=Depends(auth.get_current_user_writer),
):
    """Find-or-create by case-insensitive email when provided.
    201 = new pioneer; 200 = existing pioneer (caller distinguishes by status code)."""
    if payload.email:
        existing = db.find_pioneer_by_email(payload.email)
        if existing:
            response.status_code = 200
            return db.get_pioneer_with_metrics(existing["id"])
    user_id = user.get("sub")
    pid = db.create_pioneer(
        first_name=payload.first_name,
        last_name=payload.last_name,
        email=payload.email,
        notes=payload.notes,
        created_by=user_id,
    )
    return db.get_pioneer_with_metrics(pid)


@app.put("/api/pioneers/{pioneer_id}")
def update_pioneer_endpoint(
    pioneer_id: int,
    payload: PioneerUpdate,
    user=Depends(auth.get_current_user_admin),
):
    pioneer = db.get_pioneer_with_metrics(pioneer_id)
    if not pioneer:
        raise HTTPException(status_code=404, detail="Pioneer not found")
    db.update_pioneer_record(
        pioneer_id,
        first_name=payload.first_name,
        last_name=payload.last_name,
        email=payload.email,
        notes=payload.notes,
    )
    return db.get_pioneer_with_metrics(pioneer_id)


@app.delete("/api/pioneers/{pioneer_id}", status_code=204)
def delete_pioneer_endpoint(
    pioneer_id: int,
    user=Depends(auth.get_current_user_admin),
):
    try:
        db.delete_pioneer(pioneer_id)
    except PioneerInUseError as e:
        raise HTTPException(
            status_code=409,
            detail=f"Pioneer is assigned to {e.project_count} project(s); remove from all projects before deleting.",
        )


# ── Static file mount — MUST BE LAST ─────────────────────────────────────────
_frontend_dir = Path(__file__).parent.parent / "frontend"
if _frontend_dir.exists():
    app.mount("/", StaticFiles(directory=str(_frontend_dir), html=True), name="static")
