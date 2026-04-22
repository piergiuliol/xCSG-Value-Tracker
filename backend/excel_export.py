"""
excel_export.py — Helpers for building the dashboard-aggregate sheets attached
to the Excel export.

The goal of these helpers is to let a user rebuild every dashboard chart
offline by opening the xlsx. Each sheet corresponds to one or more charts and
the data is always the **full portfolio** (dashboard filters are ignored).

All numeric aggregates round to 2 decimal places. Integer counts stay integer.
Percentages are 0-100 floats with 1 decimal (not "%"-suffixed strings).
"""
from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from typing import Iterable, Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook

# ── Styling ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="121F6B")  # Alira navy
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def _apply_header(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Value helpers ─────────────────────────────────────────────────────────────

def _round2(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def _round1(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return round(float(value), 1)
    except (TypeError, ValueError):
        return None


def _avg(values: Iterable) -> Optional[float]:
    clean = [float(v) for v in values if v is not None]
    if not clean:
        return None
    return sum(clean) / len(clean)


def _avg2(values: Iterable) -> Optional[float]:
    avg = _avg(values)
    return _round2(avg) if avg is not None else None


def _quarter_of(date_str: Optional[str]) -> Optional[str]:
    if not date_str:
        return None
    try:
        d = datetime.fromisoformat(str(date_str)[:10])
    except ValueError:
        return None
    return f"{d.year}-Q{(d.month - 1) // 3 + 1}"


def _parse_date(date_str: Optional[str]):
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(str(date_str)[:10])
    except ValueError:
        return None


# ── Aggregate Sheets ──────────────────────────────────────────────────────────

def build_dashboard_aggregates_sheet(wb: Workbook, aggregates: dict, metrics_defs: dict) -> None:
    """Sheet 1: Dashboard Aggregates — metric_key, label, format, value.

    Pulls scalar aggregates used by the dashboard's KPI tiles.
    """
    ws = wb.create_sheet("Dashboard Aggregates")
    ws.append(["metric_key", "label", "format", "value"])
    _apply_header(ws)

    # scale: "as_is" means the server value is already 0-100 for pct rows.
    # "x100" means the server value is 0-1 and must be multiplied for display.
    rows: list[tuple[str, str, str, str]] = [
        # (metric_key, server_key_in_aggregates, format, scale)
        ("total_projects", "total_projects", "int", "as_is"),
        ("completed_count", "completed_count", "int", "as_is"),
        ("pending_projects", "pending_projects", "int", "as_is"),
        ("delivery_speed", "average_effort_ratio", "ratio", "as_is"),
        ("output_quality", "average_quality_ratio", "ratio", "as_is"),
        ("rework_efficiency", "rework_efficiency_avg", "ratio", "as_is"),
        ("productivity_ratio", "average_productivity_ratio", "ratio", "as_is"),
        ("machine_first_score", "machine_first_avg", "ratio", "as_is"),
        ("senior_led_score", "senior_led_avg", "ratio", "as_is"),
        ("proprietary_knowledge_score", "proprietary_knowledge_avg", "ratio", "as_is"),
        ("client_impact", "client_impact_avg", "ratio", "as_is"),
        ("data_independence", "data_independence_avg", "ratio", "as_is"),
        ("overall_xcsg_avg", "overall_xcsg_avg", "ratio", "as_is"),
        ("flywheel_health", "flywheel_health", "ratio", "as_is"),
        ("reuse_intent_avg", "reuse_intent_avg", "pct", "x100"),
        ("reuse_intent_rate", "reuse_intent_rate", "pct", "as_is"),
        ("ai_survival_avg", "ai_survival_avg", "pct", "x100"),
        ("client_pulse_avg", "client_pulse_avg", "pct", "x100"),
        ("on_time_pct", "on_time_pct", "pct", "as_is"),
        ("avg_schedule_delta_days", "avg_schedule_delta_days", "days", "as_is"),
        ("schedule_tracked_count", "schedule_tracked_count", "int", "as_is"),
        ("schedule_on_time_count", "schedule_on_time_count", "int", "as_is"),
        ("scaling_gates_passed", "scaling_gates_passed", "int", "as_is"),
        ("scaling_gates_total", "scaling_gates_total", "int", "as_is"),
        ("checkpoint", "checkpoint", "int", "as_is"),
        ("projects_to_next_checkpoint", "projects_to_next_checkpoint", "int", "as_is"),
    ]
    for metric_key, server_key, fmt, scale in rows:
        meta = metrics_defs.get(metric_key, {})
        label = meta.get("label") or metric_key.replace("_", " ").title()
        value = aggregates.get(server_key)
        if fmt == "int":
            out_value = int(value) if value is not None else None
        elif fmt == "pct":
            if value is None:
                out_value = None
            else:
                scaled = float(value) * 100.0 if scale == "x100" else float(value)
                out_value = _round1(scaled)
        else:
            out_value = _round2(value) if value is not None else None
        ws.append([metric_key, label, fmt, out_value])

    _set_col_widths(ws, {"A": 32, "B": 32, "C": 12, "D": 16})


def build_by_practice_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("By Practice")
    ws.append([
        "practice_code", "practice_name", "project_count",
        "avg_delivery_speed", "avg_output_quality", "avg_value_gain",
    ])
    _apply_header(ws)

    groups: dict[tuple[str, str], list] = defaultdict(list)
    for p in complete:
        code = p.get("practice_code") or ""
        name = p.get("practice_name") or ""
        groups[(code, name)].append(p)

    for (code, name), items in sorted(groups.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "")):
        ws.append([
            code,
            name,
            len(items),
            _avg2(p.get("delivery_speed") for p in items),
            _avg2(p.get("output_quality") for p in items),
            _avg2(p.get("productivity_ratio") for p in items),
        ])
    _set_col_widths(ws, {"A": 14, "B": 32, "C": 14, "D": 18, "E": 18, "F": 18})


def build_by_category_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("By Category")
    ws.append([
        "category_id", "category_name", "project_count",
        "avg_delivery_speed", "avg_output_quality", "avg_value_gain",
    ])
    _apply_header(ws)

    groups: dict[tuple, list] = defaultdict(list)
    for p in complete:
        cat_id = p.get("category_id")
        cat_name = p.get("category_name") or ""
        groups[(cat_id, cat_name)].append(p)

    for (cat_id, cat_name), items in sorted(groups.items(), key=lambda kv: (kv[0][1] or "")):
        ws.append([
            cat_id,
            cat_name,
            len(items),
            _avg2(p.get("delivery_speed") for p in items),
            _avg2(p.get("output_quality") for p in items),
            _avg2(p.get("productivity_ratio") for p in items),
        ])
    _set_col_widths(ws, {"A": 12, "B": 32, "C": 14, "D": 18, "E": 18, "F": 18})


def build_by_pioneer_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("By Pioneer")
    ws.append([
        "pioneer_name", "project_count",
        "avg_delivery_speed", "avg_output_quality", "avg_value_gain",
    ])
    _apply_header(ws)

    groups: dict[str, list] = defaultdict(list)
    for p in complete:
        name = p.get("pioneer_name") or ""
        if not name:
            continue
        # Split comma-separated pioneer names so each pioneer's stats roll up
        # separately (matches the dashboard's "By Pioneer" chart semantics).
        for part in str(name).split(","):
            key = part.strip()
            if key:
                groups[key].append(p)

    for name, items in sorted(groups.items(), key=lambda kv: kv[0].lower()):
        ws.append([
            name,
            len(items),
            _avg2(p.get("delivery_speed") for p in items),
            _avg2(p.get("output_quality") for p in items),
            _avg2(p.get("productivity_ratio") for p in items),
        ])
    _set_col_widths(ws, {"A": 32, "B": 14, "C": 18, "D": 18, "E": 18})


def build_quarterly_trend_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Quarterly Trend")
    ws.append([
        "quarter", "project_count",
        "avg_delivery_speed", "avg_output_quality", "avg_value_gain",
    ])
    _apply_header(ws)

    groups: dict[str, list] = defaultdict(list)
    for p in complete:
        q = _quarter_of(p.get("date_delivered"))
        if q is None:
            continue
        groups[q].append(p)

    for q in sorted(groups.keys()):
        items = groups[q]
        ws.append([
            q,
            len(items),
            _avg2(p.get("delivery_speed") for p in items),
            _avg2(p.get("output_quality") for p in items),
            _avg2(p.get("productivity_ratio") for p in items),
        ])
    _set_col_widths(ws, {"A": 12, "B": 14, "C": 18, "D": 18, "E": 18})


def build_cumulative_trend_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Cumulative Trend")
    ws.append([
        "nth_project", "delivered_date", "project_name",
        "running_avg_delivery_speed", "running_avg_output_quality", "running_avg_value_gain",
    ])
    _apply_header(ws)

    dated = [p for p in complete if _parse_date(p.get("date_delivered"))]
    dated.sort(key=lambda p: _parse_date(p.get("date_delivered")))

    speed_acc: list = []
    quality_acc: list = []
    gain_acc: list = []
    for idx, p in enumerate(dated, start=1):
        speed_acc.append(p.get("delivery_speed"))
        quality_acc.append(p.get("output_quality"))
        gain_acc.append(p.get("productivity_ratio"))
        ws.append([
            idx,
            str(p.get("date_delivered") or "")[:10],
            p.get("project_name") or "",
            _avg2(speed_acc),
            _avg2(quality_acc),
            _avg2(gain_acc),
        ])
    _set_col_widths(ws, {"A": 12, "B": 14, "C": 40, "D": 22, "E": 22, "F": 22})


def build_cohort_practice_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Cohort — Practice")
    ws.append([
        "practice_code", "nth_project_in_cohort", "project_name", "delivered_date", "value_gain",
    ])
    _apply_header(ws)

    by_practice: dict[str, list] = defaultdict(list)
    for p in complete:
        code = p.get("practice_code") or ""
        if not code:
            continue
        if _parse_date(p.get("date_delivered")) is None:
            continue
        by_practice[code].append(p)

    for code in sorted(by_practice.keys()):
        items = sorted(by_practice[code], key=lambda p: _parse_date(p.get("date_delivered")))
        for idx, p in enumerate(items, start=1):
            ws.append([
                code,
                idx,
                p.get("project_name") or "",
                str(p.get("date_delivered") or "")[:10],
                _round2(p.get("productivity_ratio")),
            ])
    _set_col_widths(ws, {"A": 14, "B": 24, "C": 40, "D": 14, "E": 14})


def build_practice_quarter_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Practice × Quarter")
    ws.append(["practice_code", "quarter", "project_count", "avg_value_gain"])
    _apply_header(ws)

    cells: dict[tuple[str, str], list] = defaultdict(list)
    for p in complete:
        code = p.get("practice_code") or ""
        q = _quarter_of(p.get("date_delivered"))
        if not code or not q:
            continue
        cells[(code, q)].append(p)

    for (code, q) in sorted(cells.keys()):
        items = cells[(code, q)]
        ws.append([
            code,
            q,
            len(items),
            _avg2(p.get("productivity_ratio") for p in items),
        ])
    _set_col_widths(ws, {"A": 14, "B": 12, "C": 14, "D": 18})


def build_category_mix_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Category Mix by Quarter")
    ws.append(["quarter", "category_name", "project_count", "share_pct_of_quarter"])
    _apply_header(ws)

    per_q: dict[str, Counter] = defaultdict(Counter)
    q_totals: Counter = Counter()
    for p in complete:
        q = _quarter_of(p.get("date_delivered"))
        cat = p.get("category_name") or "Unknown"
        if not q:
            continue
        per_q[q][cat] += 1
        q_totals[q] += 1

    for q in sorted(per_q.keys()):
        total = q_totals[q] or 1
        for cat, count in sorted(per_q[q].items()):
            share = (count / total) * 100.0
            ws.append([q, cat, count, _round1(share)])
    _set_col_widths(ws, {"A": 12, "B": 32, "C": 14, "D": 22})


def build_disprove_matrix_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Disprove Matrix")
    ws.append([
        "project_name", "practice_code", "category_name",
        "delivery_speed", "output_quality", "value_gain", "quadrant",
    ])
    _apply_header(ws)

    for p in complete:
        speed = p.get("delivery_speed")
        quality = p.get("output_quality")
        if speed is None or quality is None:
            quadrant = "NA"
        else:
            if speed >= 1 and quality >= 1:
                quadrant = "top-right"
            elif speed < 1 and quality >= 1:
                quadrant = "top-left"
            elif speed >= 1 and quality < 1:
                quadrant = "bottom-right"
            else:
                quadrant = "bottom-left"
        ws.append([
            p.get("project_name") or "",
            p.get("practice_code") or "",
            p.get("category_name") or "",
            _round2(speed),
            _round2(quality),
            _round2(p.get("productivity_ratio")),
            quadrant,
        ])
    _set_col_widths(ws, {"A": 40, "B": 14, "C": 24, "D": 16, "E": 16, "F": 14, "G": 14})


def build_gains_radar_sheet(wb: Workbook, aggregates: dict) -> None:
    ws = wb.create_sheet("Gains Radar")
    ws.append(["dimension", "portfolio_avg", "baseline"])
    _apply_header(ws)

    dims = [
        ("Machine-First", "machine_first_avg"),
        ("Senior-Led", "senior_led_avg"),
        ("Knowledge", "proprietary_knowledge_avg"),
        ("Rework Eff.", "rework_efficiency_avg"),
        ("Client Impact", "client_impact_avg"),
        ("Data Ind.", "data_independence_avg"),
    ]
    for label, key in dims:
        ws.append([label, _round2(aggregates.get(key)), 1.0])
    _set_col_widths(ws, {"A": 18, "B": 16, "C": 12})


def build_client_pulse_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Client Pulse")
    ws.append(["response", "count", "pct_of_total"])
    _apply_header(ws)

    counter: Counter = Counter()
    for p in complete:
        val = p.get("client_pulse")
        if val:
            counter[str(val)] += 1
    total = sum(counter.values()) or 1
    for response, count in sorted(counter.items(), key=lambda kv: -kv[1]):
        ws.append([response, count, _round1((count / total) * 100.0)])
    _set_col_widths(ws, {"A": 28, "B": 10, "C": 16})


def build_reuse_intent_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Reuse Intent")
    ws.append(["response", "count", "pct_of_total"])
    _apply_header(ws)

    counter: Counter = Counter()
    for p in complete:
        raw = p.get("g1_reuse_intent")
        if raw:
            counter[str(raw)] += 1
            continue
        # Fallback: derive bucket from numeric reuse_intent_score
        score = p.get("reuse_intent_score")
        if score is None:
            continue
        if score == 1.0:
            counter["Yes without hesitation"] += 1
        elif score == 0.5:
            counter["Yes with reservations"] += 1
        else:
            counter["No — legacy would have been better"] += 1

    total = sum(counter.values()) or 1
    for response, count in sorted(counter.items(), key=lambda kv: -kv[1]):
        ws.append([response, count, _round1((count / total) * 100.0)])
    _set_col_widths(ws, {"A": 36, "B": 10, "C": 16})


def build_schedule_variance_sheet(wb: Workbook, all_projects: list) -> None:
    ws = wb.create_sheet("Schedule Variance")
    ws.append([
        "project_name", "practice_code", "date_expected", "date_delivered",
        "delta_days", "on_time",
    ])
    _apply_header(ws)

    for p in all_projects:
        expected = p.get("date_expected_delivered")
        actual = p.get("date_delivered")
        d_exp = _parse_date(expected)
        d_act = _parse_date(actual)
        if not d_exp or not d_act:
            continue
        delta = (d_act - d_exp).days
        ws.append([
            p.get("project_name") or "",
            p.get("practice_code") or "",
            str(expected)[:10],
            str(actual)[:10],
            delta,
            bool(delta <= 0),
        ])
    _set_col_widths(ws, {"A": 40, "B": 14, "C": 14, "D": 14, "E": 12, "F": 10})


def build_scaling_gates_sheet(wb: Workbook, scaling_gates: list) -> None:
    ws = wb.create_sheet("Scaling Gates")
    ws.append(["gate_id", "gate_name", "description", "threshold", "status", "detail"])
    _apply_header(ws)

    for gate in scaling_gates:
        ws.append([
            gate.get("id"),
            gate.get("name") or "",
            gate.get("description") or "",
            gate.get("threshold") or "",
            gate.get("status") or "",
            gate.get("detail") or "",
        ])
    _set_col_widths(ws, {"A": 10, "B": 26, "C": 60, "D": 16, "E": 12, "F": 60})


def build_top_movers_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Top Movers")
    ws.append(["rank", "project_name", "practice_code", "category_name", "value_gain"])
    _apply_header(ws)

    scored = [p for p in complete if p.get("productivity_ratio") is not None]
    scored.sort(key=lambda p: p.get("productivity_ratio") or 0, reverse=True)
    for rank, p in enumerate(scored[:5], start=1):
        ws.append([
            rank,
            p.get("project_name") or "",
            p.get("practice_code") or "",
            p.get("category_name") or "",
            _round2(p.get("productivity_ratio")),
        ])
    _set_col_widths(ws, {"A": 8, "B": 40, "C": 14, "D": 24, "E": 14})


def build_bottom_movers_sheet(wb: Workbook, complete: list) -> None:
    ws = wb.create_sheet("Bottom Movers")
    ws.append(["rank", "project_name", "practice_code", "category_name", "value_gain"])
    _apply_header(ws)

    scored = [p for p in complete if p.get("productivity_ratio") is not None]
    scored.sort(key=lambda p: p.get("productivity_ratio") or 0)
    for rank, p in enumerate(scored[:5], start=1):
        ws.append([
            rank,
            p.get("project_name") or "",
            p.get("practice_code") or "",
            p.get("category_name") or "",
            _round2(p.get("productivity_ratio")),
        ])
    _set_col_widths(ws, {"A": 8, "B": 40, "C": 14, "D": 24, "E": 14})


def build_takeaways_sheet(wb: Workbook, complete: list, aggregates: dict, scaling_gates: list, chart_configs: list) -> None:
    from backend.takeaways import compute_takeaways

    ws = wb.create_sheet("Takeaways")
    ws.append(["chart_id", "chart_title", "chart_type", "takeaway"])
    _apply_header(ws)

    takeaways = compute_takeaways(complete, aggregates, scaling_gates, chart_configs)
    for cfg in chart_configs:
        cid = cfg.get("id") or ""
        ws.append([
            cid,
            cfg.get("title") or "",
            cfg.get("type") or "",
            takeaways.get(cid, "") or "",
        ])
    _set_col_widths(ws, {"A": 22, "B": 34, "C": 26, "D": 80})

    # Wrap the takeaway column so multi-sentence text stays readable.
    wrap = Alignment(wrapText=True, vertical="top")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = wrap


# ── Orchestrator ──────────────────────────────────────────────────────────────

def build_dashboard_sheets(
    wb: Workbook,
    *,
    complete: list,
    all_projects: list,
    aggregates: dict,
    scaling_gates: list,
    chart_configs: list,
    metrics_defs: dict,
) -> None:
    """Append all 18 dashboard-aggregate sheets to `wb` in spec order.

    Safe with an empty portfolio — each sheet still writes its header row.
    """
    build_dashboard_aggregates_sheet(wb, aggregates, metrics_defs)
    build_by_practice_sheet(wb, complete)
    build_by_category_sheet(wb, complete)
    build_by_pioneer_sheet(wb, complete)
    build_quarterly_trend_sheet(wb, complete)
    build_cumulative_trend_sheet(wb, complete)
    build_cohort_practice_sheet(wb, complete)
    build_practice_quarter_sheet(wb, complete)
    build_category_mix_sheet(wb, complete)
    build_disprove_matrix_sheet(wb, complete)
    build_gains_radar_sheet(wb, aggregates)
    build_client_pulse_sheet(wb, complete)
    build_reuse_intent_sheet(wb, complete)
    build_schedule_variance_sheet(wb, all_projects)
    build_scaling_gates_sheet(wb, scaling_gates)
    build_top_movers_sheet(wb, complete)
    build_bottom_movers_sheet(wb, complete)
    build_takeaways_sheet(wb, complete, aggregates, scaling_gates, chart_configs)
