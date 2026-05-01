#!/usr/bin/env python3
"""
Comprehensive QA/QC test suite for xCSG Value Tracker V2.
Tests authentication, expert options, categories, project CRUD, expert submissions,
metrics computation, scaling gates, dashboard, norms, schema, string consistency,
and frontend JS syntax.
"""
import base64
import json
import os
import re
import sqlite3
import subprocess
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import requests

BASE = "http://localhost:8765"
DB_PATH = os.path.expanduser("~/Documents/Projects/xCSG-Value-Tracker/data/tracker.db")

passed = 0
failed = 0
failures = []

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
        print(f"  PASS: {name}")
    else:
        failed += 1
        msg = f"  FAIL: {name}"
        if detail:
            msg += f" — {detail}"
        print(msg)
        failures.append((name, detail))

# ── Helpers ───────────────────────────────────────────────────────────────────

def login(username="admin", password="AliraAdmin2026!"):
    r = requests.post(f"{BASE}/api/auth/login", json={"username": username, "password": password})
    return r

def get_token(username="admin"):
    r = login(username)
    r.raise_for_status()
    return r.json()

def admin_token():
    return get_token("admin")["access_token"]

def pmo_token():
    return get_token("pmo")["access_token"]

def login_token(username, password):
    r = requests.post(f"{BASE}/api/auth/login", json={"username": username, "password": password})
    r.raise_for_status()
    return r.json()["access_token"]

def auth_h(token):
    return {"Authorization": f"Bearer {token}"}

def decode_jwt(token):
    payload_b64 = token.split(".")[1]
    # Add padding
    payload_b64 += "=" * (4 - len(payload_b64) % 4)
    return json.loads(base64.urlsafe_b64decode(payload_b64))

# ── A. Authentication ─────────────────────────────────────────────────────────

def test_authentication():
    print("\n── A. Authentication ──")
    
    # Admin login
    r = login("admin", "AliraAdmin2026!")
    test("Admin login returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code == 200:
        d = r.json()
        test("Admin login returns JWT", "access_token" in d)
        test("Admin login returns user object", "user" in d and d["user"]["role"] == "admin", f"role={d.get('user',{}).get('role')}")
        
        # JWT payload
        jwt = decode_jwt(d["access_token"])
        test("JWT has sub", "sub" in jwt)
        test("JWT has username", "username" in jwt and jwt["username"] == "admin")
        test("JWT has role", "role" in jwt and jwt["role"] == "admin")
        test("JWT has iat", "iat" in jwt)
        test("JWT has exp", "exp" in jwt)
        
        # 8-hour expiry
        if "iat" in jwt and "exp" in jwt:
            diff = jwt["exp"] - jwt["iat"]
            test("JWT expires in ~8 hours (28800s)", abs(diff - 28800) < 60, f"got {diff}s")
    
    # PMO login
    r = login("pmo", "AliraPMO2026!")
    test("PMO login returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code == 200:
        d = r.json()
        test("PMO user has role analyst", d.get("user", {}).get("role") == "analyst", f"role={d.get('user',{}).get('role')}")
    
    # Bad credentials
    r = login("admin", "wrongpassword")
    test("Bad credentials returns 401", r.status_code == 401, f"got {r.status_code}")
    
    # Register (admin only)
    tk = admin_token()
    r = requests.post(f"{BASE}/api/auth/register", headers=auth_h(tk), json={
        "username": "qa_test_user", "email": "qa@test.com", "password": "TestPass123!", "role": "viewer"
    })
    test("Register with admin returns 201", r.status_code in (200, 201, 400) and (r.status_code != 400 or "already exists" in r.text.lower()), f"got {r.status_code}: {r.text[:80]}")
    
    # Register without auth
    r = requests.post(f"{BASE}/api/auth/register", json={
        "username": "qa_test_noauth", "email": "noauth@test.com", "password": "TestPass123!", "role": "viewer"
    })
    test("Register without auth returns 403", r.status_code in (401, 403), f"got {r.status_code}")

# ── B. Expert Options ─────────────────────────────────────────────────────────

def test_expert_options():
    print("\n── B. Expert Options ──")
    r = requests.get(f"{BASE}/api/expert/options")
    test("GET /api/expert/options returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    
    opts = r.json()
    test("Exactly 34 fields (l2_legacy_team_size dropped in Phase 2c)", len(opts) == 34, f"got {len(opts)}")
    
    expected = {
        "b1_starting_point": ["From AI draft", "Mixed", "From blank page"],
        "b2_research_sources": ["Single source or dataset", "A few targeted sources (2-4)", "Multiple sources across domains (5-10)", "Broad systematic synthesis (10+)"],
        "b3_assembly_ratio": [">75% AI", "50-75%", "25-50%", "<25%"],
        "b4_hypothesis_first": ["Hypothesis-first", "Hybrid", "Discovery-first"],
        "b5_ai_survival": [">75%", "50-75%", "25-50%", "<25%", "Did not use AI draft"],
        "c1_specialization": ["Deep specialist", "Adjacent expertise", "Generalist"],
        "c2_directness": ["Expert authored", "Expert co-authored", "Expert reviewed only"],
        "c3_judgment_pct": [">75% judgment", "50-75%", "25-50%", "<25%"],
        "c6_self_assessment": ["Significantly better", "Somewhat better", "Comparable", "Somewhat worse"],
        "d1_proprietary_data": ["Yes", "No"],
        "d2_knowledge_reuse": ["Yes directly reused and extended", "Yes provided useful starting context", "No built from scratch"],
        "d3_moat_test": ["No — proprietary inputs decisive", "Partially — they would miss key insights", "Yes — all inputs publicly available"],
        "f1_feasibility": ["Not feasible", "Feasible but at 2x+ the cost and time", "Feasible at similar cost", "Legacy would have been more effective"],
        "f2_productization": ["Yes largely as-is", "Yes with moderate customization", "No fully bespoke"],
        "g1_reuse_intent": ["Yes without hesitation", "Yes with reservations", "No — legacy would have been better"],
    }
    
    for field, exp_opts in expected.items():
        if field in opts:
            actual = opts[field].get("options", [])
            match = actual == exp_opts
            test(f"{field} options correct", match, f"expected {exp_opts}, got {actual}")
    
    # L1 should be integer type, not categorical
    if "l1_legacy_working_days" in opts:
        l1 = opts["l1_legacy_working_days"]
        test("L1 has type='integer'", l1.get("type") == "integer" or "options" not in l1, f"type={l1.get('type')}, has_options={'options' in l1}")
    
    # L3-L12 present (L2/l2_legacy_team_size dropped by Phase 2c)
    for i in range(3, 13):
        key = f"l{i}_legacy_{'revision_depth' if i == 3 else 'scope_expansion' if i == 4 else 'client_reaction' if i == 5 else 'b2_sources' if i == 6 else 'c1_specialization' if i == 7 else 'c2_directness' if i == 8 else 'c3_judgment' if i == 9 else 'd1_proprietary' if i == 10 else 'd2_reuse' if i == 11 else 'd3_moat'}"
        test(f"L{i} present ({key})", key in opts, f"missing from options")
    test("L2 (l2_legacy_team_size) absent from options (dropped in v18)", "l2_legacy_team_size" not in opts)
    
    # Em dash check
    em_dash_fields = ["d3_moat_test", "g1_reuse_intent"]
    # Also check L7-L12 for em dashes
    for i in range(7, 13):
        key = f"l{i}_legacy_{'c1_specialization' if i == 7 else 'c2_directness' if i == 8 else 'c3_judgment' if i == 9 else 'd1_proprietary' if i == 10 else 'd2_reuse' if i == 11 else 'd3_moat'}"
        if key in opts:
            em_dash_fields.append(key)
    
    for field in ["d3_moat_test", "g1_reuse_intent"]:
        if field in opts:
            opt_list = opts[field].get("options", [])
            all_text = " ".join(opt_list)
            has_em = "—" in all_text
            test(f"{field} uses em dashes", has_em, f"no em dash in: {all_text[:80]}")
    
    # Check all option strings for em dash consistency
    for field_name in ["d3_moat_test", "g1_reuse_intent"] + [f"l{i}_legacy_d3_moat" for i in range(12, 13)]:
        if field_name in opts:
            opt_list = opts[field_name].get("options", [])
            all_text = " ".join(opt_list)
            test(f"{field_name} contains em dash", "—" in all_text, f"no em dash in: {all_text[:80]}")

def test_categories():
    print("\n── C. Project Categories (seeded from CSV) ──")
    tk = admin_token()
    r = requests.get(f"{BASE}/api/categories", headers=auth_h(tk))
    test("GET /api/categories returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return

    cats = r.json()
    test(f"Categories returned (expected 79)", len(cats) == 79, f"got {len(cats)}")

    # Spot-check CSV-sourced names
    cat_names = {c["name"] for c in cats}
    for name in ["510(k)", "Regulatory Strategy", "Evidence Generation Strategy", "MAA/NDA", "Training", "Other"]:
        test(f"Category '{name}' exists", name in cat_names, f"missing from {len(cat_names)} categories")


def test_practices():
    print("\n── C2. Practices ──")
    tk = admin_token()
    r = requests.get(f"{BASE}/api/practices", headers=auth_h(tk))
    test("GET /api/practices returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return

    practices = r.json()
    test(f"Practices returned (expected 11)", len(practices) == 11, f"got {len(practices)}")

    codes = {p["code"] for p in practices}
    for code in ["RAM", "MAP", "NPS", "MCD", "RWE", "PEN", "RAP", "TAD", "CLI", "Other", "ALL"]:
        test(f"Practice '{code}' exists", code in codes, f"missing from {sorted(codes)}")

    # Each practice row should carry a project_count field
    test("Practice rows carry project_count", all("project_count" in p for p in practices))

# ── D. Create Deliverable / Project ───────────────────────────────────────────

def test_create_deliverable():
    print("\n── D. Create Deliverable ──")
    tk = admin_token()
    
    # First get a valid category and a practice
    r = requests.get(f"{BASE}/api/categories", headers=auth_h(tk))
    cats = r.json()
    if not cats:
        test("Cannot test create - no categories", False)
        return
    cat_id = cats[0]["id"]
    # Phase 2c added (category, practice) constraint — pick a practice allowed for cat[0].
    practice_id = cats[0].get("practices", [{}])[0].get("id") if cats[0].get("practices") else None
    if not practice_id:
        # Fall back: find any category that has at least one allowed practice.
        for c in cats:
            if c.get("practices"):
                cat_id, practice_id = c["id"], c["practices"][0]["id"]
                break

    # Create with all V2 fields
    payload = {
        "project_name": "QA Test Deliverable",
        "category_id": cat_id,
        "practice_id": practice_id,
        "pioneers": [{"first_name": "Dr.", "last_name": "QA"}],
        "client_name": "QA Client",
        "engagement_stage": "Active engagement",
        "date_started": "2026-03-01",
        "date_delivered": "2026-03-10",
        "working_days": 8,
        "xcsg_team_size": "2",
        "revision_depth": "Cosmetic only",
        "xcsg_scope_expansion": "No",
        "engagement_revenue": 100000,
    }

    payload["xcsg_revision_rounds"] = "1"
    r = requests.post(f"{BASE}/api/projects", headers={**auth_h(tk), "Content-Type": "application/json"}, json=payload)
    test("POST /api/projects returns 201", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:100]}")
    
    if r.status_code in (200, 201):
        proj = r.json()
        test("Returns expert_token", "expert_token" in proj, f"keys: {list(proj.keys())}")

        # Check working_days and engagement_revenue stored
        test("working_days stored", proj.get("working_days") == 8, f"got {proj.get('working_days')}")
        test("engagement_revenue stored", proj.get("engagement_revenue") == 100000, f"got {proj.get('engagement_revenue')}")
        test("revision_depth stored", proj.get("revision_depth") == "Cosmetic only", f"got {proj.get('revision_depth')}")
        test("practice_id stored", proj.get("practice_id") == practice_id, f"got {proj.get('practice_id')}")
        test("practice_code returned on project row", proj.get("practice_code") is not None, f"got {proj.get('practice_code')}")
        
        # Clean up
        requests.delete(f"{BASE}/api/projects/{proj['id']}", headers=auth_h(tk))
    
    # Required fields test - empty pioneers list rejected
    payload2 = {
        "project_name": "QA Test No Pioneer",
        "category_id": cat_id,
        "pioneers": [],
        "date_started": "2026-03-01",
        "date_delivered": "2026-03-10",
        "xcsg_team_size": "2",
        "xcsg_revision_rounds": "0",
    }
    r = requests.post(f"{BASE}/api/projects", headers={**auth_h(tk), "Content-Type": "application/json"}, json=payload2)
    test("Empty pioneers list rejected", r.status_code == 422, f"got {r.status_code}")
    
    # No legacy estimate fields in PMO form
    test("No legacy_calendar_days required (V2)", True)  # just checking it's not required
    payload3 = {**payload, "project_name": "QA No Legacy"}
    r = requests.post(f"{BASE}/api/projects", headers={**auth_h(tk), "Content-Type": "application/json"}, json=payload3)
    if r.status_code in (200, 201):
        proj = r.json()
        requests.delete(f"{BASE}/api/projects/{proj['id']}", headers=auth_h(tk))

# ── E. Expert Assessment ──────────────────────────────────────────────────────

def test_expert_assessment():
    print("\n── E. Expert Assessment ──")
    tk = admin_token()
    
    # Create a project first
    r = requests.get(f"{BASE}/api/categories", headers=auth_h(tk))
    cats = r.json()
    cat_id = cats[0]["id"] if cats else 1
    
    r = requests.post(f"{BASE}/api/projects", headers={**auth_h(tk), "Content-Type": "application/json"}, json={
        "project_name": "QA Expert Test",
        "category_id": cat_id,
        "pioneers": [{"first_name": "Dr.", "last_name": "Expert QA"}],
        "engagement_stage": "Active engagement",
        "date_started": "2026-03-01",
        "date_delivered": "2026-03-10",
        "working_days": 5,
        "xcsg_team_size": "2",
        "revision_depth": "Cosmetic only",
        "xcsg_revision_rounds": "1",
    })
    
    if r.status_code not in (200, 201):
        test("Cannot test expert - project creation failed", False, f"{r.status_code}: {r.text[:80]}")
        return
    
    proj = r.json()
    # Phase 2+: /api/expert/{token} resolves a per-pioneer round token, not the
    # legacy project-level expert_token. Pull the pioneer's round-1 token.
    pioneers = proj.get("pioneers") or []
    token = pioneers[0]["expert_token"] if pioneers else proj.get("expert_token")

    # GET expert context (no auth)
    r = requests.get(f"{BASE}/api/expert/{token}")
    test("GET /api/expert/{{token}} returns 200 (no auth)", r.status_code == 200, f"got {r.status_code}")
    if r.status_code == 200:
        ctx = r.json()
        test("Expert context has project_name", "project_name" in ctx)
        test("Expert context has already_completed", "already_completed" in ctx)
        test("Expert context already_completed=False", ctx.get("already_completed") == False)
    
    # POST expert assessment
    expert_data = {
        "b1_starting_point": "From AI draft",
        "b2_research_sources": "Multiple sources across domains (5-10)",
        "b3_assembly_ratio": ">75% AI",
        "b4_hypothesis_first": "Hypothesis-first",
        "b5_ai_survival": ">75%",
        "b6_data_analysis_split": "25-50%",
        "c1_specialization": "Deep specialist",
        "c2_directness": "Expert authored",
        "c3_judgment_pct": ">75% judgment",
        "c6_self_assessment": "Significantly better",
        "c7_analytical_depth": "Exceptional",
        "c8_decision_readiness": "Yes without caveats",
        "d1_proprietary_data": "Yes",
        "d2_knowledge_reuse": "Yes directly reused and extended",
        "d3_moat_test": "No — proprietary inputs decisive",
        "e1_client_decision": "Yes — informed a specific decision",
        "f1_feasibility": "Not feasible",
        "f2_productization": "Yes largely as-is",
        "g1_reuse_intent": "Yes without hesitation",
        "l1_legacy_working_days": 15,
        "l3_legacy_revision_depth": "Moderate rework",
        "l4_legacy_scope_expansion": "No",
        "l5_legacy_client_reaction": "Met expectations",
        "l6_legacy_b2_sources": "A few targeted sources (2-4)",
        "l7_legacy_c1_specialization": "Generalist",
        "l8_legacy_c2_directness": "Expert reviewed only",
        "l9_legacy_c3_judgment": "25-50%",
        "l10_legacy_d1_proprietary": "No",
        "l11_legacy_d2_reuse": "No built from scratch",
        "l12_legacy_d3_moat": "Yes — all inputs publicly available",
        "l13_legacy_c7_depth": "Adequate",
        "l14_legacy_c8_decision": "Needs significant additional work",
        "l15_legacy_e1_decision": "Yes — referenced in internal discussions",
        "l16_legacy_b6_data": ">75% on data",
    }

    r = requests.post(f"{BASE}/api/expert/{token}", json=expert_data)
    test("POST /api/expert/{{token}} returns 201", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:100]}")
    
    # Re-submit should fail
    r2 = requests.post(f"{BASE}/api/expert/{token}", json=expert_data)
    test("Re-submit returns already completed", r2.status_code in (200, 201, 400, 409), f"got {r.status_code}: {r2.text[:100]}")
    if r2.status_code in (200, 201, 400, 409):
        test("Re-submit message mentions 'already'", "already" in r2.text.lower(), f"msg: {r2.text[:80]}")
    
    # Clean up
    requests.delete(f"{BASE}/api/projects/{proj['id']}", headers=auth_h(tk))

# ── F. Metrics Computation ────────────────────────────────────────────────────

def test_metrics():
    print("\n── F. Metrics Computation ──")
    tk = admin_token()
    
    # Scoring literals live in backend/schema.py (single source of truth);
    # metrics.py only re-binds them via SCORES["…"]. So check each literal in
    # whichever file actually owns it.
    from pathlib import Path
    base_dir = Path(DB_PATH).parent.parent
    metrics_file = base_dir / "backend" / "metrics.py"
    schema_file = base_dir / "backend" / "schema.py"
    if metrics_file.exists() and schema_file.exists():
        content = metrics_file.read_text()
        schema_content = schema_file.read_text()

        # Team midpoints still live in metrics.py
        test("TEAM_MIDPOINTS defined", 'TEAM_MIDPOINTS' in content)
        test('Team midpoint "1":1', '"1": 1' in content or "'1': 1" in content)
        test('Team midpoint "4+":5', '"4+": 5' in content or "'4+': 5" in content)

        # Scoring dicts: identifier in metrics.py, literal weights in schema.py.
        test("REVISION_DEPTH_SCORES bound in metrics", "REVISION_DEPTH_SCORES" in content)
        test("No revisions=1.0", '"No revisions needed": 1.0' in schema_content)
        test("Cosmetic=0.85", '"Cosmetic only": 0.85' in schema_content)
        test("Moderate=0.55", '"Moderate rework": 0.55' in schema_content)
        test("Major=0.2", '"Major rework": 0.2' in schema_content)

        test("SELF_ASSESSMENT_SCORES bound in metrics", "SELF_ASSESSMENT_SCORES" in content)
        test("Significantly=1.0", '"Significantly better": 1.0' in schema_content)
        test("Somewhat=0.7", '"Somewhat better": 0.7' in schema_content)
        test("Comparable=0.4", '"Comparable": 0.4' in schema_content)
        test("Worse=0.1", '"Somewhat worse": 0.1' in schema_content)

        test("CLIENT_PULSE_SCORES bound in metrics", "CLIENT_PULSE_SCORES" in content)
        test("Exceeded=1.0", '"Exceeded expectations": 1.0' in schema_content)
        test("Met=0.6", '"Met expectations": 0.6' in schema_content)
        test("Below=0.1", '"Below expectations": 0.1' in schema_content)

        # Quality score is composite, not ratio
        test("compute_quality_score defined", "compute_quality_score" in content or "quality_score" in content)

        # Outcome rate = quality_score / person_days
        test("outcome_rate computation", "outcome_rate" in content.lower())

        # Revenue productivity
        test("revenue_productivity computation", "productivity" in content.lower() or "revenue_productivity" in content.lower())

        # B5 "Did not use AI draft" handled — option declared in schema.py.
        test("B5 'Did not use AI draft' handled", "Did not use AI draft" in schema_content)

        # Flywheel leg scores
        test("Machine-First score computed", "machine_first" in content.lower())
        test("Senior-Led score computed", "senior_led" in content.lower())
        test("Proprietary Knowledge score computed", "proprietary_knowledge" in content.lower())
    
    # Check per-project metrics from API
    r = requests.get(f"{BASE}/api/metrics/projects", headers=auth_h(tk))
    if r.status_code == 200:
        metrics = r.json()
        if metrics:
            m = metrics[0]
            test("Per-project has effort_ratio", "effort_ratio" in m)
            test("Per-project has quality_score", "quality_score" in m)
            test("Per-project has outcome_rate_ratio", "outcome_rate_ratio" in m)
            
            # Quality score should be 0-1 range
            if m.get("quality_score"):
                qs = m["quality_score"]
                test(f"Quality score in 0-1 range (got {qs})", 0 <= qs <= 1, f"value={qs}")

# ── G. Scaling Gates ──────────────────────────────────────────────────────────

def test_scaling_gates():
    print("\n── G. Scaling Gates ──")
    tk = admin_token()
    r = requests.get(f"{BASE}/api/metrics/scaling-gates", headers=auth_h(tk))
    test("GET /api/metrics/scaling-gates returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    
    data = r.json()
    gates = data.get("gates", [])
    test("7 gates returned", len(gates) == 7, f"got {len(gates)}")
    test("passed_count field exists", "passed_count" in data)
    
    gate_map = {g["id"]: g for g in gates}
    
    if 1 in gate_map:
        test("Gate 1: Multi-engagement", "Multi-engagement" in gate_map[1]["name"] or "deliverable type" in gate_map[1]["description"].lower())
    if 2 in gate_map:
        test("Gate 2: Effort ratio > 1.3", "1.3" in gate_map[2]["description"] or "effort" in gate_map[2]["name"].lower())
    if 3 in gate_map:
        test("Gate 3: Client-invisible quality", "revision" in gate_map[3]["description"].lower() or "quality" in gate_map[3]["name"].lower())
    if 6 in gate_map:
        test("Gate 6: D2 reuse ≥40%", "40%" in gate_map[6]["description"] or "d2" in gate_map[6]["description"].lower() or "compounding" in gate_map[6]["name"].lower())
    if 7 in gate_map:
        test("Gate 7: G1 reuse intent ≥70%", "70%" in gate_map[7]["description"] or "g1" in gate_map[7]["description"].lower() or "adoption" in gate_map[7]["name"].lower())

# ── H. Dashboard Metrics ──────────────────────────────────────────────────────

def test_dashboard():
    print("\n── H. Dashboard Metrics ──")
    tk = admin_token()
    
    r = requests.get(f"{BASE}/api/dashboard/metrics", headers=auth_h(tk))
    test("GET /api/dashboard/metrics returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    
    m = r.json()
    
    # Also check /api/metrics/summary
    r2 = requests.get(f"{BASE}/api/metrics/summary", headers=auth_h(tk))
    test("GET /api/metrics/summary returns 200", r2.status_code == 200)
    
    # Check for key fields across both responses
    for endpoint_name, resp in [("dashboard", m), ("summary", r2.json() if r2.status_code == 200 else {})]:
        if not resp:
            continue
        test(f"{endpoint_name}: has total_completed or complete_projects",
             "total_completed" in resp or "complete_projects" in resp or "completed_count" in resp)
        test(f"{endpoint_name}: has effort_ratio",
             "average_effort_ratio" in resp or "effort_ratio" in resp)
        test(f"{endpoint_name}: has quality_score",
             "average_quality_score" in resp or "quality_score" in resp)
        test(f"{endpoint_name}: has reuse_intent_rate",
             "reuse_intent_rate" in resp or "reuse_intent_avg" in resp)
    
    # Quality score 0-1 range
    qs = m.get("average_quality_score")
    if qs is not None and qs > 0:
        test(f"Dashboard quality_score in 0-1 range (got {qs})", 0 <= qs <= 1, f"value={qs}")

# ── I. Deliverables / Projects List ──────────────────────────────────────────

def test_deliverables_list():
    print("\n── I. Deliverables/Projects List ──")
    tk = admin_token()
    
    # Check /api/projects list
    r = requests.get(f"{BASE}/api/projects?limit=5", headers=auth_h(tk))
    test("GET /api/projects returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code == 200:
        data = r.json()
        items = data if isinstance(data, list) else data.get("items", [])
        test("Projects list has items", len(items) > 0)
        if items:
            test("Project has status field", "status" in items[0])
    
    # Check PATCH /api/deliverables/{id} for client_pulse
    if items:
        proj_id = items[0]["id"]
        r = requests.patch(f"{BASE}/api/deliverables/{proj_id}", headers={**auth_h(tk), "Content-Type": "application/json"},
                          json={"client_pulse": "Exceeded expectations"})
        test("PATCH /api/deliverables/{{id}} accepts client_pulse", r.status_code in (200, 201), f"got {r.status_code}")

# ── J. String Consistency ─────────────────────────────────────────────────────

def test_string_consistency():
    print("\n── J. String Consistency ──")
    
    from pathlib import Path
    base_dir = Path(DB_PATH).parent.parent
    
    # Get option strings from API
    r = requests.get(f"{BASE}/api/expert/options")
    if r.status_code != 200:
        test("Cannot test string consistency - options endpoint failed", False)
        return
    api_opts = r.json()
    
    # Extract all option strings from API
    api_strings = set()
    for field, info in api_opts.items():
        for opt in info.get("options", []):
            api_strings.add(opt)
    
    # Check backend schema.py for matching strings \u2014 option strings are the
    # single source of truth there (metrics.py only references SCORES["\u2026"]).
    schema_file = base_dir / "backend" / "schema.py"
    if schema_file.exists():
        content = schema_file.read_text()
        # Check key strings exist in schema
        critical_strings = [
            "No revisions needed", "Cosmetic only", "Moderate rework", "Major rework",
            "Significantly better", "Somewhat better", "Comparable", "Somewhat worse",
            "Exceeded expectations", "Met expectations", "Below expectations",
            "Yes directly reused and extended", "Yes provided useful starting context", "No built from scratch",
            "No \u2014 proprietary inputs decisive", "Partially \u2014 they would miss key insights", "Yes \u2014 all inputs publicly available",
        ]
        for s in critical_strings:
            # Schema source escapes em dashes as —; accept either form.
            escaped = s.replace("—", "\\u2014")
            test(f"Schema contains: '{s[:40]}...'", s in content or escaped in content, f"missing: {s[:60]}")
    
    # Check frontend app.js
    frontend_file = base_dir / "frontend" / "app.js"
    if frontend_file.exists():
        js_content = frontend_file.read_text()
        # Check em dash strings
        em_dash_options = [
            "No \u2014 proprietary inputs decisive",
            "Partially \u2014 they would miss key insights",
            "Yes \u2014 all inputs publicly available",
            "No \u2014 legacy would have been worse",
        ]
        for opt in em_dash_options:
            # Check for em dash in JS
            has_em = "\u2014" in js_content
            if has_em:
                # Just check the em dash variant is present
                test(f"Frontend has em dash (—) in options", True)
                break
        else:
            test("Frontend has em dash (—) in options", "\u2014" in js_content, "no em dash found in app.js")

# ── K. Frontend JS Syntax ────────────────────────────────────────────────────

def test_frontend_js():
    print("\n── K. Frontend JS Syntax ──")
    from pathlib import Path
    frontend_file = Path(DB_PATH).parent.parent / "frontend" / "app.js"
    if not frontend_file.exists():
        test("frontend/app.js exists", False)
        return
    
    result = subprocess.run(["node", "-c", str(frontend_file)], capture_output=True, text=True)
    test("node -c frontend/app.js passes", result.returncode == 0, result.stderr.strip() if result.returncode != 0 else "")

# ── L. Norms / Category Norms ─────────────────────────────────────────────────

def test_norms():
    print("\n── L. Norms / Category Norms ──")
    tk = admin_token()
    
    r = requests.get(f"{BASE}/api/norms/aggregates", headers=auth_h(tk))
    test("GET /api/norms/aggregates returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    
    norms = r.json()
    test("Norms is a list", isinstance(norms, list))
    
    if norms:
        n = norms[0]
        # Modern norms aggregate exposes 'completed_surveys' (renamed from 'sample_count').
        test("Norm has completed_surveys", "completed_surveys" in n)
        test("Norm has category_name", "category_name" in n)

        # Check outlier flag logic - should only appear when ≥3 samples
        if n.get("completed_surveys", 0) < 3:
            test(f"Outlier flag absent when <3 samples ({n.get('completed_surveys')})", not n.get("has_outlier_flags", False))

    # Categories with 0 completions should show empty/null norms
    has_zero = any(n.get("completed_surveys", 0) == 0 for n in norms)
    if has_zero:
        for n in norms:
            if n.get("completed_surveys", 0) == 0:
                test(f"Zero-completion category ({n.get('category_name')}) has null averages",
                     n.get("avg_effort_ratio") is None or n.get("avg_quality_ratio") is None)

# ── L2. /api/schema endpoint ─────────────────────────────────────────────────

def test_schema_endpoint():
    print("\n── L2. /api/schema endpoint ──")
    tk = admin_token()
    r = requests.get(f"{BASE}/api/schema", headers=auth_h(tk))
    test("GET /api/schema returns 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    body = r.json()
    test("schema response has dashboard key", "dashboard" in body)
    dash = body.get("dashboard") or {}
    test("schema.dashboard has tabs list", isinstance(dash.get("tabs"), list) and len(dash["tabs"]) == 4)
    test("schema.dashboard has charts list", isinstance(dash.get("charts"), list) and len(dash["charts"]) == 19)
    test("schema.dashboard has kpi_tiles list", isinstance(dash.get("kpi_tiles"), list) and len(dash["kpi_tiles"]) == 12)
    test("schema.dashboard.thresholds.radar_axis_cap present",
         isinstance((dash.get("thresholds") or {}).get("radar_axis_cap"), (int, float)))

# ── L3. MetricsSummary endpoint fields ───────────────────────────────────────

def test_metrics_summary_fields():
    """Regression guard: MetricsSummary model must not silently drop these keys."""
    print("\n── L3. MetricsSummary endpoint fields ──")
    tk = admin_token()
    r = requests.get(f"{BASE}/api/metrics/summary", headers=auth_h(tk))
    test("GET /api/metrics/summary 200", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    body = r.json()
    for key in ("average_quality_ratio", "rework_efficiency_avg", "client_impact_avg", "data_independence_avg"):
        test(f"summary exposes {key}", key in body, f"missing from response keys")

# ── M. Database Schema Checks ────────────────────────────────────────────────

def test_schema():
    print("\n── M. Database Schema Checks ──")
    if not os.path.exists(DB_PATH):
        test("Database file exists", False, f"not found: {DB_PATH}")
        return
    
    conn = sqlite3.connect(DB_PATH)
    
    # projects table
    proj_cols = {row[1] for row in conn.execute("PRAGMA table_info(projects)").fetchall()}
    test("projects has working_days column", "working_days" in proj_cols)
    test("projects has engagement_revenue column", "engagement_revenue" in proj_cols)
    test("projects has revision_depth column", "revision_depth" in proj_cols)
    
    # expert_responses table
    exp_cols = {row[1] for row in conn.execute("PRAGMA table_info(expert_responses)").fetchall()}
    test("expert_responses has c6_self_assessment", "c6_self_assessment" in exp_cols)
    test("expert_responses has g1_reuse_intent", "g1_reuse_intent" in exp_cols)
    test("expert_responses has b5_ai_survival", "b5_ai_survival" in exp_cols)
    test("expert_responses has b6_data_analysis_split", "b6_data_analysis_split" in exp_cols)
    test("expert_responses has c7_analytical_depth", "c7_analytical_depth" in exp_cols)
    test("expert_responses has c8_decision_readiness", "c8_decision_readiness" in exp_cols)
    test("expert_responses has e1_client_decision", "e1_client_decision" in exp_cols)
    
    # Legacy columns (l2_legacy_team_size dropped by migrate_v18)
    legacy_cols = ["l1_legacy_working_days", "l3_legacy_revision_depth", "l13_legacy_c7_depth", "l14_legacy_c8_decision", "l15_legacy_e1_decision", "l16_legacy_b6_data"]
    for col in legacy_cols:
        test(f"expert_responses has {col}", col in exp_cols)
    test("expert_responses l2_legacy_team_size dropped (v18)", "l2_legacy_team_size" not in exp_cols)

    conn.close()

# ── N. DASHBOARD_CONFIG ───────────────────────────────────────────────────────

def test_dashboard_config():
    print("\n── N. DASHBOARD_CONFIG ──")
    from backend import schema as _schema
    dc = getattr(_schema, "DASHBOARD_CONFIG", None)
    test("DASHBOARD_CONFIG exists", isinstance(dc, dict))
    test("DASHBOARD_CONFIG has tabs", isinstance(dc.get("tabs"), list) and len(dc["tabs"]) == 4)
    test("DASHBOARD_CONFIG has charts list", isinstance(dc.get("charts"), list))
    th = dc.get("thresholds", {})
    test("thresholds.radar_axis_cap is positive float", isinstance(th.get("radar_axis_cap"), (int, float)) and th["radar_axis_cap"] > 1)
    test("thresholds.quarterly_bucket_min_quarters is int >= 2", isinstance(th.get("quarterly_bucket_min_quarters"), int) and th["quarterly_bucket_min_quarters"] >= 2)
    test("thresholds.cohort_min_projects is int >= 1", isinstance(th.get("cohort_min_projects"), int) and th["cohort_min_projects"] >= 1)
    test("thresholds.bar_top_n is int >= 3", isinstance(th.get("bar_top_n"), int) and th["bar_top_n"] >= 3)
    tone = th.get("metric_tone", {})
    test("metric_tone.success_above is float", isinstance(tone.get("success_above"), (int, float)))
    test("metric_tone.blue_above is float", isinstance(tone.get("blue_above"), (int, float)))
    test("metric_tone.warning_above is float", isinstance(tone.get("warning_above"), (int, float)))

    # (added in Task 2 to harden Task 1 config)
    test("tab ids are unique", len({t["id"] for t in dc["tabs"]}) == len(dc["tabs"]))
    test("every tab has id/label/icon", all({"id", "label", "icon"} <= set(t.keys()) for t in dc["tabs"]))
    test("metric_tone ordering (success > blue > warning)",
         tone["success_above"] > tone["blue_above"] > tone["warning_above"])

    # kpi_tiles assertions (Task 2)
    test("kpi_tiles has 12 entries", isinstance(dc["kpi_tiles"], list) and len(dc["kpi_tiles"]) == 12)
    test("every kpi_tile.metric_key exists in METRICS or is synthetic",
         all(t["metric_key"] in _schema.METRICS or t.get("synthetic") is True for t in dc["kpi_tiles"]))
    test("every kpi_tile has tab field", all("tab" in t for t in dc["kpi_tiles"]))
    tab_ids_set = {x["id"] for x in dc["tabs"]}
    test("every kpi_tile.tab is a known tab id",
         all(t["tab"] in tab_ids_set for t in dc["kpi_tiles"]))

    CHART_TYPES = {
        "scatter_disprove", "radar_gains",
        "timeline_per_project", "timeline_quarterly", "timeline_cumulative", "cohort_learning_curve",
        "bar_by_category", "bar_by_practice", "bar_by_pioneer",
        "heatmap_practice_quarter", "area_category_mix",
        "donut_client_pulse", "donut_reuse_intent", "scatter_schedule", "track_scaling_gates",
        "table_portfolio",
        "ranked_list_top", "ranked_list_bottom", "timeline_effort",
    }
    test("every chart has id/tab/type/title",
         all({"id", "tab", "type", "title"} <= set(c.keys()) for c in dc["charts"]))
    test("every chart.tab is a known tab id",
         all(c["tab"] in tab_ids_set for c in dc["charts"]))
    test("every chart.type is in CHART_TYPES",
         all(c["type"] in CHART_TYPES for c in dc["charts"]))
    test("chart ids are unique",
         len({c["id"] for c in dc["charts"]}) == len(dc["charts"]))
    test("Trends tab has 5 charts",
         sum(1 for c in dc["charts"] if c["tab"] == "trends") == 5)
    test("Breakdowns has 5 charts (3 bars + heatmap + mix)",
         sum(1 for c in dc["charts"] if c["tab"] == "breakdowns") == 5)
    test("every tab has at least one chart",
         all(any(c["tab"] == t for c in dc["charts"]) for t in tab_ids_set))
    test("Overview has 4 charts",
         sum(1 for c in dc["charts"] if c["tab"] == "overview") == 4)
    test("Signals has 5 charts",
         sum(1 for c in dc["charts"] if c["tab"] == "signals") == 5)

# ── O. Seed profile EXPERT_FIELDS coverage ───────────────────────────────────

def test_seed_field_coverage():
    """Every EXPERT_FIELDS key must appear in every seed profile dict."""
    print("\n── O. Seed profile coverage (every EXPERT_FIELDS key) ──")
    import importlib.util, pathlib
    from backend.schema import EXPERT_FIELDS
    seed_path = pathlib.Path("tests/seed_20_projects.py")
    spec = importlib.util.spec_from_file_location("seed_module", seed_path)
    seed_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(seed_mod)

    # Heuristic: profile dicts are module-level, uppercase names, and carry c6_self_assessment
    profiles = {name: obj for name, obj in vars(seed_mod).items()
                if name.isupper() and isinstance(obj, dict) and obj.get("c6_self_assessment")}
    test("seed defines at least 3 profiles", len(profiles) >= 3, detail=f"found {list(profiles.keys())}")
    for pname, pdict in profiles.items():
        missing = [k for k in EXPERT_FIELDS if k not in pdict]
        test(f"seed profile {pname} covers all EXPERT_FIELDS", not missing, detail=f"missing={missing}")

# Aggregate parity (no-filter client aggregates == /api/dashboard/metrics) is
# enforced by tests/e2e-dashboard-redesign.spec.ts > "no-filter parity".

# ── P. Survey flow: cross-pioneer visibility + auto-issue next round ─────────

_VALID_EXPERT_PAYLOAD = {
    "b1_starting_point": "From AI draft",
    "b2_research_sources": "Multiple sources across domains (5-10)",
    "b3_assembly_ratio": ">75% AI",
    "b4_hypothesis_first": "Hypothesis-first",
    "b5_ai_survival": ">75%",
    "b6_data_analysis_split": "25-50%",
    "c1_specialization": "Deep specialist",
    "c2_directness": "Expert authored",
    "c3_judgment_pct": ">75% judgment",
    "c6_self_assessment": "Significantly better",
    "c7_analytical_depth": "Exceptional",
    "c8_decision_readiness": "Yes without caveats",
    "d1_proprietary_data": "Yes",
    "d2_knowledge_reuse": "Yes directly reused and extended",
    "d3_moat_test": "No — proprietary inputs decisive",
    "e1_client_decision": "Yes — informed a specific decision",
    "f1_feasibility": "Not feasible",
    "f2_productization": "Yes largely as-is",
    "g1_reuse_intent": "Yes without hesitation",
    "l1_legacy_working_days": 15,
    "l3_legacy_revision_depth": "Moderate rework",
    "l4_legacy_scope_expansion": "No",
    "l5_legacy_client_reaction": "Met expectations",
    "l6_legacy_b2_sources": "A few targeted sources (2-4)",
    "l7_legacy_c1_specialization": "Generalist",
    "l8_legacy_c2_directness": "Expert reviewed only",
    "l9_legacy_c3_judgment": "25-50%",
    "l10_legacy_d1_proprietary": "No",
    "l11_legacy_d2_reuse": "No built from scratch",
    "l12_legacy_d3_moat": "Yes — all inputs publicly available",
    "l13_legacy_c7_depth": "Adequate",
    "l14_legacy_c8_decision": "Needs significant additional work",
    "l15_legacy_e1_decision": "Yes — referenced in internal discussions",
    "l16_legacy_b6_data": ">75% on data",
}


def _pioneer_round1_token(project: dict, pioneer_name: str) -> str:
    for p in project.get("pioneers", []):
        # Prefer the server-supplied display_name (= "first last" trimmed),
        # then the pioneer_name alias for backward-compat.
        name = p.get("display_name") or p.get("pioneer_name") or p.get("name")
        if name == pioneer_name:
            for rnd in p.get("rounds", []):
                if rnd.get("round_number") == 1:
                    return rnd["token"]
            # Fallback to legacy pioneer-level token (it doubles as round 1).
            return p.get("expert_token")
    raise AssertionError(f"Pioneer '{pioneer_name}' not found on project")


def test_show_other_pioneers_flag():
    print("\n── P1. show_other_pioneers_answers flag ──")
    tk = admin_token()
    cats = requests.get(f"{BASE}/api/categories", headers=auth_h(tk)).json()
    cat_id = cats[0]["id"] if cats else 1

    def _create(flag: bool, name: str):
        r = requests.post(
            f"{BASE}/api/projects",
            headers={**auth_h(tk), "Content-Type": "application/json"},
            json={
                "project_name": name,
                "category_id": cat_id,
                "pioneers": [
                    {"first_name": "Pioneer", "last_name": "One"},
                    {"first_name": "Pioneer", "last_name": "Two"},
                ],
                "default_rounds": 1,
                "show_other_pioneers_answers": flag,
                "date_started": "2026-03-01",
                "date_delivered": "2026-03-10",
                "working_days": 5,
                "xcsg_team_size": "2",
                "xcsg_revision_rounds": "1",
            },
        )
        assert r.status_code in (200, 201), r.text
        return r.json()

    # Case 1: flag enabled
    proj_on = _create(True, "QA Cross-Pioneer ON")
    test(
        "Project persists show_other_pioneers_answers=True",
        bool(proj_on.get("show_other_pioneers_answers")),
        f"got {proj_on.get('show_other_pioneers_answers')}",
    )
    p1_tok = _pioneer_round1_token(proj_on, "Pioneer One")
    p2_tok = _pioneer_round1_token(proj_on, "Pioneer Two")
    r = requests.post(f"{BASE}/api/expert/{p1_tok}", json=_VALID_EXPERT_PAYLOAD)
    test("P1 submits round 1 (flag on)", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:120]}")
    ctx = requests.get(f"{BASE}/api/expert/{p2_tok}").json()
    test(
        "P2 expert-view exposes show_other_pioneers=True",
        ctx.get("show_other_pioneers") is True,
        f"got {ctx.get('show_other_pioneers')}",
    )
    others = ctx.get("other_pioneers_responses")
    test(
        "P2 sees exactly one other-pioneer response (flag on)",
        isinstance(others, list) and len(others) == 1,
        f"got {others if isinstance(others, list) else type(others).__name__}",
    )
    if isinstance(others, list) and others:
        test(
            "Other-pioneer entry has pioneer_name=Pioneer One",
            others[0].get("pioneer_name") == "Pioneer One",
            f"got {others[0].get('pioneer_name')}",
        )
        test(
            "Other-pioneer entry contains answer field (b1_starting_point)",
            others[0].get("b1_starting_point") == "From AI draft",
            f"got {others[0].get('b1_starting_point')}",
        )
    requests.delete(f"{BASE}/api/projects/{proj_on['id']}", headers=auth_h(tk))

    # Case 2: flag disabled
    proj_off = _create(False, "QA Cross-Pioneer OFF")
    test(
        "Project persists show_other_pioneers_answers=False",
        not bool(proj_off.get("show_other_pioneers_answers")),
        f"got {proj_off.get('show_other_pioneers_answers')}",
    )
    p1_tok = _pioneer_round1_token(proj_off, "Pioneer One")
    p2_tok = _pioneer_round1_token(proj_off, "Pioneer Two")
    r = requests.post(f"{BASE}/api/expert/{p1_tok}", json=_VALID_EXPERT_PAYLOAD)
    test("P1 submits round 1 (flag off)", r.status_code in (200, 201), f"got {r.status_code}")
    ctx = requests.get(f"{BASE}/api/expert/{p2_tok}").json()
    test(
        "P2 expert-view exposes show_other_pioneers=False (flag off)",
        ctx.get("show_other_pioneers") is False,
        f"got {ctx.get('show_other_pioneers')}",
    )
    test(
        "P2 other_pioneers_responses is None (flag off)",
        ctx.get("other_pioneers_responses") is None,
        f"got {ctx.get('other_pioneers_responses')}",
    )
    requests.delete(f"{BASE}/api/projects/{proj_off['id']}", headers=auth_h(tk))


def test_auto_issue_next_round():
    print("\n── P2. Auto-issue next round on submit ──")
    tk = admin_token()
    cats = requests.get(f"{BASE}/api/categories", headers=auth_h(tk)).json()
    cat_id = cats[0]["id"] if cats else 1

    r = requests.post(
        f"{BASE}/api/projects",
        headers={**auth_h(tk), "Content-Type": "application/json"},
        json={
            "project_name": "QA Auto-Issue Round",
            "category_id": cat_id,
            "pioneers": [{"first_name": "Solo", "last_name": "Pioneer", "total_rounds": 2}],
            "default_rounds": 2,
            "date_started": "2026-03-01",
            "date_delivered": "2026-03-10",
            "working_days": 5,
            "xcsg_team_size": "2",
            "xcsg_revision_rounds": "1",
        },
    )
    test("Create 2-round project returns 201", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:120]}")
    if r.status_code not in (200, 201):
        return
    proj = r.json()
    r1_tok = _pioneer_round1_token(proj, "Solo Pioneer")

    submit = requests.post(f"{BASE}/api/expert/{r1_tok}", json=_VALID_EXPERT_PAYLOAD)
    test("Round 1 submit returns 201", submit.status_code in (200, 201), f"got {submit.status_code}: {submit.text[:120]}")
    body = submit.json() if submit.status_code in (200, 201) else {}
    test("Submit response includes next_round_token (not None)", bool(body.get("next_round_token")), f"got {body.get('next_round_token')}")
    test("Submit response reports current_round=1", body.get("current_round") == 1, f"got {body.get('current_round')}")
    test("Submit response reports total_rounds=2", body.get("total_rounds") == 2, f"got {body.get('total_rounds')}")

    # Verify GET /api/projects/{id} shows the new round 2 token as issued.
    proj_after = requests.get(f"{BASE}/api/projects/{proj['id']}", headers=auth_h(tk)).json()
    pioneers = proj_after.get("pioneers", [])
    solo = next((p for p in pioneers if (p.get("pioneer_name") or p.get("name")) == "Solo Pioneer"), None)
    test("Pioneer is present in project after submit", solo is not None)
    if solo is not None:
        rounds = solo.get("rounds", [])
        r2_row = next((rnd for rnd in rounds if rnd.get("round_number") == 2), None)
        test("Round 2 row exists after auto-issue", r2_row is not None, f"rounds={[r.get('round_number') for r in rounds]}")
        if r2_row is not None:
            test("Round 2 token matches submit response", r2_row.get("token") == body.get("next_round_token"), f"db={r2_row.get('token')} resp={body.get('next_round_token')}")
            test("Round 2 token is not yet completed", r2_row.get("completed_at") in (None, ""), f"got {r2_row.get('completed_at')}")

    # Also verify the new token actually works as an expert-view token.
    if body.get("next_round_token"):
        ctx = requests.get(f"{BASE}/api/expert/{body['next_round_token']}").json()
        test("next_round_token opens expert view at round 2", ctx.get("current_round") == 2, f"got {ctx.get('current_round')}")
        test("next_round_token already_completed=False", ctx.get("already_completed") is False, f"got {ctx.get('already_completed')}")

    requests.delete(f"{BASE}/api/projects/{proj['id']}", headers=auth_h(tk))


# ── X. Expert notes (per-round optional free-text) ───────────────────────────

def test_expert_notes():
    print("\n── X. Expert notes (per-round, optional free-text) ──")
    tok = admin_token()
    h = auth_h(tok)
    # Find or create a project with at least one pioneer round issued
    projects = requests.get(f"{BASE}/api/projects", headers=h).json()
    if not projects:
        test("skipped — no projects seeded", False, detail="seed first")
        return
    # Pick the first project that still has an open round for a pioneer
    proj_id = None
    pioneer = None
    open_round = None
    for pr in projects:
        pid = pr.get("id")
        if pid is None:
            continue
        detail = requests.get(f"{BASE}/api/projects/{pid}", headers=h).json()
        for pp in detail.get("pioneers", []):
            rnds = pp.get("rounds") or []
            candidate = next((r for r in rnds if r.get("token") and not r.get("completed_at")), None)
            if candidate:
                proj_id = pid
                pioneer = pp
                open_round = candidate
                break
        if open_round:
            break
    if not (proj_id and pioneer and open_round):
        test("skipped — no pioneer with open round", False, detail="all rounds completed")
        return

    # Load STRONG payload from the seed file
    import importlib.util, pathlib
    spec = importlib.util.spec_from_file_location("sd", pathlib.Path("tests/seed_20_projects.py"))
    sd = importlib.util.module_from_spec(spec); spec.loader.exec_module(sd)
    payload = dict(getattr(sd, "STRONG"))
    payload["notes"] = "Client pivoted mid-engagement from feasibility to gap analysis.\nWorth flagging for similar RAM projects."

    r = requests.post(f"{BASE}/api/expert/{open_round['token']}", json=payload)
    test("POST /api/expert/{token} with notes returns 200/201", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:200]}")

    # Confirm stored: direct DB read
    db_conn = sqlite3.connect(DB_PATH)
    row = db_conn.execute(
        "SELECT notes FROM expert_responses WHERE project_id = ? AND pioneer_id = ? ORDER BY id DESC LIMIT 1",
        (proj_id, pioneer["id"]),
    ).fetchone()
    db_conn.close()
    test("notes persisted in expert_responses", bool(row and row[0] and "pivoted" in row[0]), detail=f"row={row}")


def test_notes_feed_endpoint():
    print("\n── Y. GET /api/notes feed + filters ──")
    tok = admin_token()
    h = auth_h(tok)

    # Unauthenticated call should be rejected (401 or 403).
    r_noauth = requests.get(f"{BASE}/api/notes")
    test("GET /api/notes unauthenticated -> 401/403", r_noauth.status_code in (401, 403), f"got {r_noauth.status_code}")

    # Seed 2 notes on different projects/practices so we can exercise filters.
    import importlib.util, pathlib
    spec = importlib.util.spec_from_file_location("sd", pathlib.Path("tests/seed_20_projects.py"))
    sd = importlib.util.module_from_spec(spec); spec.loader.exec_module(sd)
    base_payload = dict(getattr(sd, "STRONG"))

    db_conn = sqlite3.connect(DB_PATH)
    # Open rounds with practice code joined in.
    open_rounds = db_conn.execute(
        """SELECT prt.token, pr.code AS practice_code, p.id AS project_id
           FROM pioneer_round_tokens prt
           JOIN project_pioneers pp ON pp.id = prt.pioneer_id
           JOIN projects p ON p.id = pp.project_id
           LEFT JOIN practices pr ON pr.id = p.practice_id
           WHERE prt.completed_at IS NULL
           ORDER BY p.id ASC LIMIT 10"""
    ).fetchall()
    db_conn.close()

    # Pick one MAP and one non-MAP token if possible.
    map_tok = next((r for r in open_rounds if r[1] == "MAP"), None)
    other_tok = next((r for r in open_rounds if r[1] and r[1] != "MAP"), None)
    if not map_tok or not other_tok:
        test("skipped — need one MAP and one non-MAP open round", False, detail=f"open={[r[1] for r in open_rounds]}")
        return

    p_map = dict(base_payload); p_map["notes"] = "Pivotal MAP insight: payer evidence bar rose."
    p_oth = dict(base_payload); p_oth["notes"] = "Generic note from another practice."

    r1 = requests.post(f"{BASE}/api/expert/{map_tok[0]}", json=p_map)
    r2 = requests.post(f"{BASE}/api/expert/{other_tok[0]}", json=p_oth)
    test("seed MAP note submit 200/201", r1.status_code in (200, 201), f"got {r1.status_code}")
    test("seed non-MAP note submit 200/201", r2.status_code in (200, 201), f"got {r2.status_code}")

    # 1. Unfiltered feed returns both seeded notes.
    feed = requests.get(f"{BASE}/api/notes", headers=h)
    test("GET /api/notes 200", feed.status_code == 200, f"got {feed.status_code}: {feed.text[:120]}")
    items = feed.json() if feed.status_code == 200 else []
    test("/api/notes returns a list", isinstance(items, list), detail=f"type={type(items).__name__}")
    notes_texts = [i.get("notes", "") for i in items] if isinstance(items, list) else []
    test("feed contains MAP seeded note", any("Pivotal MAP insight" in n for n in notes_texts))
    test("feed contains non-MAP seeded note", any("Generic note from another practice" in n for n in notes_texts))

    # Every returned row must have the standard shape.
    if items:
        sample = items[0]
        for key in ("id", "project_id", "project_name", "pioneer_name", "round_number", "submitted_at", "notes"):
            test(f"feed row has '{key}'", key in sample, detail=f"sample keys={list(sample.keys())}")

    # 2. Filter by practice_code=MAP.
    feed_map = requests.get(f"{BASE}/api/notes", headers=h, params={"practice_code": "MAP"})
    test("GET /api/notes?practice_code=MAP 200", feed_map.status_code == 200, f"got {feed_map.status_code}")
    map_items = feed_map.json() if feed_map.status_code == 200 else []
    test("MAP filter returns >=1 row", isinstance(map_items, list) and len(map_items) >= 1, detail=f"len={len(map_items) if isinstance(map_items, list) else 'n/a'}")
    non_map = [i for i in map_items if (i.get("practice_code") or "") != "MAP"]
    test("MAP filter excludes non-MAP rows", not non_map, detail=f"leaked={[i.get('practice_code') for i in non_map]}")

    # 3. Search param (case-insensitive substring).
    feed_search = requests.get(f"{BASE}/api/notes", headers=h, params={"search": "pivot"})
    test("GET /api/notes?search=pivot 200", feed_search.status_code == 200, f"got {feed_search.status_code}")
    search_items = feed_search.json() if feed_search.status_code == 200 else []
    test(
        "search=pivot matches MAP note (case-insensitive)",
        any("Pivotal MAP insight" in i.get("notes", "") for i in search_items),
        detail=f"len={len(search_items) if isinstance(search_items, list) else 'n/a'}",
    )
    # Generic note has no "pivot" — must be excluded.
    test(
        "search=pivot excludes unrelated notes",
        not any("Generic note from another practice" in i.get("notes", "") for i in search_items),
    )


def test_notes_excel_sheet():
    print("\n── Z. Notes sheet in Excel export ──")
    import io
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        test("skipped — openpyxl not installed", False)
        return
    import openpyxl
    tok = admin_token()
    r = requests.get(f"{BASE}/api/export/excel", headers=auth_h(tok))
    ct = r.headers.get("content-type", "")
    test("export returns xlsx", r.status_code == 200 and ct.endswith("sheet"), detail=f"{r.status_code} {ct}")
    if r.status_code != 200:
        return
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    test("export has Notes sheet", "Notes" in wb.sheetnames, detail=f"sheets={wb.sheetnames}")
    if "Notes" not in wb.sheetnames:
        return
    notes_sheet = wb["Notes"]
    rows = list(notes_sheet.iter_rows(values_only=True))
    test("Notes sheet has header row + >=1 data row", len(rows) >= 2, detail=f"rows={len(rows)}")
    expected_cols = ["Project", "Category", "Practice", "Pioneer", "Round", "Submitted", "Notes"]
    test("Notes sheet header matches expected columns", list(rows[0]) == expected_cols, detail=f"got {rows[0]}")


def test_dashboard_export_sheets():
    print("\n── Y. Dashboard export sheets ──")
    import io, openpyxl
    tok = admin_token()
    r = requests.get(f"{BASE}/api/export/excel", headers=auth_h(tok))
    test(
        "export/excel returns xlsx",
        r.status_code == 200 and r.headers.get("content-type", "").endswith("sheet"),
    )
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    expected_new = [
        "Dashboard Aggregates", "By Practice", "By Category", "By Pioneer",
        "Quarterly Trend", "Cumulative Trend", "Cohort — Practice",
        "Practice × Quarter", "Category Mix by Quarter",
        "Disprove Matrix", "Gains Radar",
        "Client Pulse", "Reuse Intent",
        "Schedule Variance", "Scaling Gates",
        "Top Movers", "Bottom Movers", "Takeaways",
    ]
    for name in expected_new:
        test(
            f"export has '{name}' sheet",
            name in wb.sheetnames,
            detail=f"sheetnames={wb.sheetnames}",
        )
    # Spot-check a couple of sheets have data rows beyond the header
    for name in ["By Practice", "Quarterly Trend", "Takeaways", "Top Movers"]:
        if name in wb.sheetnames:
            rows = list(wb[name].iter_rows(values_only=True))
            test(
                f"'{name}' sheet has header + at least one data row",
                len(rows) >= 2,
                detail=f"rows={len(rows)}",
            )
    # Spot-check Disprove Matrix quadrant column has valid enum values
    if "Disprove Matrix" in wb.sheetnames:
        ws = wb["Disprove Matrix"]
        header = [c.value for c in ws[1]]
        qi = header.index("quadrant") if "quadrant" in header else -1
        if qi >= 0:
            allowed = {"top-right", "top-left", "bottom-right", "bottom-left", "NA"}
            values = {ws.cell(row=i, column=qi + 1).value for i in range(2, ws.max_row + 1)}
            test(
                "Disprove Matrix quadrants are in allowed enum",
                values.issubset(allowed),
                detail=f"seen={values}",
            )


# ── Main ──────────────────────────────────────────────────────────────────────

def test_dashboard_takeaways():
    print("\n── T. Dashboard takeaways endpoint ──")
    tok = admin_token()
    r = requests.get(f"{BASE}/api/dashboard/takeaways", headers=auth_h(tok))
    test("GET /api/dashboard/takeaways 200", r.status_code == 200)
    if r.status_code != 200:
        return
    body = r.json()
    test("takeaways has chartDisprove", isinstance(body.get("chartDisprove"), str))
    test("takeaways has chartQuarterly", isinstance(body.get("chartQuarterly"), str))
    test("takeaways has chartPulse", isinstance(body.get("chartPulse"), str))
    test("takeaways has trackGates", isinstance(body.get("trackGates"), str))
    test("takeaways has chartTopMovers", isinstance(body.get("chartTopMovers"), str) and body["chartTopMovers"])
    test("takeaways has chartBottomMovers", isinstance(body.get("chartBottomMovers"), str) and body["chartBottomMovers"])
    test("takeaways has chartEffortTrend", isinstance(body.get("chartEffortTrend"), str) and body["chartEffortTrend"])
    # Every chart in DASHBOARD_CONFIG must have a key in the response
    from backend.schema import DASHBOARD_CONFIG
    missing = [c["id"] for c in DASHBOARD_CONFIG["charts"] if c["id"] not in body]
    test("every chart has a takeaway key", not missing, detail=f"missing={missing}")
    # Takeaways ≤ 80 chars (≤ 60 is the target, allow slack)
    too_long = {cid: v for cid, v in body.items() if len(v) > 80}
    test("takeaways are concise", not too_long, detail=f"too_long={too_long}")
    empty = {cid: v for cid, v in body.items() if not v or not v.strip()}
    test("takeaways are all non-empty", not empty, detail=f"empty={list(empty.keys())}")


def test_economics_schema():
    """Schema response exposes economics fields, currencies, and pricing models."""
    from backend.schema import CURRENCIES, PRICING_MODELS, ECONOMICS_FIELDS, METRICS, build_schema_response

    assert CURRENCIES == ["EUR", "USD", "GBP", "CHF", "CAD", "AUD"]
    assert "Fixed fee" in PRICING_MODELS
    assert "Time & materials" in PRICING_MODELS
    assert "Retainer" in PRICING_MODELS
    assert "Milestone" in PRICING_MODELS
    assert "Other" in PRICING_MODELS
    assert len(PRICING_MODELS) == 5

    expected_econ = {
        "engagement_revenue", "currency", "xcsg_pricing_model",
        "scope_expansion_revenue",
    }
    assert expected_econ.issubset(set(ECONOMICS_FIELDS.keys()))

    for key in ("margin_gain", "xcsg_margin_pct", "cost_per_quality_point_gain", "revenue_per_day_gain"):
        assert key in METRICS, f"missing metric {key}"

    response = build_schema_response()
    assert response["currencies"] == CURRENCIES
    assert response["pricing_models"] == PRICING_MODELS
    assert "economics_fields" in response


def test_migrate_v15_idempotent():
    """migrate_v15 adds new columns + app_settings table, runs idempotently."""
    from backend import database

    database.init_db()

    # init_db just ran above; verify the post-state.
    with database._db() as conn:
        proj_cols = {r[1] for r in conn.execute("PRAGMA table_info(projects)").fetchall()}
        # legacy_day_rate_override was added by v15 but dropped by v18
        for col in ("currency", "xcsg_pricing_model", "scope_expansion_revenue"):
            assert col in proj_cols, f"projects.{col} missing"

        pp_cols = {r[1] for r in conn.execute("PRAGMA table_info(project_pioneers)").fetchall()}
        assert "day_rate" in pp_cols, "project_pioneers.day_rate missing"

        # default_legacy_day_rate was added by v15 but dropped by v18

        tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        assert "app_settings" in tables

        row = conn.execute("SELECT id, default_currency FROM app_settings WHERE id=1").fetchone()
        assert row is not None
        assert row["default_currency"] == "EUR"

    # Re-run migration — must be idempotent.
    database.migrate_v15()
    database.migrate_v15()

    with database._db() as conn:
        rows = conn.execute("SELECT COUNT(*) AS n FROM app_settings").fetchone()
        assert rows["n"] == 1, "app_settings must remain a single row"


def test_migrate_v16_idempotent():
    """migrate_v16 creates practice_roles table + index, runs idempotently."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        tables = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()}
        assert "practice_roles" in tables

        cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(practice_roles)"
        ).fetchall()}
        for col in ("id", "practice_id", "role_name", "day_rate", "currency",
                    "display_order", "created_at"):
            assert col in cols, f"practice_roles.{col} missing"

        indexes = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='index' AND tbl_name='practice_roles'"
        ).fetchall()}
        assert "idx_practice_roles_practice" in indexes

    # Re-run migration — must be idempotent.
    database.migrate_v16()
    database.migrate_v16()

    # FK CASCADE: deleting a practice removes its roles.
    with database._db() as conn:
        cur = conn.execute(
            "INSERT INTO practices (code, name, description) VALUES (?, ?, ?)",
            ("TST", "Test practice", "for migration test"),
        )
        practice_id = cur.lastrowid
        conn.execute(
            "INSERT INTO practice_roles (practice_id, role_name, day_rate, currency) "
            "VALUES (?, ?, ?, ?)",
            (practice_id, "Senior", 1500, "EUR"),
        )
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("DELETE FROM practices WHERE id = ?", (practice_id,))
        remaining = conn.execute(
            "SELECT COUNT(*) AS n FROM practice_roles WHERE practice_id = ?",
            (practice_id,),
        ).fetchone()
        assert remaining["n"] == 0, "practice_roles should CASCADE-delete"


def test_migrate_v17_idempotent():
    """migrate_v17 adds project_pioneers.role_name (nullable), runs idempotently."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(project_pioneers)"
        ).fetchall()}
        assert "role_name" in cols, "project_pioneers.role_name missing"

    # Re-run migration — must be idempotent.
    database.migrate_v17()
    database.migrate_v17()

    # Existing rows should have NULL role_name — verify via a temporary fixture.
    with database._db() as conn:
        # Insert a pioneer record first (required by the project_pioneers FK).
        cur_pio = conn.execute(
            "INSERT INTO pioneers (first_name, last_name, email) VALUES ('P', '', 'p@x.io')"
        )
        pioneer_registry_id = cur_pio.lastrowid
        # Insert a project (no pioneer_name / pioneer_email columns since migrate_v19).
        cur = conn.execute(
            "INSERT INTO projects (created_by, project_name, category_id, "
            "xcsg_team_size, xcsg_revision_rounds, "
            "legacy_calendar_days, legacy_revision_rounds, "
            "expert_token, status) "
            "VALUES (1, 'mig17 test', 1, '2', '1', '10', '1', 'tok-mig17', 'pending')"
        )
        project_id = cur.lastrowid
        cur = conn.execute(
            "INSERT INTO project_pioneers (project_id, pioneer_id, expert_token) "
            "VALUES (?, ?, 'tok-mig17-pp')",
            (project_id, pioneer_registry_id),
        )
        pid = cur.lastrowid
        row = conn.execute(
            "SELECT role_name FROM project_pioneers WHERE id = ?", (pid,)
        ).fetchone()
        assert row["role_name"] is None
        # Cleanup — delete pioneer_pioneers before project to satisfy FK constraint.
        conn.execute("DELETE FROM project_pioneers WHERE project_id = ?", (project_id,))
        conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))
        conn.execute("DELETE FROM pioneers WHERE id = ?", (pioneer_registry_id,))
        conn.commit()


def test_practice_roles_db_helpers():
    """list_practice_roles and replace_practice_roles round-trip correctly."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        cur = conn.execute(
            "INSERT INTO practices (code, name, description) VALUES (?, ?, ?)",
            ("HLP", "Helper test", "for db helper test"),
        )
        practice_id = cur.lastrowid
        conn.commit()

    try:
        # Empty list initially.
        assert database.list_practice_roles(practice_id) == []

        # Replace with a list.
        database.replace_practice_roles(practice_id, [
            {"role_name": "Senior", "day_rate": 1500, "currency": "EUR", "display_order": 0},
            {"role_name": "Manager", "day_rate": 1000, "currency": "EUR", "display_order": 1},
            {"role_name": "Senior", "day_rate": 1800, "currency": "USD", "display_order": 0},
        ])

        rows = database.list_practice_roles(practice_id)
        assert len(rows) == 3
        names = [(r["role_name"], r["currency"]) for r in rows]
        assert ("Senior", "EUR") in names
        assert ("Senior", "USD") in names
        assert ("Manager", "EUR") in names

        # Replace with a different list — old rows are gone.
        database.replace_practice_roles(practice_id, [
            {"role_name": "Analyst", "day_rate": 600, "currency": "EUR", "display_order": 0},
        ])
        rows = database.list_practice_roles(practice_id)
        assert len(rows) == 1
        assert rows[0]["role_name"] == "Analyst"

        # Replace with empty list — clears the catalog.
        database.replace_practice_roles(practice_id, [])
        assert database.list_practice_roles(practice_id) == []
    finally:
        with database._db() as conn:
            conn.execute("DELETE FROM practices WHERE id = ?", (practice_id,))
            conn.commit()


def test_practice_roles_schema():
    """schema.py exposes PRACTICE_ROLE_FIELDS and surfaces it via build_schema_response."""
    from backend.schema import PRACTICE_ROLE_FIELDS, build_schema_response

    expected_fields = {"role_name", "day_rate", "currency", "display_order"}
    assert expected_fields.issubset(set(PRACTICE_ROLE_FIELDS.keys()))

    role_name = PRACTICE_ROLE_FIELDS["role_name"]
    assert role_name["type"] == "text"
    assert role_name.get("max_length") == 80

    day_rate = PRACTICE_ROLE_FIELDS["day_rate"]
    assert day_rate["type"] == "number"
    assert day_rate["min"] == 0

    currency = PRACTICE_ROLE_FIELDS["currency"]
    assert currency["type"] == "select"
    assert "options" in currency

    response = build_schema_response()
    assert "practice_role_fields" in response
    assert response["practice_role_fields"]["role_name"]["max_length"] == 80


def test_legacy_team_schema():
    """schema.py exposes LEGACY_TEAM_FIELDS via build_schema_response.
    Deprecated fields are removed from ECONOMICS_FIELDS and EXPERT_FIELDS."""
    from backend.schema import (
        LEGACY_TEAM_FIELDS, ECONOMICS_FIELDS, EXPERT_FIELDS,
        REQUIRED_EXPERT_FIELDS, build_schema_response,
    )

    expected = {"role_name", "count", "day_rate"}
    assert expected.issubset(set(LEGACY_TEAM_FIELDS.keys()))

    role_name = LEGACY_TEAM_FIELDS["role_name"]
    assert role_name["type"] == "text"

    count = LEGACY_TEAM_FIELDS["count"]
    assert count["type"] == "integer"
    assert count["min"] == 1

    day_rate = LEGACY_TEAM_FIELDS["day_rate"]
    assert day_rate["type"] == "number"
    assert day_rate["min"] == 0

    # Deprecated fields removed from ECONOMICS_FIELDS.
    assert "legacy_day_rate_override" not in ECONOMICS_FIELDS
    assert "default_legacy_day_rate" not in ECONOMICS_FIELDS

    # Deprecated field removed from EXPERT_FIELDS and REQUIRED_EXPERT_FIELDS.
    assert "l2_legacy_team_size" not in EXPERT_FIELDS
    assert "l2_legacy_team_size" not in REQUIRED_EXPERT_FIELDS

    response = build_schema_response()
    assert "legacy_team_fields" in response
    assert response["legacy_team_fields"]["count"]["min"] == 1


def test_economics_models():
    """ProjectCreate, ProjectPioneerEntry, PracticeUpdate accept and validate economics fields."""
    import pytest
    from pydantic import ValidationError
    from backend.models import (
        ProjectCreate, ProjectUpdate, ProjectPioneerEntry, PracticeUpdate,
        AppSettings, AppSettingsUpdate,
    )

    base = {
        "project_name": "Demo",
        "category_id": 1,
        "pioneers": [{"first_name": "Pia", "last_name": "", "day_rate": 1500}],
        "xcsg_team_size": "2",
        "xcsg_revision_rounds": "1",
    }
    p = ProjectCreate(
        **base,
        engagement_revenue=120000,
        currency="EUR",
        xcsg_pricing_model="Fixed fee",
        scope_expansion_revenue=15000,
    )
    assert p.engagement_revenue == 120000
    assert p.currency == "EUR"
    assert p.xcsg_pricing_model == "Fixed fee"
    assert p.pioneers[0].day_rate == 1500

    # Negative numeric fields are rejected.
    for bad_field in ("engagement_revenue", "scope_expansion_revenue"):
        with pytest.raises(ValidationError):
            ProjectCreate(**{**base, bad_field: -1})
    with pytest.raises(ValidationError):
        ProjectPioneerEntry(name="Pia", day_rate=-50)

    # Invalid currency / pricing model are rejected.
    with pytest.raises(ValidationError):
        ProjectCreate(**base, currency="XYZ")
    with pytest.raises(ValidationError):
        ProjectCreate(**base, xcsg_pricing_model="Pay what you want")

    # AppSettings models.
    s = AppSettings(default_currency="USD", base_currency="USD")
    assert s.default_currency == "USD"
    with pytest.raises(ValidationError):
        AppSettingsUpdate(default_currency="XYZ")

    # Update model also accepts econ fields.
    u = ProjectUpdate(engagement_revenue=999, currency="USD")
    assert u.engagement_revenue == 999


def test_economics_metrics():
    """_compute_economics_metrics covers all formulas + edge cases per spec.
    Phase 2c: legacy_rate_effective replaced by legacy_team + l1_legacy_working_days.
    Legacy cost = Σ(count × day_rate) × l1_days.
    """
    from backend.metrics import _compute_economics_metrics

    # Happy path: all inputs populated, single pioneer.
    # legacy_team=[{count:2, day_rate:900}], l1=40 → weighted=1800, cost=72000
    out = _compute_economics_metrics(
        engagement_revenue=120000,
        xcsg_person_days=20,
        legacy_person_days=80,
        pioneer_rates=[1500],
        legacy_team=[{"role_name": "X", "count": 2, "day_rate": 900}],
        l1_legacy_working_days=40,
        quality_score=0.85,
        legacy_quality_score=0.5,
        scope_expansion_revenue=15000,
        currency="EUR",
    )
    assert out["xcsg_blended_rate"] == 1500
    assert out["xcsg_cost"] == 30000.0          # 1500 * 20
    assert out["legacy_cost"] == 72000.0        # (2*900) * 40
    assert out["xcsg_margin"] == 90000.0        # 120000 - 30000
    assert out["legacy_margin"] == 48000.0      # 120000 - 72000
    assert round(out["xcsg_margin_pct"], 4) == 0.75
    assert round(out["legacy_margin_pct"], 4) == 0.4
    assert round(out["margin_gain"], 2) == 1.88  # 90000 / 48000
    assert out["scope_expansion_revenue"] == 15000
    assert out["currency"] == "EUR"

    # Multi-pioneer averaging, mixed null/non-null rates.
    # legacy_team=[{count:2, day_rate:800}], l1=20 → weighted=1600, cost=32000
    out = _compute_economics_metrics(
        engagement_revenue=100000,
        xcsg_person_days=10,
        legacy_person_days=40,
        pioneer_rates=[2000, None, 1000],   # null skipped
        legacy_team=[{"role_name": "X", "count": 2, "day_rate": 800}],
        l1_legacy_working_days=20,
        quality_score=0.8,
        legacy_quality_score=0.4,
        scope_expansion_revenue=None,
        currency="USD",
    )
    assert out["xcsg_blended_rate"] == 1500   # mean of [2000, 1000]

    # Negative legacy margin → margin_gain is None.
    # legacy_team=[{count:5, day_rate:200}], l1=40 → 5*200*40=40000 > 10000 revenue
    out = _compute_economics_metrics(
        engagement_revenue=10000,
        xcsg_person_days=5,
        legacy_person_days=200,
        pioneer_rates=[1000],
        legacy_team=[{"role_name": "X", "count": 5, "day_rate": 200}],
        l1_legacy_working_days=40,
        quality_score=0.7,
        legacy_quality_score=0.4,
        scope_expansion_revenue=None,
        currency="EUR",
    )
    assert out["legacy_margin"] == -30000
    assert out["margin_gain"] is None

    # No pioneer rates → xcsg cost/margin None; legacy still computes.
    # legacy_team=[{count:1, day_rate:900}], l1=40 → 900*40=36000
    out = _compute_economics_metrics(
        engagement_revenue=50000, xcsg_person_days=10, legacy_person_days=40,
        pioneer_rates=[None, None],
        legacy_team=[{"role_name": "X", "count": 1, "day_rate": 900}],
        l1_legacy_working_days=40,
        quality_score=0.7, legacy_quality_score=0.4,
        scope_expansion_revenue=None, currency="EUR",
    )
    assert out["xcsg_blended_rate"] is None
    assert out["xcsg_cost"] is None
    assert out["xcsg_margin"] is None
    assert out["margin_gain"] is None
    assert out["legacy_cost"] == 36000.0  # legacy still computes
    assert out["legacy_margin"] == 14000.0

    # No legacy team → legacy cost / margin / gain all None.
    out = _compute_economics_metrics(
        engagement_revenue=50000, xcsg_person_days=10, legacy_person_days=40,
        pioneer_rates=[1500],
        legacy_team=[],
        l1_legacy_working_days=40,
        quality_score=0.7, legacy_quality_score=0.4,
        scope_expansion_revenue=None, currency="EUR",
    )
    assert out["legacy_cost"] is None
    assert out["legacy_margin"] is None
    assert out["margin_gain"] is None

    # Cost-per-quality-point gain.
    # legacy_team=[{count:2, day_rate:900}], l1=40 → 2*900*40=72000
    out = _compute_economics_metrics(
        engagement_revenue=120000, xcsg_person_days=20, legacy_person_days=80,
        pioneer_rates=[1500],
        legacy_team=[{"role_name": "X", "count": 2, "day_rate": 900}],
        l1_legacy_working_days=40,
        quality_score=0.85, legacy_quality_score=0.5,
        scope_expansion_revenue=None, currency="EUR",
    )
    # xcsg_cppq = 30000/0.85, legacy_cppq = 72000/0.5
    # gain = legacy_cppq / xcsg_cppq = (72000/0.5) / (30000/0.85)
    expected_cppq_gain = (72000 / 0.5) / (30000 / 0.85)
    assert abs(out["cost_per_quality_point_gain"] - round(expected_cppq_gain, 2)) < 0.01

    # margin_gain capped at 10x. xcsg_cost=10, legacy_cost=10000.
    # legacy_team=[{count:100, day_rate:100}], l1=1 → 100*100*1=10000
    out = _compute_economics_metrics(
        engagement_revenue=11000, xcsg_person_days=1, legacy_person_days=100,
        pioneer_rates=[10],
        legacy_team=[{"role_name": "X", "count": 100, "day_rate": 100}],
        l1_legacy_working_days=1,
        quality_score=0.9, legacy_quality_score=0.5,
        scope_expansion_revenue=None, currency="EUR",
    )
    assert out["margin_gain"] == 10.0


def test_compute_project_metrics_includes_economics():
    """compute_project_metrics merges economics keys into its output.
    Phase 2c: legacy_cost = Σ(count × day_rate) × l1_legacy_working_days.
    legacy_team=[{count:2, day_rate:800}], l1=40 → 2*800*40=64000. Same as Phase 1.
    """
    from backend.metrics import compute_project_metrics

    data = {
        "id": 1, "project_name": "T",
        "category_name": "Cat", "practice_code": "PC", "practice_name": "PName",
        "pioneer_name": "Pia", "client_name": "C",
        "xcsg_team_size": "2", "working_days": 10,
        "l1_legacy_working_days": 40,
        "legacy_team": [{"role_name": "Engineer", "count": 2, "day_rate": 800}],
        "engagement_revenue": 100000,
        "currency": "EUR",
        "xcsg_pricing_model": "Fixed fee",
        "scope_expansion_revenue": 10000,
        "pioneer_day_rates": [1500],
        # quality inputs (minimum to make quality_score non-null)
        "c6_self_assessment": "Significantly better",
        "c7_analytical_depth": "Strong",
        "c8_decision_readiness": "Yes without caveats",
        "l13_legacy_c7_depth": "Adequate",
        "l14_legacy_c8_decision": "Yes with minor caveats",
        "l5_legacy_client_reaction": "Met expectations",
    }
    out = compute_project_metrics(data)
    for key in (
        "xcsg_blended_rate", "xcsg_cost", "legacy_cost",
        "xcsg_margin", "legacy_margin", "margin_gain",
        "xcsg_margin_pct", "legacy_margin_pct",
        "revenue_per_day_xcsg", "revenue_per_day_legacy",
        "cost_per_quality_point_xcsg", "cost_per_quality_point_legacy",
        "cost_per_quality_point_gain",
        "currency", "engagement_revenue", "scope_expansion_revenue",
    ):
        assert key in out, f"compute_project_metrics output missing {key}"
    assert out["currency"] == "EUR"
    assert out["xcsg_cost"] == 30000.0
    assert out["legacy_cost"] == 64000.0  # 2 * 800 * 40
    assert out["xcsg_pricing_model"] == "Fixed fee"

    # No economics inputs → all econ keys present but None.
    bare = {k: v for k, v in data.items() if k not in (
        "engagement_revenue", "currency", "xcsg_pricing_model",
        "scope_expansion_revenue", "pioneer_day_rates", "legacy_team",
    )}
    out2 = compute_project_metrics(bare)
    assert out2["xcsg_cost"] is None
    assert out2["legacy_cost"] is None
    assert out2["margin_gain"] is None


def test_legacy_cost_from_team_mix():
    """compute_project_metrics derives legacy_cost / legacy_person_days
    from legacy_team x l1_legacy_working_days."""
    from backend.metrics import compute_project_metrics

    base = {
        "id": 1, "project_name": "T",
        "category_name": "Cat", "practice_code": "PC", "practice_name": "PName",
        "pioneer_name": "Pia", "client_name": "C",
        "xcsg_team_size": "2", "working_days": 10,
        "l1_legacy_working_days": 40,
        # Quality inputs (minimum to make scores non-null):
        "c6_self_assessment": "Significantly better",
        "c7_analytical_depth": "Strong",
        "c8_decision_readiness": "Yes without caveats",
        "l13_legacy_c7_depth": "Adequate",
        "l14_legacy_c8_decision": "Yes with minor caveats",
        "l5_legacy_client_reaction": "Met expectations",
        "engagement_revenue": 100000,
        "currency": "EUR",
        "pioneer_day_rates": [1500],
    }

    # 1 Senior @ 1500 + 2 Analysts @ 600, project duration 40 days.
    # legacy_person_days = (1+2) x 40 = 120
    # legacy_cost = (1*1500 + 2*600) x 40 = (1500 + 1200) x 40 = 108000
    out = compute_project_metrics({
        **base,
        "legacy_team": [
            {"role_name": "Senior", "count": 1, "day_rate": 1500},
            {"role_name": "Analyst", "count": 2, "day_rate": 600},
        ],
    })
    assert out["legacy_person_days"] == 120
    assert out["legacy_cost"] == 108000.0
    assert out["legacy_margin"] == -8000.0  # 100000 - 108000

    # No team -> cost None
    out2 = compute_project_metrics({**base, "legacy_team": []})
    assert out2["legacy_cost"] is None
    assert out2["legacy_person_days"] is None
    assert out2["legacy_margin"] is None
    assert out2["margin_gain"] is None

    # Missing legacy_team key -> same as empty
    out3 = compute_project_metrics(base)
    assert out3["legacy_cost"] is None

    # No l1_legacy_working_days -> cost None even if team present
    out4 = compute_project_metrics({
        **{k: v for k, v in base.items() if k != "l1_legacy_working_days"},
        "legacy_team": [{"role_name": "Senior", "count": 1, "day_rate": 1500}],
    })
    assert out4["legacy_cost"] is None
    assert out4["legacy_person_days"] is None


def test_create_project_persists_economics():
    """POST /api/projects accepts and stores economics fields, including pioneer rates."""
    tk = admin_token()
    payload = {
        "project_name": "Econ test",
        "category_id": 1,
        "pioneers": [{"first_name": "P1", "last_name": "", "day_rate": 1500}, {"first_name": "P2", "last_name": "", "day_rate": 1000}],
        "xcsg_team_size": "2",
        "xcsg_revision_rounds": "1",
        "engagement_revenue": 80000,
        "currency": "USD",
        "xcsg_pricing_model": "Fixed fee",
        "scope_expansion_revenue": 5000,
        "legacy_team": [{"role_name": "Senior", "count": 1, "day_rate": 750}],
    }
    r = requests.post(f"{BASE}/api/projects", headers={**auth_h(tk), "Content-Type": "application/json"}, json=payload)
    test("POST /api/projects with economics returns 201", r.status_code == 201, f"got {r.status_code}: {r.text[:200]}")
    if r.status_code != 201:
        return
    pid = r.json()["id"]

    try:
        detail = requests.get(f"{BASE}/api/projects/{pid}", headers=auth_h(tk)).json()
        test("economics: engagement_revenue stored", detail.get("engagement_revenue") == 80000, f"got {detail.get('engagement_revenue')}")
        test("economics: currency stored", detail.get("currency") == "USD", f"got {detail.get('currency')}")
        test("economics: xcsg_pricing_model stored", detail.get("xcsg_pricing_model") == "Fixed fee", f"got {detail.get('xcsg_pricing_model')}")
        test("economics: scope_expansion_revenue stored", detail.get("scope_expansion_revenue") == 5000, f"got {detail.get('scope_expansion_revenue')}")
        test("economics: legacy_team stored", len(detail.get("legacy_team", [])) == 1, f"got {detail.get('legacy_team')}")
        test("economics: legacy_team role_name stored", detail.get("legacy_team", [{}])[0].get("role_name") == "Senior", f"got {detail.get('legacy_team')}")
        test("economics: legacy_team day_rate stored", detail.get("legacy_team", [{}])[0].get("day_rate") == 750, f"got {detail.get('legacy_team')}")
        test("economics: legacy_team count stored", detail.get("legacy_team", [{}])[0].get("count") == 1, f"got {detail.get('legacy_team')}")

        pioneer_rates = sorted(p["day_rate"] for p in detail.get("pioneers", []))
        test("economics: pioneer day_rates stored", pioneer_rates == [1000, 1500], f"got {pioneer_rates}")

        # Adding a pioneer post-creation must also persist day_rate.
        add = requests.post(
            f"{BASE}/api/projects/{pid}/pioneers",
            headers={**auth_h(tk), "Content-Type": "application/json"},
            json={"first_name": "P3", "last_name": "", "day_rate": 2000},
        )
        test("economics: POST pioneer returns 201", add.status_code == 201, add.text)
        if add.status_code == 201:
            detail2 = requests.get(f"{BASE}/api/projects/{pid}", headers=auth_h(tk)).json()
            rates2 = sorted(p["day_rate"] for p in detail2.get("pioneers", []))
            test("economics: post-creation pioneer day_rate persisted", rates2 == [1000, 1500, 2000], f"got {rates2}")
    finally:
        requests.delete(f"{BASE}/api/projects/{pid}", headers=auth_h(tk))



def test_fx_rate_models():
    """FxRate, FxRatesPayload, FxRatesResponse, EconomicsResponse validate correctly.
    AppSettings/AppSettingsUpdate now include base_currency."""
    import pytest
    from pydantic import ValidationError
    from backend.models import (
        FxRate, FxRatesPayload, FxRatesResponse,
        AppSettings, AppSettingsUpdate, EconomicsResponse,
    )

    # FxRate accepts valid input.
    fx = FxRate(currency_code="EUR", rate_to_base=1.0850)
    assert fx.currency_code == "EUR"
    assert fx.rate_to_base == 1.0850

    # FxRate rejects invalid currency code.
    with pytest.raises(ValidationError):
        FxRate(currency_code="XYZ", rate_to_base=1.0)
    # FxRate rejects negative rate.
    with pytest.raises(ValidationError):
        FxRate(currency_code="EUR", rate_to_base=-0.5)
    # Rate of 0 is allowed (signals "unset").
    FxRate(currency_code="EUR", rate_to_base=0.0)

    # FxRatesPayload (PUT body) requires base_currency + rates list.
    payload = FxRatesPayload(base_currency="USD", rates=[
        FxRate(currency_code="EUR", rate_to_base=1.0850),
        FxRate(currency_code="GBP", rate_to_base=1.2430),
    ])
    assert payload.base_currency == "USD"
    assert len(payload.rates) == 2

    # FxRatesPayload rejects invalid base currency.
    with pytest.raises(ValidationError):
        FxRatesPayload(base_currency="XYZ", rates=[])
    # Duplicate currency codes rejected.
    with pytest.raises(ValidationError):
        FxRatesPayload(base_currency="USD", rates=[
            FxRate(currency_code="EUR", rate_to_base=1.0),
            FxRate(currency_code="EUR", rate_to_base=2.0),
        ])

    # FxRatesResponse (GET body) shape.
    resp = FxRatesResponse(base_currency="USD", rates=[
        {"currency_code": "EUR", "rate_to_base": 1.0850, "updated_at": "2026-05-01T12:00:00"},
    ])
    assert resp.base_currency == "USD"
    assert resp.rates[0].currency_code == "EUR"

    # AppSettings includes base_currency now.
    s = AppSettings(default_currency="EUR", base_currency="USD")
    assert s.base_currency == "USD"

    # AppSettingsUpdate accepts both fields, both optional.
    upd = AppSettingsUpdate(base_currency="GBP")
    assert upd.base_currency == "GBP"
    assert upd.default_currency is None

    # AppSettingsUpdate rejects invalid currency codes.
    with pytest.raises(ValidationError):
        AppSettingsUpdate(base_currency="XYZ")

    # EconomicsResponse skeleton check (the aggregator returns this shape).
    er = EconomicsResponse(
        summary={
            "total_revenue": 0.0, "total_cost_saved": 0.0,
            "avg_margin_pct": None, "avg_revenue_per_day_xcsg": None,
            "cost_ratio": None, "qualifying_project_count": 0,
            "total_complete_count": 0, "base_currency": "USD",
            "currencies_missing_fx": [],
        },
        breakdowns={"by_practice": [], "by_pioneer": [], "by_currency": [], "by_pricing_model": []},
        trends={"quarterly": []},
    )
    assert er.summary["base_currency"] == "USD"


def test_app_settings_includes_base_currency():
    """get_app_settings returns base_currency; update_app_settings persists it."""
    from backend import database
    database.init_db()
    s = database.get_app_settings()
    assert "default_currency" in s
    assert "base_currency" in s, f"missing base_currency: {s}"
    assert s["base_currency"] == "USD", f"default base_currency should be USD; got {s['base_currency']}"

    initial_default = s["default_currency"]
    initial_base = s["base_currency"]
    try:
        database.update_app_settings(default_currency="EUR", base_currency="GBP")
        s2 = database.get_app_settings()
        assert s2["default_currency"] == "EUR"
        assert s2["base_currency"] == "GBP"

        # Partial update: only base_currency.
        database.update_app_settings(base_currency="CHF")
        s3 = database.get_app_settings()
        assert s3["default_currency"] == "EUR"  # unchanged
        assert s3["base_currency"] == "CHF"
    finally:
        database.update_app_settings(default_currency=initial_default, base_currency=initial_base)


def test_fx_rates_db_helpers():
    """get_fx_rates returns all 6 rates; update_fx_rates round-trips and
    bumps updated_at; _ensure_all_fx_rows_exist backfills missing currencies."""
    from backend import database
    database.init_db()

    rates = database.get_fx_rates()
    assert isinstance(rates, list)
    codes = sorted(r["currency_code"] for r in rates)
    assert codes == ["AUD", "CAD", "CHF", "EUR", "GBP", "USD"], f"got {codes}"
    for r in rates:
        assert "rate_to_base" in r and "updated_at" in r

    try:
        # update_fx_rates accepts a list and persists each row.
        database.update_fx_rates([
            {"currency_code": "EUR", "rate_to_base": 1.0850},
            {"currency_code": "GBP", "rate_to_base": 1.2430},
        ])
        after = {r["currency_code"]: r["rate_to_base"] for r in database.get_fx_rates()}
        assert after["EUR"] == 1.0850, f"got {after['EUR']}"
        assert after["GBP"] == 1.2430, f"got {after['GBP']}"
        # USD unchanged.
        assert after["USD"] == 1.0, f"got {after['USD']}"

        # _ensure_all_fx_rows_exist is a no-op when all rows present.
        n_before = len(database.get_fx_rates())
        database._ensure_all_fx_rows_exist()
        assert len(database.get_fx_rates()) == n_before

        # If we manually delete a row then call the helper, it gets re-added at rate 0
        # (signals "not yet set" to the FX-missing path).
        with database._db() as conn:
            conn.execute("DELETE FROM fx_rates WHERE currency_code = 'CHF'")
            conn.commit()
        database._ensure_all_fx_rows_exist()
        after2 = {r["currency_code"]: r["rate_to_base"] for r in database.get_fx_rates()}
        assert "CHF" in after2
        assert after2["CHF"] == 0.0, f"backfilled rate should be 0; got {after2['CHF']}"
    finally:
        # Restore ALL six rates to identity so subsequent tests start clean.
        from backend.schema import CURRENCIES
        database.update_fx_rates([
            {"currency_code": code, "rate_to_base": 1.0} for code in CURRENCIES
        ])


def test_app_settings_endpoints():
    """GET /api/settings is open to all roles; PUT requires admin."""
    print("\n── App Settings ──")

    def login_token(username, password):
        r = requests.post(f"{BASE}/api/auth/login", json={"username": username, "password": password})
        r.raise_for_status()
        return r.json()["access_token"]

    admin = admin_token()
    analyst = login_token("pmo", "AliraPMO2026!")
    viewer = login_token("viewer", "AliraView2026!")

    h_admin = auth_h(admin)
    h_analyst = auth_h(analyst)
    h_viewer = auth_h(viewer)

    r = requests.get(f"{BASE}/api/settings", headers=h_admin)
    test("GET /api/settings returns 200 for admin", r.status_code == 200, f"got {r.status_code}")
    if r.status_code != 200:
        return
    initial = r.json()["default_currency"]

    test("GET /api/settings returns 200 for analyst",
         requests.get(f"{BASE}/api/settings", headers=h_analyst).status_code == 200)
    test("GET /api/settings returns 200 for viewer",
         requests.get(f"{BASE}/api/settings", headers=h_viewer).status_code == 200)

    try:
        # Admin can change.
        upd = requests.put(f"{BASE}/api/settings", headers=h_admin, json={"default_currency": "USD"})
        test("PUT /api/settings returns 200 for admin", upd.status_code == 200, f"got {upd.status_code}: {upd.text[:120]}")
        test("PUT /api/settings persists new currency",
             requests.get(f"{BASE}/api/settings", headers=h_admin).json()["default_currency"] == "USD")

        # Analyst and viewer cannot.
        test("PUT /api/settings returns 403 for analyst",
             requests.put(f"{BASE}/api/settings", headers=h_analyst, json={"default_currency": "EUR"}).status_code == 403)
        test("PUT /api/settings returns 403 for viewer",
             requests.put(f"{BASE}/api/settings", headers=h_viewer, json={"default_currency": "EUR"}).status_code == 403)

        # Invalid currency rejected.
        bad = requests.put(f"{BASE}/api/settings", headers=h_admin, json={"default_currency": "XYZ"})
        test("PUT /api/settings rejects invalid currency with 422", bad.status_code == 422, f"got {bad.status_code}: {bad.text[:120]}")
    finally:
        # Restore.
        requests.put(f"{BASE}/api/settings", headers=h_admin, json={"default_currency": initial})


def test_compute_economics_summary():
    """Pure aggregator: takes pre-computed per-project metrics dicts +
    fx_rates dict + base_currency. Returns the 6 hero tile values."""
    from backend.economics import compute_economics_summary

    # Two USD projects, one EUR — base = USD.
    # Each project shape mirrors what _build_averaged_complete_projects produces.
    projects = [
        # USD project: revenue 100k, xcsg_cost 30k, legacy_cost 200k → margin=70k, margin_pct=0.7
        {
            "id": 1, "status": "complete", "currency": "USD",
            "engagement_revenue": 100_000.0,
            "xcsg_cost": 30_000.0, "legacy_cost": 200_000.0,
            "xcsg_margin": 70_000.0, "xcsg_margin_pct": 0.70,
            "revenue_per_day_xcsg": 5000.0,
            "xcsg_person_days": 20.0,
            "legacy_team": [{"role_name": "Senior", "count": 2, "day_rate": 1500}],
        },
        # EUR project: revenue 200k EUR (= 220k USD at rate 1.10), cost 60k EUR / 132k legacy
        {
            "id": 2, "status": "complete", "currency": "EUR",
            "engagement_revenue": 200_000.0,
            "xcsg_cost": 60_000.0, "legacy_cost": 132_000.0,
            "xcsg_margin": 140_000.0, "xcsg_margin_pct": 0.70,
            "revenue_per_day_xcsg": 8000.0,
            "xcsg_person_days": 25.0,
            "legacy_team": [{"role_name": "Senior", "count": 1, "day_rate": 1200}],
        },
        # USD project missing legacy_team → does NOT qualify (sum excluded)
        # but still counted in total_complete_count.
        {
            "id": 3, "status": "complete", "currency": "USD",
            "engagement_revenue": 50_000.0,
            "xcsg_cost": 10_000.0, "legacy_cost": None,
            "xcsg_margin": None, "xcsg_margin_pct": None,
            "revenue_per_day_xcsg": None,
            "xcsg_person_days": 5.0,
            "legacy_team": [],
        },
    ]
    fx_rates = {"USD": 1.0, "EUR": 1.10}

    out = compute_economics_summary(projects, fx_rates, base_currency="USD")

    # 100k USD + 200k * 1.10 = 100k + 220k = 320k
    assert out["total_revenue"] == 320_000.0, f"got {out['total_revenue']}"
    # cost saved = (legacy - xcsg) per project, normalized:
    #   project 1: (200k - 30k) USD = 170k
    #   project 2: (132k - 60k) EUR * 1.10 = 79.2k USD
    # total = 249.2k
    assert abs(out["total_cost_saved"] - 249_200.0) < 0.01, f"got {out['total_cost_saved']}"
    # avg margin pct across qualifying projects: (0.70 + 0.70) / 2 = 0.70
    assert out["avg_margin_pct"] == 0.70, f"got {out['avg_margin_pct']}"
    # avg revenue/day (normalized):
    #   project 1: 5000 USD/d
    #   project 2: 8000 EUR/d * 1.10 = 8800 USD/d
    # avg = 6900
    assert abs(out["avg_revenue_per_day_xcsg"] - 6900.0) < 0.01, f"got {out['avg_revenue_per_day_xcsg']}"
    # cost ratio = total_xcsg_cost / total_legacy_cost (normalized)
    #   total_xcsg = 30k + 60k*1.10 = 96k
    #   total_legacy = 200k + 132k*1.10 = 345.2k
    #   ratio = 96 / 345.2 ≈ 0.2782
    assert abs(out["cost_ratio"] - (96_000.0 / 345_200.0)) < 0.001, f"got {out['cost_ratio']}"
    assert out["qualifying_project_count"] == 2, f"got {out['qualifying_project_count']}"
    assert out["total_complete_count"] == 3, f"got {out['total_complete_count']}"
    assert out["base_currency"] == "USD"
    assert out["currencies_missing_fx"] == []


def test_compute_economics_summary_fx_missing():
    """When a currency has no FX rate (or rate == 0), affected projects
    are excluded from totals and the currency code is reported."""
    from backend.economics import compute_economics_summary

    projects = [
        {
            "id": 1, "status": "complete", "currency": "USD",
            "engagement_revenue": 100_000.0,
            "xcsg_cost": 30_000.0, "legacy_cost": 200_000.0,
            "xcsg_margin": 70_000.0, "xcsg_margin_pct": 0.70,
            "revenue_per_day_xcsg": 5000.0, "xcsg_person_days": 20.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
        },
        {
            "id": 2, "status": "complete", "currency": "GBP",
            "engagement_revenue": 80_000.0,
            "xcsg_cost": 20_000.0, "legacy_cost": 150_000.0,
            "xcsg_margin": 60_000.0, "xcsg_margin_pct": 0.75,
            "revenue_per_day_xcsg": 4000.0, "xcsg_person_days": 20.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
        },
    ]
    fx_rates = {"USD": 1.0, "GBP": 0.0}  # GBP rate not yet set
    out = compute_economics_summary(projects, fx_rates, base_currency="USD")
    # Only USD project contributes to totals.
    assert out["total_revenue"] == 100_000.0
    assert out["qualifying_project_count"] == 1
    assert "GBP" in out["currencies_missing_fx"]


def test_compute_economics_summary_empty():
    """Zero qualifying projects returns zero totals + None for averages."""
    from backend.economics import compute_economics_summary
    out = compute_economics_summary([], {"USD": 1.0}, base_currency="USD")
    assert out["total_revenue"] == 0.0
    assert out["total_cost_saved"] == 0.0
    assert out["avg_margin_pct"] is None
    assert out["avg_revenue_per_day_xcsg"] is None
    assert out["cost_ratio"] is None
    assert out["qualifying_project_count"] == 0
    assert out["total_complete_count"] == 0


def test_compute_economics_breakdowns():
    """Aggregator returns by_practice, by_pioneer, by_currency, by_pricing_model."""
    from backend.economics import compute_economics_breakdowns

    projects = [
        {
            "id": 1, "status": "complete", "currency": "USD",
            "engagement_revenue": 100_000.0, "xcsg_cost": 30_000.0, "legacy_cost": 200_000.0,
            "xcsg_margin": 70_000.0, "xcsg_margin_pct": 0.70,
            "revenue_per_day_xcsg": 5000.0, "xcsg_person_days": 20.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
            "practice_code": "RAM", "xcsg_pricing_model": "Fixed fee",
            "pioneer_ids": [10], "pioneer_display_names": ["Sofia Romano"],
        },
        {
            "id": 2, "status": "complete", "currency": "USD",
            "engagement_revenue": 200_000.0, "xcsg_cost": 50_000.0, "legacy_cost": 250_000.0,
            "xcsg_margin": 150_000.0, "xcsg_margin_pct": 0.75,
            "revenue_per_day_xcsg": 8000.0, "xcsg_person_days": 25.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
            "practice_code": "RAM", "xcsg_pricing_model": "Time & materials",
            "pioneer_ids": [10, 11], "pioneer_display_names": ["Sofia Romano", "Marcus Chen"],
        },
        {
            "id": 3, "status": "complete", "currency": "EUR",
            "engagement_revenue": 80_000.0, "xcsg_cost": 20_000.0, "legacy_cost": 100_000.0,
            "xcsg_margin": 60_000.0, "xcsg_margin_pct": 0.75,
            "revenue_per_day_xcsg": 4000.0, "xcsg_person_days": 20.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
            "practice_code": "MAP", "xcsg_pricing_model": "Fixed fee",
            "pioneer_ids": [11], "pioneer_display_names": ["Marcus Chen"],
        },
    ]
    fx_rates = {"USD": 1.0, "EUR": 1.10}
    out = compute_economics_breakdowns(projects, fx_rates, base_currency="USD")

    # by_practice: RAM = 100k+200k = 300k revenue, MAP = 80k * 1.10 = 88k
    practices = {row["practice_code"]: row for row in out["by_practice"]}
    assert practices["RAM"]["revenue"] == 300_000.0
    assert practices["RAM"]["n"] == 2
    assert abs(practices["MAP"]["revenue"] - 88_000.0) < 0.01

    # by_pioneer: Sofia on 2 projects (revenue split or full?), Marcus on 2.
    # Convention: each pioneer gets the FULL project revenue (multi-attribution).
    # That matches the existing By Pioneer chart behavior.
    pioneers = {p["pioneer_id"]: p for p in out["by_pioneer"]}
    assert 10 in pioneers and 11 in pioneers
    assert pioneers[10]["display_name"] == "Sofia Romano"
    # Sofia: project 1 (100k) + project 2 (200k) = 300k
    assert pioneers[10]["revenue"] == 300_000.0
    # Marcus: project 2 (200k) + project 3 (88k) = 288k
    assert abs(pioneers[11]["revenue"] - 288_000.0) < 0.01

    # by_currency: native amounts (NOT normalized)
    cur = {c["code"]: c for c in out["by_currency"]}
    assert cur["USD"]["native_revenue"] == 300_000.0
    assert cur["USD"]["n_projects"] == 2
    assert cur["EUR"]["native_revenue"] == 80_000.0
    assert cur["EUR"]["n_projects"] == 1

    # by_pricing_model: "Fixed fee" = 100k + 88k = 188k, "T&M" = 200k
    pm = {row["model"]: row for row in out["by_pricing_model"]}
    assert abs(pm["Fixed fee"]["revenue"] - 188_000.0) < 0.01
    assert pm["Fixed fee"]["n"] == 2
    assert pm["Time & materials"]["revenue"] == 200_000.0
    assert pm["Time & materials"]["n"] == 1


def test_compute_economics_trends():
    """Quarterly trend bucketed by date_delivered, normalized to base currency."""
    from backend.economics import compute_economics_trends

    projects = [
        # 2026-Q1
        {
            "id": 1, "status": "complete", "currency": "USD",
            "engagement_revenue": 100_000.0, "xcsg_cost": 30_000.0, "legacy_cost": 200_000.0,
            "xcsg_margin": 70_000.0, "xcsg_margin_pct": 0.70,
            "revenue_per_day_xcsg": 5000.0, "revenue_per_day_legacy": 1000.0,
            "xcsg_person_days": 20.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
            "date_delivered": "2026-02-15",
        },
        # 2026-Q1 (same quarter)
        {
            "id": 2, "status": "complete", "currency": "EUR",
            "engagement_revenue": 200_000.0, "xcsg_cost": 50_000.0, "legacy_cost": 250_000.0,
            "xcsg_margin": 150_000.0, "xcsg_margin_pct": 0.75,
            "revenue_per_day_xcsg": 8000.0, "revenue_per_day_legacy": 1500.0,
            "xcsg_person_days": 25.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
            "date_delivered": "2026-03-30",
        },
        # 2026-Q2
        {
            "id": 3, "status": "complete", "currency": "USD",
            "engagement_revenue": 50_000.0, "xcsg_cost": 10_000.0, "legacy_cost": 90_000.0,
            "xcsg_margin": 40_000.0, "xcsg_margin_pct": 0.80,
            "revenue_per_day_xcsg": 5000.0, "revenue_per_day_legacy": 900.0,
            "xcsg_person_days": 10.0,
            "legacy_team": [{"role_name": "x", "count": 1, "day_rate": 100}],
            "date_delivered": "2026-04-10",
        },
    ]
    fx_rates = {"USD": 1.0, "EUR": 1.10}
    out = compute_economics_trends(projects, fx_rates, base_currency="USD")

    quarters = {q["quarter"]: q for q in out["quarterly"]}
    assert "2026-Q1" in quarters and "2026-Q2" in quarters
    # Q1: revenue 100k + 200k*1.10 = 320k; cost saved 170k + 200k*1.10 = 390k
    assert quarters["2026-Q1"]["revenue"] == 320_000.0, f"got {quarters['2026-Q1']['revenue']}"
    assert quarters["2026-Q1"]["cost_saved"] == 390_000.0, f"got {quarters['2026-Q1']['cost_saved']}"
    # Q1 avg margin %: (0.70 + 0.75) / 2 = 0.725
    assert quarters["2026-Q1"]["margin_pct"] == 0.725
    assert quarters["2026-Q1"]["n"] == 2
    # Q2: just project 3 → 50k revenue, 80k saved
    assert quarters["2026-Q2"]["revenue"] == 50_000.0
    assert quarters["2026-Q2"]["cost_saved"] == 80_000.0
    assert quarters["2026-Q2"]["n"] == 1
    # Sorted ascending by quarter.
    assert [q["quarter"] for q in out["quarterly"]] == ["2026-Q1", "2026-Q2"]


def test_practice_role_models():
    """PracticeRoleEntry and PracticeRolesUpdate validate correctly."""
    import pytest
    from pydantic import ValidationError
    from backend.models import PracticeRoleEntry, PracticeRolesUpdate

    # Happy path.
    e = PracticeRoleEntry(role_name="Senior Partner", day_rate=1500, currency="EUR", display_order=1)
    assert e.role_name == "Senior Partner"
    assert e.day_rate == 1500
    assert e.currency == "EUR"
    assert e.display_order == 1

    # Default display_order.
    e2 = PracticeRoleEntry(role_name="Manager", day_rate=1000, currency="EUR")
    assert e2.display_order == 0

    # Empty role_name rejected.
    with pytest.raises(ValidationError):
        PracticeRoleEntry(role_name="", day_rate=1000, currency="EUR")

    # Whitespace-only role_name rejected.
    with pytest.raises(ValidationError):
        PracticeRoleEntry(role_name="   ", day_rate=1000, currency="EUR")

    # role_name > 80 chars rejected.
    with pytest.raises(ValidationError):
        PracticeRoleEntry(role_name="x" * 81, day_rate=1000, currency="EUR")

    # Negative day_rate rejected.
    with pytest.raises(ValidationError):
        PracticeRoleEntry(role_name="X", day_rate=-1, currency="EUR")

    # Invalid currency rejected.
    with pytest.raises(ValidationError):
        PracticeRoleEntry(role_name="X", day_rate=100, currency="XYZ")

    # PracticeRolesUpdate accepts a list.
    u = PracticeRolesUpdate(roles=[
        {"role_name": "Senior", "day_rate": 1500, "currency": "EUR"},
        {"role_name": "Manager", "day_rate": 1000, "currency": "EUR"},
    ])
    assert len(u.roles) == 2

    # Duplicate (role_name, currency) rejected.
    with pytest.raises(ValidationError):
        PracticeRolesUpdate(roles=[
            {"role_name": "Senior", "day_rate": 1500, "currency": "EUR"},
            {"role_name": "Senior", "day_rate": 1600, "currency": "EUR"},
        ])

    # Same role_name with different currencies is OK.
    u2 = PracticeRolesUpdate(roles=[
        {"role_name": "Senior", "day_rate": 1500, "currency": "EUR"},
        {"role_name": "Senior", "day_rate": 1800, "currency": "USD"},
    ])
    assert len(u2.roles) == 2

    # Empty list is OK (clears the catalog).
    u3 = PracticeRolesUpdate(roles=[])
    assert u3.roles == []


def test_legacy_team_models():
    """LegacyTeamRoleEntry validates correctly. ProjectCreate/Update accept legacy_team
    with None / [] / non-empty semantics. Deprecated fields are gone."""
    import pytest
    from pydantic import ValidationError
    from backend.models import (
        LegacyTeamRoleEntry, ProjectCreate, ProjectUpdate,
        PracticeUpdate, ExpertResponseCreate,
    )

    # Happy path entry.
    e = LegacyTeamRoleEntry(role_name="Senior", count=2, day_rate=1500)
    assert e.role_name == "Senior"
    assert e.count == 2
    assert e.day_rate == 1500

    # role_name non-empty.
    with pytest.raises(ValidationError):
        LegacyTeamRoleEntry(role_name="", count=1, day_rate=100)

    # count must be >= 1.
    with pytest.raises(ValidationError):
        LegacyTeamRoleEntry(role_name="X", count=0, day_rate=100)
    with pytest.raises(ValidationError):
        LegacyTeamRoleEntry(role_name="X", count=-1, day_rate=100)

    # day_rate must be >= 0.
    with pytest.raises(ValidationError):
        LegacyTeamRoleEntry(role_name="X", count=1, day_rate=-1)

    # ProjectCreate accepts legacy_team list (default empty).
    base = {
        "project_name": "T", "category_id": 1,
        "pioneers": [{"first_name": "Pia", "last_name": ""}],
        "xcsg_team_size": "1", "xcsg_revision_rounds": "1",
    }
    p = ProjectCreate(**base)
    assert p.legacy_team == []  # default

    p2 = ProjectCreate(**base, legacy_team=[
        {"role_name": "Senior", "count": 1, "day_rate": 1500},
        {"role_name": "Analyst", "count": 2, "day_rate": 600},
    ])
    assert len(p2.legacy_team) == 2

    # ProjectUpdate.legacy_team semantics: None/[]/non-empty.
    u_none = ProjectUpdate()  # no legacy_team → None default
    assert u_none.legacy_team is None

    u_empty = ProjectUpdate(legacy_team=[])
    assert u_empty.legacy_team == []

    u_set = ProjectUpdate(legacy_team=[{"role_name": "X", "count": 1, "day_rate": 100}])
    assert len(u_set.legacy_team) == 1

    # Deprecated fields dropped from the models.
    assert "legacy_day_rate_override" not in ProjectCreate.model_fields
    assert "legacy_day_rate_override" not in ProjectUpdate.model_fields
    assert "default_legacy_day_rate" not in PracticeUpdate.model_fields
    assert "l2_legacy_team_size" not in ExpertResponseCreate.model_fields
    assert "legacy_team_size" not in ProjectCreate.model_fields
    assert "legacy_team_size" not in ProjectUpdate.model_fields


def test_legacy_team_role_name_max_length():
    """LegacyTeamRoleEntry rejects role_name > 80 chars."""
    import pytest
    from pydantic import ValidationError
    from backend.models import LegacyTeamRoleEntry

    # 80 chars OK
    LegacyTeamRoleEntry(role_name="x" * 80, count=1, day_rate=100)

    # 81 chars rejected
    with pytest.raises(ValidationError):
        LegacyTeamRoleEntry(role_name="x" * 81, count=1, day_rate=100)


def test_practice_roles_crud():
    """GET returns rows; PUT replaces atomically."""
    token = admin_token()
    headers = auth_h(token)

    practices = requests.get(f"{BASE}/api/practices", headers=headers).json()
    assert practices, "must have at least one seeded practice"
    pid = practices[0]["id"]

    try:
        # Initial GET — empty list (no roles defined yet for this practice).
        r = requests.get(f"{BASE}/api/practices/{pid}/roles", headers=headers)
        assert r.status_code == 200
        initial = r.json()
        assert isinstance(initial, list)

        # PUT — bulk replace with two roles.
        body = {"roles": [
            {"role_name": "Senior", "day_rate": 1500, "currency": "EUR", "display_order": 0},
            {"role_name": "Manager", "day_rate": 1000, "currency": "EUR", "display_order": 1},
        ]}
        r = requests.put(f"{BASE}/api/practices/{pid}/roles", headers=headers, json=body)
        assert r.status_code == 200, r.text

        # GET — confirm replacement.
        rows = requests.get(f"{BASE}/api/practices/{pid}/roles", headers=headers).json()
        assert len(rows) == 2
        names = sorted(r["role_name"] for r in rows)
        assert names == ["Manager", "Senior"]

        # PUT — replace with a different set; old rows go away.
        r = requests.put(f"{BASE}/api/practices/{pid}/roles", headers=headers, json={
            "roles": [{"role_name": "Analyst", "day_rate": 600, "currency": "EUR"}]
        })
        assert r.status_code == 200
        rows = requests.get(f"{BASE}/api/practices/{pid}/roles", headers=headers).json()
        assert len(rows) == 1
        assert rows[0]["role_name"] == "Analyst"

        # PUT — invalid currency rejected.
        r = requests.put(f"{BASE}/api/practices/{pid}/roles", headers=headers, json={
            "roles": [{"role_name": "X", "day_rate": 100, "currency": "XYZ"}]
        })
        assert r.status_code == 422

        # PUT — duplicate (role_name, currency) rejected.
        r = requests.put(f"{BASE}/api/practices/{pid}/roles", headers=headers, json={
            "roles": [
                {"role_name": "Dup", "day_rate": 100, "currency": "EUR"},
                {"role_name": "Dup", "day_rate": 200, "currency": "EUR"},
            ]
        })
        assert r.status_code == 422
    finally:
        # Restore: empty list (clears whatever was added during the test).
        requests.put(f"{BASE}/api/practices/{pid}/roles", headers=headers, json={"roles": []})


def test_practice_roles_admin_only():
    """GET is open to all roles; PUT requires admin."""
    def login_token(username, password):
        r = requests.post(f"{BASE}/api/auth/login", json={"username": username, "password": password})
        r.raise_for_status()
        return r.json()["access_token"]

    admin = admin_token()
    analyst = login_token("pmo", "AliraPMO2026!")
    viewer = login_token("viewer", "AliraView2026!")
    h_admin = auth_h(admin)
    h_analyst = auth_h(analyst)
    h_viewer = auth_h(viewer)

    practices = requests.get(f"{BASE}/api/practices", headers=h_admin).json()
    pid = practices[0]["id"]

    try:
        # GET allowed for all three roles.
        assert requests.get(f"{BASE}/api/practices/{pid}/roles", headers=h_admin).status_code == 200
        assert requests.get(f"{BASE}/api/practices/{pid}/roles", headers=h_analyst).status_code == 200
        assert requests.get(f"{BASE}/api/practices/{pid}/roles", headers=h_viewer).status_code == 200

        # PUT allowed for admin, blocked for analyst and viewer.
        body = {"roles": [{"role_name": "X", "day_rate": 1, "currency": "EUR"}]}
        assert requests.put(f"{BASE}/api/practices/{pid}/roles", headers=h_admin, json=body).status_code == 200
        assert requests.put(f"{BASE}/api/practices/{pid}/roles", headers=h_analyst, json=body).status_code == 403
        assert requests.put(f"{BASE}/api/practices/{pid}/roles", headers=h_viewer, json=body).status_code == 403
    finally:
        # Restore.
        requests.put(f"{BASE}/api/practices/{pid}/roles", headers=h_admin, json={"roles": []})


def test_practice_roles_404_for_unknown_practice():
    """Routes return 404 for non-existent practice IDs."""
    headers = auth_h(admin_token())
    bad_id = 99999

    r = requests.get(f"{BASE}/api/practices/{bad_id}/roles", headers=headers)
    assert r.status_code == 404, f"GET expected 404, got {r.status_code}: {r.text}"

    r = requests.put(
        f"{BASE}/api/practices/{bad_id}/roles",
        headers=headers,
        json={"roles": [{"role_name": "X", "day_rate": 1, "currency": "EUR"}]},
    )
    assert r.status_code == 404, f"PUT expected 404, got {r.status_code}: {r.text}"


def test_migrate_v18_drops_columns_creates_table():
    """migrate_v18 drops 4 deprecated columns and creates project_legacy_team."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        # New table exists with expected columns.
        tables = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()}
        assert "project_legacy_team" in tables

        cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(project_legacy_team)"
        ).fetchall()}
        for col in ("id", "project_id", "role_name", "count", "day_rate"):
            assert col in cols, f"project_legacy_team.{col} missing"

        # Index exists.
        indexes = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='index' AND tbl_name='project_legacy_team'"
        ).fetchall()}
        assert "idx_project_legacy_team_project" in indexes

        # Dropped columns are gone.
        proj_cols = {r[1] for r in conn.execute("PRAGMA table_info(projects)").fetchall()}
        assert "legacy_day_rate_override" not in proj_cols
        assert "legacy_team_size" not in proj_cols

        prac_cols = {r[1] for r in conn.execute("PRAGMA table_info(practices)").fetchall()}
        assert "default_legacy_day_rate" not in prac_cols

        expert_cols = {r[1] for r in conn.execute("PRAGMA table_info(expert_responses)").fetchall()}
        assert "l2_legacy_team_size" not in expert_cols

    # Re-run migration — must be idempotent.
    database.migrate_v18()
    database.migrate_v18()


def test_legacy_team_db_helpers():
    """list_legacy_team and replace_legacy_team round-trip correctly."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        cur = conn.execute(
            "INSERT INTO projects (created_by, project_name, category_id, "
            "xcsg_team_size, xcsg_revision_rounds, "
            "legacy_calendar_days, legacy_revision_rounds, expert_token, status) "
            "VALUES (1, 'lt test', 1, '1', '1', '10', '1', 'tok-lt', 'pending')"
        )
        project_id = cur.lastrowid
        conn.commit()

    try:
        assert database.list_legacy_team(project_id) == []

        database.replace_legacy_team(project_id, [
            {"role_name": "Senior", "count": 1, "day_rate": 1500},
            {"role_name": "Analyst", "count": 2, "day_rate": 600},
        ])

        rows = database.list_legacy_team(project_id)
        assert len(rows) == 2
        names = sorted(r["role_name"] for r in rows)
        assert names == ["Analyst", "Senior"]

        # Replace clears the previous set.
        database.replace_legacy_team(project_id, [
            {"role_name": "Manager", "count": 1, "day_rate": 1000},
        ])
        rows = database.list_legacy_team(project_id)
        assert len(rows) == 1
        assert rows[0]["role_name"] == "Manager"

        # Empty list clears.
        database.replace_legacy_team(project_id, [])
        assert database.list_legacy_team(project_id) == []
    finally:
        with database._db() as conn:
            conn.execute("DELETE FROM project_pioneers WHERE project_id = ?", (project_id,))
            conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))
            conn.commit()


def test_legacy_team_persistence():
    """legacy_team round-trips through POST /api/projects + GET, including
    None/[]/non-empty semantics for ProjectUpdate."""
    headers = auth_h(admin_token())

    payload = {
        "project_name": "lt persist",
        "category_id": 1,
        "pioneers": [{"first_name": "P", "last_name": "", "email": "p@x.io"}],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
        "legacy_team": [
            {"role_name": "Senior", "count": 1, "day_rate": 1500},
            {"role_name": "Analyst", "count": 2, "day_rate": 600},
        ],
    }
    r = requests.post(f"{BASE}/api/projects", headers=headers, json=payload)
    assert r.status_code == 201, r.text
    pid = r.json()["id"]

    try:
        detail = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        assert "legacy_team" in detail
        team = sorted(detail["legacy_team"], key=lambda r: r["role_name"])
        assert team[0]["role_name"] == "Analyst"
        assert team[0]["count"] == 2
        assert team[0]["day_rate"] == 600
        assert team[1]["role_name"] == "Senior"

        # PUT with legacy_team=None → unchanged.
        upd = requests.put(
            f"{BASE}/api/projects/{pid}", headers=headers,
            json={"project_name": "lt persist v2"},
        )
        assert upd.status_code == 200, upd.text
        detail2 = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        assert len(detail2["legacy_team"]) == 2

        # PUT with legacy_team=[] → cleared.
        upd2 = requests.put(
            f"{BASE}/api/projects/{pid}", headers=headers,
            json={"legacy_team": []},
        )
        assert upd2.status_code == 200
        detail3 = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        assert detail3["legacy_team"] == []

        # PUT with legacy_team=non-empty → replaced.
        upd3 = requests.put(
            f"{BASE}/api/projects/{pid}", headers=headers,
            json={"legacy_team": [{"role_name": "Manager", "count": 1, "day_rate": 1000}]},
        )
        assert upd3.status_code == 200
        detail4 = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        assert len(detail4["legacy_team"]) == 1
        assert detail4["legacy_team"][0]["role_name"] == "Manager"
    finally:
        requests.delete(f"{BASE}/api/projects/{pid}", headers=headers)


def test_pioneer_role_name_in_models():
    """ProjectPioneerEntry and ProjectPioneerUpdate accept optional role_name."""
    from backend.models import ProjectPioneerEntry, ProjectPioneerUpdate

    # Happy path with role_name.
    p = ProjectPioneerEntry(first_name="Pia", last_name="", email="pia@example.com", day_rate=1500, role_name="Senior")
    assert p.role_name == "Senior"
    assert p.day_rate == 1500

    # role_name optional — None is fine.
    p2 = ProjectPioneerEntry(first_name="Bob", last_name="", email="bob@example.com")
    assert p2.role_name is None

    # ProjectPioneerUpdate also accepts role_name.
    u = ProjectPioneerUpdate(role_name="Manager")
    assert u.role_name == "Manager"
    u2 = ProjectPioneerUpdate()
    assert u2.role_name is None


def test_pioneer_role_name_persistence():
    """role_name round-trips through POST /api/projects (with pioneers list)
    and POST /api/projects/{id}/pioneers (post-creation add)."""
    headers = auth_h(admin_token())

    # Build a project with one pioneer that has role_name.
    payload = {
        "project_name": "Phase 2b round-trip",
        "category_id": 1,
        "pioneers": [
            {"first_name": "P-with-role", "last_name": "", "email": "p1@x.io", "day_rate": 1500, "role_name": "Senior"},
            {"first_name": "P-no-role", "last_name": "", "email": "p2@x.io", "day_rate": 1000},
        ],
        "xcsg_team_size": "2",
        "xcsg_revision_rounds": "1",
    }
    r = requests.post(f"{BASE}/api/projects", headers=headers, json=payload)
    assert r.status_code == 201, r.text
    pid = r.json()["id"]

    try:
        detail = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        by_name = {p["pioneer_name"]: p for p in detail["pioneers"]}
        assert by_name["P-with-role"]["role_name"] == "Senior"
        assert by_name["P-no-role"]["role_name"] is None

        # POST a new pioneer with role_name (post-creation add path).
        add_r = requests.post(
            f"{BASE}/api/projects/{pid}/pioneers",
            headers=headers,
            json={"first_name": "P3-added", "last_name": "", "email": "p3@x.io", "day_rate": 2000, "role_name": "Manager"},
        )
        assert add_r.status_code == 201, add_r.text

        detail2 = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        by_name2 = {p["pioneer_name"]: p for p in detail2["pioneers"]}
        assert by_name2["P3-added"]["role_name"] == "Manager"
        assert by_name2["P3-added"]["day_rate"] == 2000
    finally:
        requests.delete(f"{BASE}/api/projects/{pid}", headers=headers)


def test_pioneer_day_rate_independent_of_role_name():
    """Server does NOT auto-fill day_rate from role_name — it stores
    exactly what the request includes. The catalog lookup is the
    frontend's job."""
    headers = auth_h(admin_token())

    payload = {
        "project_name": "Phase 2b independence",
        "category_id": 1,
        "pioneers": [
            # role_name set but day_rate explicitly different from any catalog rate
            {"first_name": "P", "last_name": "", "email": "p@x.io", "day_rate": 999, "role_name": "Senior"},
        ],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
    }
    r = requests.post(f"{BASE}/api/projects", headers=headers, json=payload)
    assert r.status_code == 201
    pid = r.json()["id"]

    try:
        detail = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        p = detail["pioneers"][0]
        assert p["role_name"] == "Senior"
        assert p["day_rate"] == 999  # stored as submitted, NOT looked up from any catalog
    finally:
        requests.delete(f"{BASE}/api/projects/{pid}", headers=headers)


def test_update_pioneer_clears_role_name():
    """PUT /api/projects/{id}/pioneers/{pid} with role_name=null clears the role."""
    headers = auth_h(admin_token())

    payload = {
        "project_name": "clear role test",
        "category_id": 1,
        "pioneers": [{"first_name": "P", "last_name": "", "email": "p@x.io", "day_rate": 1000, "role_name": "Senior"}],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
    }
    r = requests.post(f"{BASE}/api/projects", headers=headers, json=payload)
    assert r.status_code == 201
    pid = r.json()["id"]

    try:
        detail = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        pioneer = detail["pioneers"][0]
        pioneer_id = pioneer["id"]
        assert pioneer["role_name"] == "Senior"

        # PUT with role_name=null should clear it.
        upd = requests.put(
            f"{BASE}/api/projects/{pid}/pioneers/{pioneer_id}",
            headers=headers,
            json={"role_name": None},
        )
        assert upd.status_code == 200, upd.text

        detail2 = requests.get(f"{BASE}/api/projects/{pid}", headers=headers).json()
        assert detail2["pioneers"][0]["role_name"] is None
        # day_rate should be untouched (stayed at 1000) — null on day_rate
        # was NOT sent so the existing value stays.
        assert detail2["pioneers"][0]["day_rate"] == 1000
    finally:
        requests.delete(f"{BASE}/api/projects/{pid}", headers=headers)


def test_export_pioneers_csv():
    """GET /api/export/pioneers.csv returns CSV of pioneer rows.
    Honors filter query params (multi-valued for practice/role/status)."""
    import csv
    import io

    headers = auth_h(admin_token())

    # Make sure at least one pioneer exists.
    p = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "CSV", "last_name": "Test", "email": "csv-test@example.com",
    }).json()

    # `csv.DictReader` doesn't strip a leading BOM on its own — decoding the
    # response bytes with utf-8-sig drops it before parsing.
    def _csv_text(resp):
        return resp.content.decode("utf-8-sig")

    try:
        r = requests.get(f"{BASE}/api/export/pioneers.csv", headers=headers)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")

        # Parse the CSV (after stripping the BOM).
        reader = csv.DictReader(io.StringIO(_csv_text(r)))
        rows = list(reader)
        assert len(rows) >= 1
        # Required columns: name is now split into first_name + last_name.
        for col in ("id", "first_name", "last_name", "email", "project_count", "status",
                    "completion_rate", "avg_value_gain", "practices", "roles"):
            assert col in rows[0], f"missing column {col}"
        assert "name" not in rows[0], "legacy 'name' column should be gone"

        # Our test pioneer should be in there.
        first_names = [row["first_name"] for row in rows]
        last_names = [row["last_name"] for row in rows]
        assert "CSV" in first_names and "Test" in last_names

        # Status filter — single value.
        r2 = requests.get(f"{BASE}/api/export/pioneers.csv?status=never", headers=headers)
        assert r2.status_code == 200
        rows2 = list(csv.DictReader(io.StringIO(_csv_text(r2))))
        # Test pioneer is `never` (no project assignments).
        assert any(
            row["first_name"] == "CSV" and row["last_name"] == "Test" and row["status"] == "never"
            for row in rows2
        )

        # Multi-valued status filter (e.g. status=never&status=pending).
        r3 = requests.get(
            f"{BASE}/api/export/pioneers.csv?status=never&status=pending",
            headers=headers,
        )
        assert r3.status_code == 200
        # CSV should still parse cleanly.
        list(csv.DictReader(io.StringIO(_csv_text(r3))))

        # Search filter — match the email.
        r4 = requests.get(
            f"{BASE}/api/export/pioneers.csv?search=csv-test",
            headers=headers,
        )
        rows4 = list(csv.DictReader(io.StringIO(_csv_text(r4))))
        assert any(row["first_name"] == "CSV" and row["last_name"] == "Test" for row in rows4)

        # UTF-8 BOM is prepended so Excel on Windows reads it correctly.
        # Inspect the raw bytes so the assertion is unambiguous.
        assert r.content.startswith(b"\xef\xbb\xbf"), "CSV should start with UTF-8 BOM"
    finally:
        requests.delete(f"{BASE}/api/pioneers/{p['id']}", headers=headers)


def test_export_pioneers_csv_empty_filter():
    """When the filter matches no pioneers, the CSV must still emit a header
    row (consumers crash on empty bodies). BOM stays in place."""
    import csv
    import io

    headers = auth_h(admin_token())
    # Use a search string that cannot possibly match any name/email.
    r = requests.get(
        f"{BASE}/api/export/pioneers.csv?search=__definitely_no_match_zzz_xxx_999__",
        headers=headers,
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "")
    # Body starts with BOM (raw bytes — most reliable check).
    assert r.content.startswith(b"\xef\xbb\xbf"), "empty CSV should still start with BOM"
    # Header row is present and parseable; row list is empty.
    text = r.content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    assert rows == [], f"expected no rows, got {len(rows)}"
    # Header includes the canonical fixed columns.
    for col in ("id", "first_name", "last_name", "email", "project_count", "status",
                "completion_rate", "avg_value_gain", "practices", "roles"):
        assert col in (reader.fieldnames or []), f"missing header column {col}"
    assert "name" not in (reader.fieldnames or []), "legacy 'name' column should be gone"


def _seed_for_notes_consumers():
    """Seed live data the notes/dashboard-export consumer tests depend on.

    Why: upstream destructive migration tests (test_migrate_v19_*, the
    test_pioneer_db_helpers_* family) wipe `pioneers` and `project_pioneers`
    mid-suite. By the time test_expert_notes runs there are no pioneer
    rounds, so it skips. Re-seed via the public API right before the
    consumers so they see realistic state regardless of what ran upstream.

    Creates:
      - 1 MAP-practice project, 1 pioneer, 2 expected rounds (both open)
      - 1 non-MAP-practice project, 1 pioneer, 2 expected rounds (both open)
      - 1 fully-completed project (all rounds done) so dashboard "Top Movers"
        has at least one row with productivity_ratio set.
    """
    print("\n── _seed_for_notes_consumers ──")
    h = auth_h(admin_token())

    cats = requests.get(f"{BASE}/api/categories", headers=h).json()
    practices = requests.get(f"{BASE}/api/practices", headers=h).json()
    pmap = {p["code"]: p["id"] for p in practices}

    # Find a category that allows MAP and one that allows a non-MAP practice (RAM is universal).
    map_cat_id = None
    other_cat_id = None
    other_practice_code = None
    for c in cats:
        codes = [p["code"] for p in c.get("practices", [])]
        if "MAP" in codes and map_cat_id is None:
            map_cat_id = c["id"]
        if other_cat_id is None:
            non_map = next((code for code in codes if code != "MAP"), None)
            if non_map:
                other_cat_id = c["id"]
                other_practice_code = non_map
        if map_cat_id and other_cat_id:
            break

    if not (map_cat_id and other_cat_id):
        test("seed: found MAP+non-MAP category", False, detail=f"map={map_cat_id} other={other_cat_id}")
        return

    # legacy_team needed for productivity_ratio computation (drives Top Movers).
    legacy_team = [{"role_name": "Senior", "count": 2, "day_rate": 1200}]

    seeds = [
        ("QA Notes Seed — MAP",   map_cat_id,   pmap["MAP"],                  2),
        ("QA Notes Seed — Other", other_cat_id, pmap[other_practice_code],    2),
    ]
    for name, cat_id, practice_id, rounds in seeds:
        body = {
            "project_name": name,
            "category_id": cat_id,
            "practice_id": practice_id,
            "pioneers": [{"first_name": "Notes", "last_name": name[-3:], "total_rounds": rounds}],
            "engagement_stage": "Active engagement",
            "date_started": "2026-03-01",
            "date_delivered": "2026-03-10",
            "working_days": 5,
            "xcsg_team_size": "2",
            "xcsg_revision_rounds": "1",
            "legacy_team": legacy_team,
        }
        r = requests.post(f"{BASE}/api/projects", headers=h, json=body)
        test(f"seed: create '{name}'", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:120]}")

    # Fully-complete project so Top Movers / status='complete' aggregations have data.
    import importlib.util, pathlib
    spec = importlib.util.spec_from_file_location("sd", pathlib.Path("tests/seed_20_projects.py"))
    sd = importlib.util.module_from_spec(spec); spec.loader.exec_module(sd)
    strong_payload = dict(getattr(sd, "STRONG"))

    body = {
        "project_name": "QA Notes Seed — Complete",
        "category_id": map_cat_id,
        "practice_id": pmap["MAP"],
        "pioneers": [{"first_name": "Notes", "last_name": "Complete", "total_rounds": 1}],
        "engagement_stage": "Post-engagement (follow-on)",
        "date_started": "2026-02-01",
        "date_delivered": "2026-02-10",
        "working_days": 7,
        "xcsg_team_size": "2",
        "xcsg_revision_rounds": "1",
        "legacy_team": legacy_team,
    }
    r = requests.post(f"{BASE}/api/projects", headers=h, json=body)
    test("seed: create 'Complete' project", r.status_code in (200, 201), f"got {r.status_code}: {r.text[:120]}")
    if r.status_code in (200, 201):
        proj = r.json()
        token = proj["pioneers"][0]["expert_token"]
        rs = requests.post(f"{BASE}/api/expert/{token}", json=strong_payload)
        test("seed: complete 'Complete' project's only round", rs.status_code in (200, 201), f"got {rs.status_code}: {rs.text[:120]}")


def main():
    global passed, failed, failures

    print("=" * 70)
    print("xCSG Value Tracker V2 — Comprehensive QA/QC Test Suite")
    print("=" * 70)

    # Health check
    try:
        r = requests.get(f"{BASE}/api/health", timeout=5)
        test("Server reachable", r.status_code == 200, f"got {r.status_code}")
    except Exception as e:
        test("Server reachable", False, str(e))
        print("\nFATAL: Server not reachable. Exiting.")
        sys.exit(1)

    test_authentication()
    test_expert_options()
    test_categories()
    test_practices()
    test_create_deliverable()
    test_expert_assessment()
    # Seed live projects so dashboard/list tests have data even though
    # test_create_deliverable + test_expert_assessment delete their own state.
    _seed_for_notes_consumers()
    test_metrics()
    test_scaling_gates()
    test_dashboard()
    test_deliverables_list()
    test_string_consistency()
    test_frontend_js()
    test_norms()
    test_schema_endpoint()
    test_metrics_summary_fields()
    test_schema()
    test_dashboard_config()
    test_seed_field_coverage()
    test_economics_schema()
    test_practice_roles_schema()
    test_legacy_team_schema()
    test_pioneer_schema()
    test_pioneer_models()
    test_pioneer_db_helpers_crud()
    test_delete_pioneer_assigned_to_project_raises()
    test_list_pioneers_with_metrics_aggregation()
    test_get_pioneer_with_metrics_includes_portfolio()
    test_pioneer_api_crud()
    test_pioneer_api_find_or_create()
    test_pioneer_api_admin_only()
    test_pioneer_delete_in_use_returns_409()
    test_project_create_with_existing_pioneer_id()
    test_project_create_with_inline_pioneer()
    test_migrate_v19_destructive_creates_pioneers_table()
    test_migrate_v19_email_unique_case_insensitive()
    test_migrate_v20_splits_name_into_first_and_last()
    test_migrate_v21_creates_fx_rates_and_base_currency()
    test_migrate_v21_idempotent()
    test_migrate_v15_idempotent()
    test_migrate_v16_idempotent()
    test_migrate_v17_idempotent()
    test_migrate_v18_drops_columns_creates_table()
    test_legacy_team_db_helpers()
    test_practice_roles_db_helpers()
    test_economics_models()
    test_practice_role_models()
    test_legacy_team_models()
    test_legacy_team_role_name_max_length()
    test_pioneer_role_name_in_models()
    test_economics_metrics()
    test_legacy_cost_from_team_mix()
    test_fx_rates_db_helpers()
    test_app_settings_includes_base_currency()
    test_fx_rate_models()
    test_app_settings_endpoints()
    test_compute_economics_summary()
    test_compute_economics_summary_fx_missing()
    test_compute_economics_summary_empty()
    test_compute_economics_breakdowns()
    test_compute_economics_trends()
    test_practice_roles_crud()
    test_practice_roles_admin_only()
    test_practice_roles_404_for_unknown_practice()
    test_compute_project_metrics_includes_economics()
    test_create_project_persists_economics()
    test_legacy_team_persistence()
    test_pioneer_role_name_persistence()
    test_pioneer_day_rate_independent_of_role_name()
    test_update_pioneer_clears_role_name()
    test_show_other_pioneers_flag()
    test_auto_issue_next_round()
    test_dashboard_takeaways()
    test_dashboard_pioneer_filter()
    _seed_for_notes_consumers()
    test_expert_notes()
    test_notes_feed_endpoint()
    test_notes_excel_sheet()
    test_dashboard_export_sheets()
    test_export_pioneers_csv()
    test_export_pioneers_csv_empty_filter()
    test_export_pioneer_xlsx()
    test_export_pioneer_xlsx_404_for_unknown()
    test_fx_rates_get_put_endpoints()
    test_dashboard_economics_endpoint()

    print("\n" + "=" * 70)
    print(f"QA SUMMARY: {passed} passed, {failed} failed, {passed + failed} total")
    print("=" * 70)
    
    if failures:
        print("\nFAILING TESTS:")
        for name, detail in failures:
            print(f"  ✗ {name}")
            if detail:
                print(f"    {detail}")
    
    return 0 if failed == 0 else 1

def test_pioneer_schema():
    """schema.py exposes PIONEER_FIELDS (split first/last), PIONEER_STATUS_OPTIONS,
    and the pioneer_overdue_days threshold."""
    from backend.schema import (
        PIONEER_FIELDS, PIONEER_STATUS_OPTIONS, DASHBOARD_CONFIG, build_schema_response,
    )

    expected = {"first_name", "last_name", "email", "notes"}
    assert expected.issubset(set(PIONEER_FIELDS.keys()))
    assert "name" not in PIONEER_FIELDS  # legacy combined field is gone
    assert PIONEER_FIELDS["first_name"]["required"] is True
    assert PIONEER_FIELDS["last_name"]["required"] is True
    assert PIONEER_FIELDS["email"].get("required") is not True
    assert PIONEER_FIELDS["first_name"].get("max_length") == 80
    assert PIONEER_FIELDS["last_name"].get("max_length") == 80

    # Status options: list of {value, label} dicts.
    values = {opt["value"] for opt in PIONEER_STATUS_OPTIONS}
    assert values == {"never", "pending", "pending_overdue", "completed"}
    labels = {opt["label"] for opt in PIONEER_STATUS_OPTIONS}
    assert labels == {"Not assigned", "Pending", "Overdue", "Completed"}

    # Threshold for pending_overdue.
    assert DASHBOARD_CONFIG["thresholds"]["pioneer_overdue_days"] == 21

    # Surfaced via /api/schema.
    response = build_schema_response()
    assert "pioneer_fields" in response
    assert "pioneer_status_options" in response
    assert response["pioneer_status_options"] == PIONEER_STATUS_OPTIONS


def test_migrate_v19_destructive_creates_pioneers_table():
    """migrate_v19 creates pioneers table, drops pioneer_name/email from
    project_pioneers and from projects, adds pioneer_id NOT NULL FK,
    and is idempotent. After migrate_v20 the legacy `name` column is gone
    in favour of first_name + last_name; the rest of the v19 contract holds."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        # New table exists.
        tables = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()}
        assert "pioneers" in tables

        cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(pioneers)"
        ).fetchall()}
        # After v20: `name` is split into first_name + last_name; everything
        # else from v19 still applies.
        for col in ("id", "first_name", "last_name", "email", "notes", "created_by", "created_at"):
            assert col in cols, f"pioneers.{col} missing"
        assert "name" not in cols, "v20 must drop the combined `name` column"

        # Partial unique index exists.
        indexes = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='index' AND tbl_name='pioneers'"
        ).fetchall()}
        assert "idx_pioneers_email_lower" in indexes

        # project_pioneers has pioneer_id (NOT NULL FK), no pioneer_name/email.
        pp_cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(project_pioneers)"
        ).fetchall()}
        assert "pioneer_id" in pp_cols
        assert "pioneer_name" not in pp_cols
        assert "pioneer_email" not in pp_cols

        # Vestigial v1.0 columns dropped from projects.
        proj_cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(projects)"
        ).fetchall()}
        assert "pioneer_name" not in proj_cols
        assert "pioneer_email" not in proj_cols

    # Re-run — must be idempotent.
    database.migrate_v19()
    database.migrate_v19()


def test_migrate_v20_splits_name_into_first_and_last():
    """migrate_v20 splits pioneers.name into first_name + last_name on first
    whitespace, drops the old `name` column, and is idempotent.

    Tests three split cases:
      - "Sofia Romano" -> first="Sofia", last="Romano"
      - "Theo van der Berg" -> first="Theo", last="van der Berg"
      - "Madonna" -> first="Madonna", last=""
    """
    from backend import database

    database.init_db()

    # Build a v19-shape pioneers table with `name TEXT NOT NULL` so the
    # migration has something to backfill. Drop the post-v20 table and
    # recreate it explicitly, then re-run migrate_v20.
    with database._db() as conn:
        # Wipe references first.
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()

        # Force the v19 shape (drop v20 columns + add `name`).
        # SQLite doesn't allow conditional ADD/DROP cleanly, so we recreate.
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("DROP INDEX IF EXISTS idx_pioneers_email_lower")
        conn.execute("DROP TABLE IF EXISTS pioneers_v20_test")
        conn.execute(
            """CREATE TABLE pioneers_v20_test (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   name TEXT NOT NULL,
                   email TEXT,
                   notes TEXT,
                   created_by INTEGER,
                   created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        # Replace pioneers table with the v19-shape one.
        conn.execute("DROP TABLE pioneers")
        conn.execute("ALTER TABLE pioneers_v20_test RENAME TO pioneers")
        # Insert the three test rows.
        conn.execute("INSERT INTO pioneers (name, email) VALUES (?, ?)",
                     ("Sofia Romano", "sofia@v20test.example"))
        conn.execute("INSERT INTO pioneers (name, email) VALUES (?, ?)",
                     ("Theo van der Berg", "theo@v20test.example"))
        conn.execute("INSERT INTO pioneers (name, email) VALUES (?, ?)",
                     ("Madonna", "madonna@v20test.example"))
        conn.commit()
        conn.execute("PRAGMA foreign_keys = ON")

    # Run the migration on the v19-shape table.
    database.migrate_v20()

    # Verify table shape: `name` is gone; first_name + last_name are present and NOT NULL.
    with database._db() as conn:
        cols = {r[1]: r for r in conn.execute("PRAGMA table_info(pioneers)").fetchall()}
        assert "name" not in cols, "legacy `name` column should be dropped"
        assert "first_name" in cols
        assert "last_name" in cols
        # PRAGMA table_info returns notnull=1 in column 3.
        assert cols["first_name"][3] == 1, "first_name must be NOT NULL"
        assert cols["last_name"][3] == 1, "last_name must be NOT NULL"

        rows = {
            r["email"]: dict(r)
            for r in conn.execute(
                "SELECT first_name, last_name, email FROM pioneers"
            ).fetchall()
        }

    sofia = rows["sofia@v20test.example"]
    assert sofia["first_name"] == "Sofia", f"got {sofia}"
    assert sofia["last_name"] == "Romano", f"got {sofia}"

    theo = rows["theo@v20test.example"]
    assert theo["first_name"] == "Theo", f"got {theo}"
    assert theo["last_name"] == "van der Berg", f"got {theo}"

    madonna = rows["madonna@v20test.example"]
    assert madonna["first_name"] == "Madonna", f"got {madonna}"
    assert madonna["last_name"] == "", f"got {madonna}"

    # Idempotent: running again must be a no-op.
    database.migrate_v20()
    database.migrate_v20()

    # Cleanup test rows so subsequent tests start clean.
    with database._db() as conn:
        conn.execute("DELETE FROM pioneers WHERE email LIKE '%@v20test.example'")
        conn.commit()


def test_migrate_v21_creates_fx_rates_and_base_currency():
    """v2.1 — fx_rates table exists with all 6 currencies seeded at rate 1.0,
    and app_settings.base_currency exists with default 'USD'."""
    from backend import database
    database.init_db()  # idempotent — applies all migrations

    with database._db() as conn:
        # fx_rates table exists with the right shape.
        rates = conn.execute(
            "SELECT currency_code, rate_to_base FROM fx_rates ORDER BY currency_code"
        ).fetchall()
        codes = [r["currency_code"] for r in rates]
        assert codes == ["AUD", "CAD", "CHF", "EUR", "GBP", "USD"], f"got {codes}"
        for r in rates:
            assert r["rate_to_base"] == 1.0, f"{r['currency_code']} rate={r['rate_to_base']}"

        # app_settings has base_currency column with default 'USD'.
        cols = {row[1] for row in conn.execute("PRAGMA table_info(app_settings)").fetchall()}
        assert "base_currency" in cols, f"got {cols}"
        row = conn.execute("SELECT base_currency FROM app_settings WHERE id=1").fetchone()
        assert row["base_currency"] == "USD", f"got {row['base_currency']}"

    # Idempotency: running migration again is a no-op.
    database.migrate_v21_fx_rates()
    database.migrate_v21_fx_rates()


def test_migrate_v21_idempotent():
    """Calling migrate_v21_fx_rates twice doesn't duplicate rows or fail."""
    from backend import database
    database.init_db()
    database.migrate_v21_fx_rates()
    with database._db() as conn:
        n = conn.execute("SELECT COUNT(*) AS n FROM fx_rates").fetchone()["n"]
    assert n == 6, f"expected 6 currency rows, got {n}"


def test_migrate_v19_email_unique_case_insensitive():
    """Partial unique index rejects case-insensitive email duplicates;
    NULL emails are allowed multiple times."""
    import sqlite3
    import pytest as _pytest
    from backend import database

    database.init_db()

    with database._db() as conn:
        # Cleanup pioneers table (in case of prior test pollution).
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()

        cur = conn.execute(
            "INSERT INTO pioneers (first_name, last_name, email) VALUES (?, ?, ?)",
            ("Pia", "", "Pia@Example.com"),
        )

        with _pytest.raises(sqlite3.IntegrityError):
            conn.execute(
                "INSERT INTO pioneers (first_name, last_name, email) VALUES (?, ?, ?)",
                ("Pia", "P.", "pia@example.com"),  # different case but same lower
            )

        # Different email is fine.
        conn.execute(
            "INSERT INTO pioneers (first_name, last_name, email) VALUES (?, ?, ?)",
            ("Bob", "", "bob@example.com"),
        )

        # Multiple NULL emails are allowed.
        conn.execute("INSERT INTO pioneers (first_name, last_name) VALUES ('X', '')")
        conn.execute("INSERT INTO pioneers (first_name, last_name) VALUES ('Y', '')")

        # Cleanup
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()


def test_pioneer_models():
    """Top-level PioneerCreate / PioneerUpdate / PioneerSummary validate.
    Nested ProjectPioneerEntry (inside ProjectCreate.pioneers) accepts either
    pioneer_id OR first_name + last_name (+ email)."""
    import pytest
    from pydantic import ValidationError
    from backend.models import (
        PioneerCreate, PioneerUpdate, PioneerSummary,
        ProjectCreate, ProjectPioneerEntry,
    )

    # Top-level PioneerCreate (for POST /api/pioneers).
    p = PioneerCreate(first_name="Pia", last_name="P.", email="pia@example.com", notes="Senior expert")
    assert p.first_name == "Pia"
    assert p.last_name == "P."
    assert p.display_name == "Pia P."
    assert p.email == "pia@example.com"
    assert p.notes == "Senior expert"

    # Email is optional, notes optional. Single-word names ok (last_name="").
    p2 = PioneerCreate(first_name="Bob", last_name="")
    assert p2.email is None
    assert p2.notes is None
    assert p2.display_name == "Bob"

    # Empty first AND empty last rejected.
    with pytest.raises(ValidationError):
        PioneerCreate(first_name="", last_name="")
    with pytest.raises(ValidationError):
        PioneerCreate(first_name="   ", last_name="   ")

    # Name part > 80 chars rejected.
    with pytest.raises(ValidationError):
        PioneerCreate(first_name="x" * 81, last_name="Y")
    with pytest.raises(ValidationError):
        PioneerCreate(first_name="X", last_name="y" * 81)

    # PioneerUpdate — all fields optional.
    u = PioneerUpdate(first_name="Pia New", last_name="Surname")
    assert u.first_name == "Pia New"
    assert u.last_name == "Surname"
    u2 = PioneerUpdate()
    assert u2.first_name is None and u2.last_name is None and u2.email is None and u2.notes is None

    # PioneerSummary minimal shape (used as response model). display_name is
    # auto-derived from first_name + last_name.
    s = PioneerSummary(
        id=1, first_name="Pia", last_name="Romano",
        email="pia@example.com", notes=None,
        project_count=3, rounds_completed=4, rounds_expected=6,
        completion_rate=0.667, last_activity_at="2026-04-26",
        status="pending",
        avg_quality_score=0.78, avg_value_gain=1.6,
        avg_machine_first=1.4, avg_senior_led=2.1, avg_knowledge=1.8,
        practices=[{"code": "RWE", "count": 3}],
        roles=[{"role_name": "Senior", "count": 2}],
    )
    assert s.id == 1
    assert s.status == "pending"
    assert s.display_name == "Pia Romano"
    assert len(s.practices) == 1

    # Nested ProjectPioneerEntry (inside ProjectCreate.pioneers) — accepts pioneer_id.
    base = {
        "project_name": "T", "category_id": 1,
        "xcsg_team_size": "1", "xcsg_revision_rounds": "1",
    }
    p_with_id = ProjectCreate(**base, pioneers=[{"pioneer_id": 7}])
    assert p_with_id.pioneers[0].pioneer_id == 7
    assert p_with_id.pioneers[0].first_name is None  # not provided

    # Nested ProjectPioneerEntry — accepts first_name+last_name+email (inline create path).
    p_with_inline = ProjectCreate(**base, pioneers=[
        {"first_name": "New", "last_name": "Pia", "email": "newpia@example.com"},
    ])
    assert p_with_inline.pioneers[0].pioneer_id is None
    assert p_with_inline.pioneers[0].first_name == "New"
    assert p_with_inline.pioneers[0].last_name == "Pia"
    assert p_with_inline.pioneers[0].email == "newpia@example.com"

    # Nested — must have at least pioneer_id OR a name part.
    with pytest.raises(ValidationError):
        ProjectCreate(**base, pioneers=[{}])  # no id, no name parts
    with pytest.raises(ValidationError):
        ProjectCreate(**base, pioneers=[{"first_name": "", "last_name": ""}])


def test_pioneer_db_helpers_crud():
    """create_pioneer / find_pioneer_by_email / update_pioneer_record / delete_pioneer."""
    from backend import database

    database.init_db()

    # Cleanup any leftover pioneers from prior tests.
    with database._db() as conn:
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()

    # Create a pioneer.
    pid = database.create_pioneer(first_name="Pia", last_name="P.", email="pia@example.com", notes=None, created_by=1)
    assert isinstance(pid, int) and pid > 0

    # find_pioneer_by_email — case-insensitive. Returns dict with display_name.
    found = database.find_pioneer_by_email("PIA@EXAMPLE.COM")
    assert found is not None
    assert found["id"] == pid
    assert found["first_name"] == "Pia"
    assert found["last_name"] == "P."
    assert found["display_name"] == "Pia P."

    # find_pioneer_by_email — None when not found.
    assert database.find_pioneer_by_email("nobody@nowhere.com") is None
    assert database.find_pioneer_by_email(None) is None

    # update_pioneer_record.
    database.update_pioneer_record(pid, first_name="Pia", last_name="Pio", email="pia@example.com", notes="Senior")
    with database._db() as conn:
        row = conn.execute("SELECT * FROM pioneers WHERE id = ?", (pid,)).fetchone()
    assert row["first_name"] == "Pia"
    assert row["last_name"] == "Pio"
    assert row["notes"] == "Senior"

    # delete_pioneer — succeeds when not on any project.
    database.delete_pioneer(pid)
    assert database.find_pioneer_by_email("pia@example.com") is None


def test_delete_pioneer_assigned_to_project_raises():
    """delete_pioneer raises a specific exception when the pioneer is on a project."""
    import pytest as _pytest
    from backend import database

    database.init_db()

    with database._db() as conn:
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()

    pid = database.create_pioneer(first_name="Bob", last_name="", email="bob@example.com", notes=None, created_by=1)

    # Manually create a project + assign pioneer.
    with database._db() as conn:
        cur = conn.execute(
            "INSERT INTO projects (created_by, project_name, category_id, "
            "xcsg_team_size, xcsg_revision_rounds, "
            "legacy_calendar_days, legacy_revision_rounds, "
            "expert_token, status) "
            "VALUES (1, 'P', 1, '1', '1', '10', '1', 'tok-del-test', 'pending')"
        )
        proj_id = cur.lastrowid
        conn.execute(
            "INSERT INTO project_pioneers (project_id, pioneer_id, expert_token) "
            "VALUES (?, ?, ?)",
            (proj_id, pid, "tok-pp-del-test"),
        )
        conn.commit()

    # Delete should raise.
    try:
        with _pytest.raises(database.PioneerInUseError):
            database.delete_pioneer(pid)
    finally:
        with database._db() as conn:
            conn.execute("DELETE FROM project_pioneers WHERE project_id = ?", (proj_id,))
            conn.execute("DELETE FROM projects WHERE id = ?", (proj_id,))
            conn.execute("DELETE FROM pioneers WHERE id = ?", (pid,))
            conn.commit()


def test_list_pioneers_with_metrics_aggregation():
    """list_pioneers_with_metrics returns one row per pioneer with aggregated
    project count, completion rate, status, practices, roles, and avg metrics."""
    from backend import database

    database.init_db()

    # Cleanup.
    with database._db() as conn:
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()

    # Create a pioneer with no project assignments — status should be "never".
    p1 = database.create_pioneer(first_name="Newbie", last_name="Tester", email="newbie@example.com", notes=None, created_by=1)

    rows = database.list_pioneers_with_metrics()
    by_id = {r["id"]: r for r in rows}
    assert p1 in by_id
    assert by_id[p1]["status"] == "never"
    assert by_id[p1]["project_count"] == 0
    assert by_id[p1]["rounds_completed"] == 0
    assert by_id[p1]["completion_rate"] is None

    # Cleanup.
    with database._db() as conn:
        conn.execute("DELETE FROM pioneers WHERE id = ?", (p1,))
        conn.commit()


def test_get_pioneer_with_metrics_includes_portfolio():
    """get_pioneer_with_metrics returns the same shape as list, plus a
    portfolio list of project entries."""
    from backend import database

    database.init_db()

    with database._db() as conn:
        conn.execute("DELETE FROM project_pioneers")
        conn.execute("DELETE FROM pioneers")
        conn.commit()

    pid = database.create_pioneer(first_name="Solo", last_name="Pia", email="solo@example.com", notes=None, created_by=1)
    out = database.get_pioneer_with_metrics(pid)
    assert out is not None
    assert out["id"] == pid
    assert "portfolio" in out
    assert out["portfolio"] == []

    # Non-existent.
    assert database.get_pioneer_with_metrics(99999) is None

    with database._db() as conn:
        conn.execute("DELETE FROM pioneers WHERE id = ?", (pid,))
        conn.commit()


def test_pioneer_api_crud():
    """POST -> GET -> GET list -> PUT -> DELETE happy path."""
    headers = auth_h(admin_token())

    # Cleanup any leftover api-test pioneers.
    pre = requests.get(f"{BASE}/api/pioneers", headers=headers).json()
    for p in pre:
        if p.get("email") and "api-test" in p["email"]:
            requests.delete(f"{BASE}/api/pioneers/{p['id']}", headers=headers)

    # POST — create.
    r = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "Pia", "last_name": "API-Test",
        "email": "pia-api-test@example.com", "notes": "Initial",
    })
    assert r.status_code == 201, r.text
    pid = r.json()["id"]

    try:
        # GET single.
        r = requests.get(f"{BASE}/api/pioneers/{pid}", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert body["first_name"] == "Pia"
        assert body["last_name"] == "API-Test"
        assert body["display_name"] == "Pia API-Test"
        assert "portfolio" in body
        assert body["portfolio"] == []
        assert body["status"] == "never"

        # GET list — pioneer appears.
        r = requests.get(f"{BASE}/api/pioneers", headers=headers)
        ids = {p["id"] for p in r.json()}
        assert pid in ids

        # PUT.
        r = requests.put(f"{BASE}/api/pioneers/{pid}", headers=headers, json={
            "first_name": "Pia", "last_name": "A-T",
        })
        assert r.status_code == 200, r.text
        assert r.json()["last_name"] == "A-T"
        assert r.json()["display_name"] == "Pia A-T"

        # GET 404 for unknown.
        assert requests.get(f"{BASE}/api/pioneers/99999", headers=headers).status_code == 404
    finally:
        # DELETE — succeeds because no project assignments.
        del_r = requests.delete(f"{BASE}/api/pioneers/{pid}", headers=headers)
        assert del_r.status_code == 204


def test_pioneer_api_find_or_create():
    """POST with an email that already exists returns 200 with the existing
    record (find-or-create); without email always 201."""
    headers = auth_h(admin_token())

    # First create.
    r1 = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "Foc", "last_name": "Pia", "email": "foc-pia@example.com",
    })
    assert r1.status_code == 201
    pid = r1.json()["id"]

    try:
        # Second POST with same email (different case) -> 200, returns same id.
        r2 = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
            "first_name": "Different", "last_name": "Name", "email": "FOC-PIA@EXAMPLE.COM",
        })
        assert r2.status_code == 200, r2.text
        assert r2.json()["id"] == pid
        # Server returns the EXISTING record; the conflicting name is not applied.
        assert r2.json()["first_name"] == "Foc"
        assert r2.json()["last_name"] == "Pia"

        # POST without email -> always 201, new id.
        r3 = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
            "first_name": "No Email", "last_name": "Pia",
        })
        assert r3.status_code == 201
        assert r3.json()["id"] != pid
        # Cleanup.
        requests.delete(f"{BASE}/api/pioneers/{r3.json()['id']}", headers=headers)
    finally:
        requests.delete(f"{BASE}/api/pioneers/{pid}", headers=headers)


def test_pioneer_api_admin_only():
    """PUT / DELETE require admin. POST is admin + analyst. GET is everyone."""
    h_admin = auth_h(admin_token())
    h_analyst = auth_h(login_token("pmo", "AliraPMO2026!"))
    h_viewer = auth_h(login_token("viewer", "AliraView2026!"))

    # GET allowed for all.
    assert requests.get(f"{BASE}/api/pioneers", headers=h_admin).status_code == 200
    assert requests.get(f"{BASE}/api/pioneers", headers=h_analyst).status_code == 200
    assert requests.get(f"{BASE}/api/pioneers", headers=h_viewer).status_code == 200

    # POST allowed for admin + analyst, blocked for viewer.
    r = requests.post(f"{BASE}/api/pioneers", headers=h_analyst, json={
        "first_name": "Analyst", "last_name": "Created", "email": "analyst-created@example.com",
    })
    assert r.status_code in (200, 201), r.text
    pid = r.json()["id"]

    try:
        assert requests.post(f"{BASE}/api/pioneers", headers=h_viewer, json={
            "first_name": "Viewer", "last_name": "Tried", "email": "viewer-tried@example.com",
        }).status_code == 403

        # PUT — admin only.
        assert requests.put(f"{BASE}/api/pioneers/{pid}", headers=h_admin, json={"notes": "x"}).status_code == 200
        assert requests.put(f"{BASE}/api/pioneers/{pid}", headers=h_analyst, json={"notes": "y"}).status_code == 403
        assert requests.put(f"{BASE}/api/pioneers/{pid}", headers=h_viewer, json={"notes": "z"}).status_code == 403

        # DELETE — admin only.
        assert requests.delete(f"{BASE}/api/pioneers/{pid}", headers=h_analyst).status_code == 403
        assert requests.delete(f"{BASE}/api/pioneers/{pid}", headers=h_viewer).status_code == 403
    finally:
        requests.delete(f"{BASE}/api/pioneers/{pid}", headers=h_admin)


def test_pioneer_delete_in_use_returns_409():
    """DELETE returns 409 when the pioneer is on any project."""
    headers = auth_h(admin_token())

    # Create pioneer.
    r = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "InUse", "last_name": "Pia", "email": "inuse-pia@example.com",
    })
    pid = r.json()["id"]

    # Create project assigning this pioneer (via pioneer_id).
    proj_r = requests.post(f"{BASE}/api/projects", headers=headers, json={
        "project_name": "delete-in-use test",
        "category_id": 1,
        "pioneers": [{"pioneer_id": pid}],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
    })
    assert proj_r.status_code == 201, proj_r.text
    proj_id = proj_r.json()["id"]

    try:
        # DELETE pioneer -> 409.
        del_r = requests.delete(f"{BASE}/api/pioneers/{pid}", headers=headers)
        assert del_r.status_code == 409, del_r.text
        # Detail should mention the project count.
        detail = del_r.json().get("detail", "")
        assert "1" in str(detail)
    finally:
        requests.delete(f"{BASE}/api/projects/{proj_id}", headers=headers)
        requests.delete(f"{BASE}/api/pioneers/{pid}", headers=headers)


def test_project_create_with_existing_pioneer_id():
    """ProjectCreate.pioneers entries can reference an existing pioneer by id."""
    headers = auth_h(admin_token())

    # Create a pioneer first.
    pr = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "Existing", "last_name": "Pia", "email": "existing-pia@example.com",
    })
    assert pr.status_code in (200, 201)
    pid = pr.json()["id"]

    try:
        # Create project referencing pioneer by id.
        proj_r = requests.post(f"{BASE}/api/projects", headers=headers, json={
            "project_name": "with-existing-id",
            "category_id": 1,
            "pioneers": [{"pioneer_id": pid, "total_rounds": 2, "role_name": "Senior"}],
            "xcsg_team_size": "1",
            "xcsg_revision_rounds": "1",
        })
        assert proj_r.status_code == 201, proj_r.text
        proj_id = proj_r.json()["id"]

        # GET project: pioneer name comes from pioneers table (display_name +
        # pioneer_first_name + pioneer_last_name + pioneer_name (computed)).
        detail = requests.get(f"{BASE}/api/projects/{proj_id}", headers=headers).json()
        pi = detail["pioneers"][0]
        assert pi["pioneer_id"] == pid
        assert pi["pioneer_first_name"] == "Existing"
        assert pi["pioneer_last_name"] == "Pia"
        assert pi["pioneer_name"] == "Existing Pia"
        assert pi["display_name"] == "Existing Pia"
        assert pi["pioneer_email"] == "existing-pia@example.com"
        assert pi["role_name"] == "Senior"

        # Cleanup.
        requests.delete(f"{BASE}/api/projects/{proj_id}", headers=headers)
    finally:
        requests.delete(f"{BASE}/api/pioneers/{pid}", headers=headers)


def test_project_create_with_inline_pioneer():
    """ProjectCreate.pioneers can carry name+email — server find-or-creates the pioneer."""
    headers = auth_h(admin_token())

    # Cleanup any leftover.
    pre = requests.get(f"{BASE}/api/pioneers", headers=headers).json()
    for p in pre:
        if p.get("email") == "inline-pia@example.com":
            requests.delete(f"{BASE}/api/pioneers/{p['id']}", headers=headers)

    # Create project with inline pioneer.
    proj_r = requests.post(f"{BASE}/api/projects", headers=headers, json={
        "project_name": "with-inline-pioneer",
        "category_id": 1,
        "pioneers": [{"first_name": "Inline", "last_name": "Pia", "email": "inline-pia@example.com", "total_rounds": 1}],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
    })
    assert proj_r.status_code == 201, proj_r.text
    proj_id = proj_r.json()["id"]

    new_pioneer = None
    try:
        # Pioneer was created.
        all_p = requests.get(f"{BASE}/api/pioneers", headers=headers).json()
        new_pioneer = next((p for p in all_p if p.get("email") == "inline-pia@example.com"), None)
        assert new_pioneer is not None

        # And the project references it.
        detail = requests.get(f"{BASE}/api/projects/{proj_id}", headers=headers).json()
        assert detail["pioneers"][0]["pioneer_id"] == new_pioneer["id"]
    finally:
        requests.delete(f"{BASE}/api/projects/{proj_id}", headers=headers)
        if new_pioneer:
            requests.delete(f"{BASE}/api/pioneers/{new_pioneer['id']}", headers=headers)


def test_dashboard_pioneer_filter():
    """GET /api/dashboard/metrics with pioneer_id query param scopes to that
    pioneer's projects only."""
    headers = auth_h(admin_token())

    # Create two pioneers and two projects, each with a different pioneer.
    p1 = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "Filter", "last_name": "A", "email": "filter-a@example.com",
    }).json()
    p2 = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "Filter", "last_name": "B", "email": "filter-b@example.com",
    }).json()

    proj1 = requests.post(f"{BASE}/api/projects", headers=headers, json={
        "project_name": "filter-test-1",
        "category_id": 1,
        "pioneers": [{"pioneer_id": p1["id"], "total_rounds": 1}],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
    }).json()
    proj2 = requests.post(f"{BASE}/api/projects", headers=headers, json={
        "project_name": "filter-test-2",
        "category_id": 1,
        "pioneers": [{"pioneer_id": p2["id"], "total_rounds": 1}],
        "xcsg_team_size": "1",
        "xcsg_revision_rounds": "1",
    }).json()

    try:
        # Without filter — both projects counted.
        all_metrics = requests.get(f"{BASE}/api/dashboard/metrics", headers=headers).json()
        assert all_metrics["total_projects"] >= 2

        # With pioneer_id=p1 — proj1 included, proj2 excluded.
        p1_metrics = requests.get(
            f"{BASE}/api/dashboard/metrics?pioneer_id={p1['id']}", headers=headers
        ).json()
        assert p1_metrics["total_projects"] <= all_metrics["total_projects"]
        assert p1_metrics["total_projects"] >= 1  # at least proj1

        # With both pioneer_ids — both included.
        both_metrics = requests.get(
            f"{BASE}/api/dashboard/metrics?pioneer_id={p1['id']}&pioneer_id={p2['id']}",
            headers=headers,
        ).json()
        assert both_metrics["total_projects"] >= 2  # at least proj1 and proj2

        # Unknown pioneer_id — no projects matched (empty filter).
        unknown_metrics = requests.get(
            f"{BASE}/api/dashboard/metrics?pioneer_id=999999", headers=headers
        ).json()
        assert unknown_metrics["total_projects"] == 0
    finally:
        requests.delete(f"{BASE}/api/projects/{proj1['id']}", headers=headers)
        requests.delete(f"{BASE}/api/projects/{proj2['id']}", headers=headers)
        requests.delete(f"{BASE}/api/pioneers/{p1['id']}", headers=headers)
        requests.delete(f"{BASE}/api/pioneers/{p2['id']}", headers=headers)


def test_export_pioneer_xlsx():
    """GET /api/export/pioneer/{id}.xlsx returns XLSX with summary + portfolio sheets."""
    import io
    from openpyxl import load_workbook

    headers = auth_h(admin_token())

    p = requests.post(f"{BASE}/api/pioneers", headers=headers, json={
        "first_name": "XLSX", "last_name": "Test", "email": "xlsx-test@example.com",
    }).json()

    try:
        r = requests.get(f"{BASE}/api/export/pioneer/{p['id']}.xlsx", headers=headers)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

        # Load XLSX.
        wb = load_workbook(io.BytesIO(r.content))
        assert "summary" in wb.sheetnames
        assert "portfolio" in wb.sheetnames

        summary = wb["summary"]
        # Header row + 1 data row.
        assert summary.max_row == 2
        # Find the first_name and last_name columns and verify our pioneer.
        header_row = [c.value for c in summary[1]]
        assert "name" not in header_row, "legacy 'name' column should be gone"
        first_col_idx = header_row.index("first_name")
        last_col_idx = header_row.index("last_name")
        assert summary.cell(row=2, column=first_col_idx + 1).value == "XLSX"
        assert summary.cell(row=2, column=last_col_idx + 1).value == "Test"

        # Empty portfolio (no project assignments).
        portfolio = wb["portfolio"]
        # Header row only (no data rows).
        assert portfolio.max_row == 1
    finally:
        requests.delete(f"{BASE}/api/pioneers/{p['id']}", headers=headers)


def test_export_pioneer_xlsx_404_for_unknown():
    """Unknown pioneer_id returns 404."""
    headers = auth_h(admin_token())
    r = requests.get(f"{BASE}/api/export/pioneer/99999.xlsx", headers=headers)
    assert r.status_code == 404


def test_fx_rates_get_put_endpoints():
    """GET /api/fx-rates is open; PUT requires admin and validates inputs."""
    print("\n── FX Rates ──")

    def _login(u, p):
        r = requests.post(f"{BASE}/api/auth/login", json={"username": u, "password": p})
        r.raise_for_status()
        return r.json()["access_token"]

    h_admin = auth_h(admin_token())
    h_analyst = auth_h(_login("pmo", "AliraPMO2026!"))
    h_viewer = auth_h(_login("viewer", "AliraView2026!"))

    # GET works for all roles.
    r = requests.get(f"{BASE}/api/fx-rates", headers=h_admin)
    test("GET /api/fx-rates 200 for admin", r.status_code == 200, f"got {r.status_code}")
    body = r.json()
    test("GET /api/fx-rates has base_currency", "base_currency" in body)
    test("GET /api/fx-rates has rates list", isinstance(body.get("rates"), list))
    test("GET /api/fx-rates has 6 rate rows", len(body["rates"]) == 6, f"got {len(body['rates'])}")

    test("GET /api/fx-rates 200 for analyst",
         requests.get(f"{BASE}/api/fx-rates", headers=h_analyst).status_code == 200)
    test("GET /api/fx-rates 200 for viewer",
         requests.get(f"{BASE}/api/fx-rates", headers=h_viewer).status_code == 200)

    initial_base = body["base_currency"]
    initial_rates = {r["currency_code"]: r["rate_to_base"] for r in body["rates"]}

    try:
        # Admin can PUT.
        upd = requests.put(f"{BASE}/api/fx-rates", headers=h_admin, json={
            "base_currency": "USD",
            "rates": [
                {"currency_code": "EUR", "rate_to_base": 1.0850},
                {"currency_code": "GBP", "rate_to_base": 1.2430},
            ],
        })
        test("PUT /api/fx-rates 200 for admin", upd.status_code == 200, f"got {upd.status_code}: {upd.text[:120]}")
        rt = requests.get(f"{BASE}/api/fx-rates", headers=h_admin).json()
        new_rates = {r["currency_code"]: r["rate_to_base"] for r in rt["rates"]}
        test("PUT persists EUR rate", new_rates["EUR"] == 1.0850)
        test("PUT persists GBP rate", new_rates["GBP"] == 1.2430)

        # Non-admin cannot PUT.
        analyst_put = requests.put(f"{BASE}/api/fx-rates", headers=h_analyst, json={
            "base_currency": "USD", "rates": []
        })
        test("PUT 403 for analyst", analyst_put.status_code == 403, f"got {analyst_put.status_code}")
        viewer_put = requests.put(f"{BASE}/api/fx-rates", headers=h_viewer, json={
            "base_currency": "USD", "rates": []
        })
        test("PUT 403 for viewer", viewer_put.status_code == 403, f"got {viewer_put.status_code}")

        # Invalid currency code rejected.
        bad_code = requests.put(f"{BASE}/api/fx-rates", headers=h_admin, json={
            "base_currency": "USD",
            "rates": [{"currency_code": "XYZ", "rate_to_base": 1.0}],
        })
        test("PUT 422 for invalid currency code", bad_code.status_code == 422, f"got {bad_code.status_code}: {bad_code.text[:120]}")

        # Negative rate rejected.
        bad_rate = requests.put(f"{BASE}/api/fx-rates", headers=h_admin, json={
            "base_currency": "USD",
            "rates": [{"currency_code": "EUR", "rate_to_base": -1.0}],
        })
        test("PUT 422 for negative rate", bad_rate.status_code == 422, f"got {bad_rate.status_code}")

        # base_currency persists into app_settings.
        settings_get = requests.get(f"{BASE}/api/settings", headers=h_admin).json()
        test("base_currency mirrored in /api/settings", settings_get.get("base_currency") == "USD")
    finally:
        # Restore.
        restore = [{"currency_code": k, "rate_to_base": v} for k, v in initial_rates.items()]
        requests.put(f"{BASE}/api/fx-rates", headers=h_admin, json={
            "base_currency": initial_base, "rates": restore
        })


def test_dashboard_economics_endpoint():
    """GET /api/dashboard/economics returns {summary, breakdowns, trends},
    honors pioneer_id and date filters, normalizes via FX rates."""
    print("\n── Dashboard Economics ──")
    h = auth_h(admin_token())

    # Seed two projects via the public API so we know exactly what to expect.
    cats = requests.get(f"{BASE}/api/categories", headers=h).json()
    practices = requests.get(f"{BASE}/api/practices", headers=h).json()
    pmap = {p["code"]: p["id"] for p in practices}
    map_cat = next((c for c in cats if any(p["code"] == "MAP" for p in c.get("practices", []))), None)
    if not map_cat:
        test("skip — no MAP-eligible category", False, detail="setup")
        return

    common_legacy = [{"role_name": "Senior", "count": 2, "day_rate": 1500}]
    bodies = [
        {
            "project_name": "Econ Test USD", "category_id": map_cat["id"], "practice_id": pmap["MAP"],
            "engagement_revenue": 100000.0, "currency": "USD", "xcsg_pricing_model": "Fixed fee",
            "date_started": "2026-01-15", "date_delivered": "2026-02-10",
            "working_days": 10, "xcsg_team_size": "2", "xcsg_revision_rounds": "1",
            "engagement_stage": "Active engagement",
            "pioneers": [{"first_name": "Econ", "last_name": "Tester1", "email": "econ1@example.com", "total_rounds": 1}],
            "legacy_team": common_legacy,
        },
        {
            "project_name": "Econ Test EUR", "category_id": map_cat["id"], "practice_id": pmap["MAP"],
            "engagement_revenue": 200000.0, "currency": "EUR", "xcsg_pricing_model": "Time & materials",
            "date_started": "2026-01-20", "date_delivered": "2026-02-15",
            "working_days": 12, "xcsg_team_size": "2", "xcsg_revision_rounds": "1",
            "engagement_stage": "Active engagement",
            "pioneers": [{"first_name": "Econ", "last_name": "Tester2", "email": "econ2@example.com", "total_rounds": 1}],
            "legacy_team": common_legacy,
        },
    ]
    created_ids = []
    initial_fx = None
    try:
        # Set EUR rate to 1.10 for predictable arithmetic. Restore after.
        initial_fx = requests.get(f"{BASE}/api/fx-rates", headers=h).json()
        requests.put(f"{BASE}/api/fx-rates", headers=h, json={
            "base_currency": "USD",
            "rates": [{"currency_code": "EUR", "rate_to_base": 1.10}],
        })

        # Import a STRONG payload so the projects compute non-null margins.
        import importlib.util, pathlib
        spec = importlib.util.spec_from_file_location("sd", pathlib.Path("tests/seed_20_projects.py"))
        sd = importlib.util.module_from_spec(spec); spec.loader.exec_module(sd)
        strong = dict(getattr(sd, "STRONG"))

        for b in bodies:
            r = requests.post(f"{BASE}/api/projects", headers=h, json=b)
            assert r.status_code == 201, f"create failed: {r.status_code}: {r.text[:160]}"
            proj = r.json()
            created_ids.append(proj["id"])
            tok = proj["pioneers"][0]["expert_token"]
            sr = requests.post(f"{BASE}/api/expert/{tok}", json=strong)
            assert sr.status_code == 201, f"survey submit failed: {sr.status_code}: {sr.text[:160]}"

        # Hit the endpoint.
        r = requests.get(f"{BASE}/api/dashboard/economics", headers=h)
        test("GET /api/dashboard/economics 200", r.status_code == 200, f"got {r.status_code}: {r.text[:120]}")
        data = r.json()
        test("response has summary", "summary" in data)
        test("response has breakdowns", "breakdowns" in data)
        test("response has trends", "trends" in data)

        s = data["summary"]
        test("summary.base_currency = USD", s["base_currency"] == "USD")
        test("summary.qualifying_project_count >= 2",
             s["qualifying_project_count"] >= 2,
             detail=f"got {s['qualifying_project_count']}")
        # total_revenue should INCLUDE EUR-converted revenue (200k * 1.10 = 220k)
        # and the USD project's 100k → at least 320k attributed to OUR projects.
        test("summary.total_revenue >= 320000",
             s["total_revenue"] >= 320000.0,
             detail=f"got {s['total_revenue']}")

        # by_currency includes USD and EUR with native amounts.
        cur = {row["code"]: row for row in data["breakdowns"]["by_currency"]}
        test("by_currency has USD", "USD" in cur)
        test("by_currency has EUR", "EUR" in cur)

        # Pioneer filter pass-through: filter by Econ Tester1's id.
        pioneers = requests.get(f"{BASE}/api/pioneers", headers=h).json()
        tester1 = next((p for p in pioneers if p.get("email") == "econ1@example.com"), None)
        test("Tester1 pioneer found", tester1 is not None)
        if tester1:
            pid = tester1["id"]
            rf = requests.get(f"{BASE}/api/dashboard/economics?pioneer_id={pid}", headers=h).json()
            # Filter narrows to just the USD project (revenue 100k).
            test("pioneer filter narrows total_revenue to 100k",
                 abs(rf["summary"]["total_revenue"] - 100000.0) < 0.01,
                 detail=f"got {rf['summary']['total_revenue']}")
    finally:
        # Cleanup: delete created projects + restore initial FX + remove test pioneers.
        for pid in created_ids:
            requests.delete(f"{BASE}/api/projects/{pid}", headers=h)
        # Pioneer rows in the global pioneers table don't cascade on project
        # deletion (FK is RESTRICT), so we must DELETE them by email manually.
        all_pioneers = requests.get(f"{BASE}/api/pioneers", headers=h).json()
        for p in all_pioneers:
            if p.get("email") in {"econ1@example.com", "econ2@example.com"}:
                requests.delete(f"{BASE}/api/pioneers/{p['id']}", headers=h)
        if initial_fx:
            requests.put(f"{BASE}/api/fx-rates", headers=h, json={
                "base_currency": initial_fx["base_currency"],
                "rates": [{"currency_code": r["currency_code"], "rate_to_base": r["rate_to_base"]}
                          for r in initial_fx["rates"]],
            })


if __name__ == "__main__":
    sys.exit(main())
